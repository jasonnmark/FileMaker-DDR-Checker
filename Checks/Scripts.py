from lxml import etree as ET
import re
from collections import defaultdict
import ahocorasick

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Script Usage"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 2  # Make this the second sheet

def get_column_widths():
    """Return column widths for this sheet"""
    widths = {
        "Script Name": 250,
        "Total Usage": 80,
        "Status": 85,
        "XML Count": 65,
        "Called from Scripts": 250,
        "Called from Buttons": 250,
        "Called from Triggers": 250,
        "Called from Menus": 200,
        "Called from Other": 200,
        "Debug Info": 85
    }
    
    # Add debug column width if in debug mode
    import sys
    if '-debug' in sys.argv or '--debug' in sys.argv:
        widths["XML Occurrences (Debug)"] = 400
    
    return widths

def apply_styling(ws):
    """Apply custom styling to the Script Usage worksheet"""
    # Get the color helper functions that were injected
    from openpyxl.styles import PatternFill, Font
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Get status value
        status_value = ""
        if "Status" in columns:
            status_cell = row[columns["Status"] - 1]
            status_value = str(status_cell.value or "")
        
        # Apply styling based on status
        if "NOT USED" in status_value:
            # Highlight entire row for unused scripts
            for cell in row:
                cell.fill = get_color_fill('error')
            
            # Make status cell bold with error font
            if "Status" in columns:
                row[columns["Status"] - 1].font = get_color_font('error', bold=True)
            
            # Also make script name bold
            if "Script Name" in columns:
                row[columns["Script Name"] - 1].font = get_color_font('error', bold=True)
        
        elif status_value in ["Server", "Debug", "Dev Playground", "Check Manually"]:
            # Highlight special scripts in warning color
            for cell in row:
                cell.fill = get_color_fill('warning')
            
            # Make status cell bold with warning font
            if "Status" in columns:
                row[columns["Status"] - 1].font = get_color_font('warning', bold=True)
            
            # Also make script name bold
            if "Script Name" in columns:
                row[columns["Script Name"] - 1].font = get_color_font('warning', bold=True)
        
        elif status_value == "Scheduled For Deletion":
            # Highlight scheduled for deletion scripts in success color
            for cell in row:
                cell.fill = get_color_fill('success')
            
            # Make status cell bold with success font
            if "Status" in columns:
                row[columns["Status"] - 1].font = get_color_font('success', bold=True)
            
            # Also make script name bold
            if "Script Name" in columns:
                row[columns["Script Name"] - 1].font = get_color_font('success', bold=True)
        
        # Color code XML Count - green if > 2
        if "XML Count" in columns:
            xml_count_cell = row[columns["XML Count"] - 1]
            try:
                count = int(xml_count_cell.value or 0)
                if count > 2:
                    xml_count_cell.font = get_color_font('success', bold=True)
                    xml_count_cell.fill = PatternFill()  # Remove any background fill
            except:
                pass
        
        # Color code Total Usage
        if "Total Usage" in columns:
            count_cell = row[columns["Total Usage"] - 1]
            try:
                count = int(count_cell.value or 0)
                if count == 0:
                    # Only color red if not a special script
                    if status_value not in ["Server", "Debug", "Dev Playground"]:
                        count_cell.font = get_color_font('error', bold=True)
                elif count >= 10:
                    count_cell.font = get_color_font('success', bold=True)
                elif count >= 5:
                    count_cell.font = Font(bold=True, color="388E3C")  # Medium green
            except:
                pass

# Keep the existing run_check function and other logic unchanged
def run_check(raw_xml):
    """
    Find all scripts and track where they are called from.
    Flag scripts that aren't called anywhere.
    Sort so unused scripts appear at the top.
    """
    # Check for debug mode from environment or sys.argv
    import sys
    debug_mode = '-debug' in sys.argv or '--debug' in sys.argv
    
    try:
        # Parse the XML
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        root = ET.fromstring(raw_xml.encode('utf-8'), parser)
        
        # Dictionary to store all scripts
        scripts = {}
        scripts_by_id = {}  # Additional lookup by ID
        
        # Dictionary to store script usage by type
        script_usage = defaultdict(lambda: {
            'from_scripts': [],
            'from_buttons': [],
            'from_triggers': [],
            'from_menus': [],
            'from_value_lists': [],
            'from_web_viewers': [],
            'from_other': [],
            'debug_unknown': []  # For tracking unknown references
        })
        
        # First, find all groups and their scripts
        script_folders = {}
        
        # Find all groups
        for group in root.findall(".//Group"):
            group_name = group.attrib.get("name", "")
            group_id = group.attrib.get("id", "")
            
            # Build the full folder path by traversing up
            folder_path = [group_name]
            current = group
            for _ in range(10):
                parent = current.getparent()
                if parent is None or parent.tag != "Group":
                    break
                folder_path.insert(0, parent.attrib.get("name", ""))
                current = parent
            
            full_path = "/".join(folder_path)
            
            # Find all scripts within this group
            for script_node in group.findall(".//Script"):
                script_id = script_node.attrib.get("id")
                if script_id:
                    script_folders[script_id] = full_path
        
        # Now find all scripts
        script_definitions = set()  # Track script definition elements
        for script_node in root.findall(".//Script"):
            script_name = script_node.attrib.get("name", "Unknown Script")
            script_id = script_node.attrib.get("id", "")
            
            # Scripts are never groups - they're Script elements, not Group elements
            is_group = False
            
            # Check if this is a script definition (has steps or is in Scripts/ScriptCatalog section)
            is_definition = False
            
            # Check if this is just a permission entry for a group (has Access but no StepList)
            has_access_only = script_node.find(".//Access") is not None and script_node.find(".//StepList") is None
            
            if has_access_only:
                # This is likely a permission entry for a group, not a real script
                is_definition = False
            elif script_node.find(".//Step") is not None:
                is_definition = True
            else:
                # Check if it's in the Scripts or ScriptCatalog section
                parent = script_node.getparent()
                for _ in range(10):  # Increased range to catch deeper nesting
                    if parent is None:
                        break
                    if parent.tag in ["Scripts", "ScriptCatalog", "ScriptList"]:
                        # But not if it's just a permission entry
                        if not has_access_only:
                            is_definition = True
                        break
                    parent = parent.getparent()
            
            if is_definition:
                script_definitions.add(id(script_node))  # Store the element ID
            
            if not is_group and script_name and is_definition:  # Only add script definitions
                # Store original name and normalized versions
                scripts[script_name] = {
                    "id": script_id,
                    "name": script_name,
                    "folder": script_folders.get(script_id, "")
                }
                
                # Also store by ID for ID-based lookups
                if script_id:
                    scripts_by_id[script_id] = script_name
        
        print(f"Found {len(scripts)} scripts")
        
        # XML occurrence counting with fast search
        print("  Counting script occurrences...")
        script_name_occurrences = {}
        
        # Use Aho-Corasick algorithm for fast multi-pattern search
        print("    Building search patterns...")
        A = ahocorasick.Automaton()
        
        # Add all script names as patterns
        pattern_id = 0
        pattern_to_script = {}  # Map pattern_id to script_name
        
        for script_name in scripts:
            # Add the script name as a pattern
            A.add_word(script_name, (pattern_id, script_name))
            pattern_to_script[pattern_id] = script_name
            pattern_id += 1
        
        print(f"    Added {pattern_id} script patterns")
        
        # Make automaton ready for searching
        A.make_automaton()
        
        # Initialize counts
        for script_name in scripts:
            script_name_occurrences[script_name] = {
                'count': 0,
                'occurrences': []  # For debug mode
            }
        
        # Search all patterns in one pass
        print(f"    Searching XML ({len(raw_xml):,} characters)...")
        matches_found = 0
        
        # For debug mode, we need to track line information
        if debug_mode:
            # Split XML into lines for context
            lines = raw_xml.split('\n')
            line_starts = [0]  # Character positions where each line starts
            pos = 0
            for line in lines:
                pos += len(line) + 1  # +1 for newline
                line_starts.append(pos)
        
        # Search for all script names
        for end_index, (pid, script_name) in A.iter(raw_xml):
            script_name_occurrences[script_name]['count'] += 1
            matches_found += 1
            
            # For debug mode, find which line this match is on
            if debug_mode and len(script_name_occurrences[script_name]['occurrences']) < 50:  # Limit debug info
                # Find the line number
                line_num = 0
                for i, start_pos in enumerate(line_starts):
                    if end_index < start_pos:
                        line_num = i - 1
                        break
                
                if 0 <= line_num < len(lines):
                    line = lines[line_num]
                    # Highlight the script name in the line
                    highlighted_line = line.replace(script_name, f'**{script_name}**')
                    
                    # Trim very long lines for readability
                    if len(highlighted_line) > 200:
                        # Find the script name position and show context around it
                        pos = highlighted_line.find(f'**{script_name}**')
                        if pos >= 0:
                            start = max(0, pos - 80)
                            end = min(len(highlighted_line), pos + len(f'**{script_name}**') + 80)
                            highlighted_line = '...' + highlighted_line[start:end] + '...'
                    
                    script_name_occurrences[script_name]['occurrences'].append(highlighted_line.strip())
            
            if matches_found % 1000 == 0:
                print(f"      Found {matches_found:,} matches so far...")
        
        # Ensure minimum of 2 for all scripts (script definition + self-reference)
        for script_name in scripts:
            if script_name_occurrences[script_name]['count'] < 2:
                script_name_occurrences[script_name]['count'] = 2
                if debug_mode:
                    if len(script_name_occurrences[script_name]['occurrences']) == 0:
                        script_name_occurrences[script_name]['occurrences'].append('(Script definition assumed)')
                        script_name_occurrences[script_name]['occurrences'].append('(Self-reference assumed)')
                    elif len(script_name_occurrences[script_name]['occurrences']) == 1:
                        script_name_occurrences[script_name]['occurrences'].append('(Self-reference assumed)')
        
        print(f"    Search complete! Found {matches_found:,} total matches")
        
        print("  Processing script references...")
        
        # Find ALL Script elements anywhere in the XML
        all_script_elements = root.findall(".//Script")
        print(f"  Found {len(all_script_elements)} total Script elements")
        
        refs_processed = 0
        for script_elem in all_script_elements:
            # Skip if this is a script definition
            if id(script_elem) in script_definitions:
                continue
            
            # This is a script reference
            referenced_script_name = script_elem.attrib.get("name")
            referenced_script_id = script_elem.attrib.get("id")
            
            # Find the actual script name
            actual_script_name = None
            if referenced_script_id and referenced_script_id in scripts_by_id:
                actual_script_name = scripts_by_id[referenced_script_id]
            elif referenced_script_name in scripts:
                actual_script_name = referenced_script_name
            
            if actual_script_name:
                process_script_reference(script_elem, actual_script_name, script_usage, scripts)
                refs_processed += 1
        
        print(f"  Processed {refs_processed} script references")
        
        # Also look for scripts called from JavaScript in Web Viewers
        # These appear as parameters in Perform JavaScript in Web Viewer steps
        js_steps = []
        for step in root.findall(".//Step"):
            if step.attrib.get("id") == "175" or step.attrib.get("name") == "Perform JavaScript in Web Viewer":
                js_steps.append(step)
        
        print(f"  Found {len(js_steps)} JavaScript in Web Viewer steps")
        
        for step in js_steps:
            # Look for script names in parameters with the comment /* fmScript: (Script Name) */
            step_text = step.find(".//StepText")
            if step_text is not None and step_text.text:
                # Look for pattern: "ScriptName" /* fmScript
                import re
                matches = re.findall(r'"([^"]+)"\s*/\*\s*fmScript', step_text.text)
                for script_name in matches:
                    if script_name in scripts:
                        # Find the context (layout/button)
                        context = find_javascript_context(step)
                        script_usage[script_name]['from_other'].append(f"JavaScript in Web Viewer - {context}")
                        if debug_mode:
                            print(f"[DEBUG] Found JavaScript call to '{script_name}' in {context}")
        
        # Build results
        results = []
        
        # Process all scripts and build complete results
        all_script_results = []
        
        for script_name in scripts.keys():
            usage = script_usage.get(script_name, {
                'from_scripts': [],
                'from_buttons': [],
                'from_triggers': [],
                'from_menus': [],
                'from_value_lists': [],
                'from_web_viewers': [],
                'from_other': [],
                'debug_unknown': []
            })
            
            total_usage = sum(len(usage[key]) for key in usage if key != 'debug_unknown')
            xml_count = script_name_occurrences.get(script_name, {}).get('count', 2)
            xml_occurrences = script_name_occurrences.get(script_name, {}).get('occurrences', [])
            
            # Determine special status
            special_status = None
            script_name_lower = script_name.lower()
            script_info = scripts.get(script_name, {})
            folder_path_lower = script_info.get("folder", "").lower()
            
            # Check for ToDelete folder
            if "todelete" in folder_path_lower or "/todelete/" in folder_path_lower:
                special_status = "Scheduled For Deletion"
            # Check for server scripts (more comprehensive check)
            elif ("server" in script_name_lower or 
                  "onserver" in script_name_lower or
                  "on_server" in script_name_lower or
                  "runonserver" in script_name_lower.replace(" ", "").replace("_", "")):
                special_status = "Server"
                if debug_mode and "server" in script_name_lower:
                    print(f"[DEBUG] '{script_name}' detected as Server script")
            # Check for debug scripts
            elif "debug" in script_name_lower:
                special_status = "Debug"
            # Check for dev playground scripts (in name or folder)
            elif "dev playground" in script_name_lower or "dev playground" in folder_path_lower:
                special_status = "Dev Playground"
            
            # Format usage lists
            def format_usage_list(usage_list, max_items=3):
                if not usage_list:
                    return ""
                
                unique_items = list(dict.fromkeys(usage_list))  # Remove duplicates while preserving order
                
                if len(unique_items) <= max_items:
                    return "\n".join(unique_items)
                else:
                    shown_items = unique_items[:max_items]
                    remaining = len(unique_items) - max_items
                    shown_items.append(f"...and {remaining} more")
                    return "\n".join(shown_items)
            
            # Combine other sources
            other_sources = usage['from_value_lists'] + usage['from_web_viewers'] + usage['from_other']
            
            # Combine debug info
            debug_text = ""
            if usage['debug_unknown']:
                debug_text = "\n".join(usage['debug_unknown'][:3])
                if len(usage['debug_unknown']) > 3:
                    debug_text += f"\n...and {len(usage['debug_unknown']) - 3} more unknown refs"
            
            # Determine final status with new logic
            if special_status == "Scheduled For Deletion":
                status = "Scheduled For Deletion"
            elif total_usage == 0:
                # Even if it's a special script with no usage, mark it appropriately
                if special_status in ["Server", "Debug", "Dev Playground"]:
                    status = special_status
                elif xml_count <= 2:  # 0, 1, or 2 means just the definition
                    status = "⚠️ NOT USED"
                else:
                    # More than 2 occurrences but no recognized usage
                    status = "Check Manually"
            else:
                # Has usage
                if special_status in ["Server", "Debug", "Dev Playground"]:
                    # Don't add ⚠️ prefix - the Excel code looks for exact match
                    status = special_status
                else:
                    status = "Active"
            
            # Build result dictionary
            result = {
                "Script Name": script_name,
                "Total Usage": total_usage,
                "Status": status,
                "XML Count": xml_count,
            }
            
            # Add debug column only in debug mode
            if debug_mode:
                result["XML Occurrences (Debug)"] = "\n".join(xml_occurrences[:10])  # Limit to 10 for readability
            
            # Add the rest of the columns
            result.update({
                "Called from Scripts": format_usage_list(usage['from_scripts']),
                "Called from Buttons": format_usage_list(usage['from_buttons']),
                "Called from Triggers": format_usage_list(usage['from_triggers']),
                "Called from Menus": format_usage_list(usage['from_menus']),
                "Called from Other": format_usage_list(other_sources),
                "Debug Info": debug_text,
                "_sort_key": (
                    0 if status == "⚠️ NOT USED" else
                    1 if status == "Check Manually" else
                    2 if status == "Server" else
                    3 if status == "Debug" else
                    4 if status == "Dev Playground" else
                    5 if status == "Scheduled For Deletion" else
                    6,  # Active
                    -total_usage,  # Within each category, sort by usage (descending)
                    script_name.lower()  # Then by name
                )
            })
            
            all_script_results.append(result)
        
        # Sort by status priority, then by usage count (descending), then by name
        all_script_results.sort(key=lambda x: x['_sort_key'])
        
        # Remove the sort key before returning
        for result in all_script_results:
            del result['_sort_key']
            # Keep _row_color in results - it should be handled by Excel generation
            results.append(result)
        
        return results
        
    except Exception as e:
        print(f"Error in UnusedScriptsCheck: {e}")
        import traceback
        traceback.print_exc()
        return []


def find_javascript_context(elem):
    """Find the context (layout/button) for a JavaScript step"""
    layout_name = "Unknown"
    obj_info = ""
    current = elem
    
    for _ in range(20):
        parent = current.getparent()
        if parent is None:
            break
        
        if parent.tag == "Layout":
            layout_name = parent.attrib.get("name", "Unknown Layout")
        elif parent.tag == "ButtonObj":
            obj_info = "Button"
        elif parent.tag == "Object":
            obj_type = parent.attrib.get("type", "")
            obj_name = parent.attrib.get("name", "")
            if obj_name:
                obj_info = f"{obj_type} '{obj_name}'"
            else:
                obj_info = obj_type
        
        current = parent
    
    context = layout_name
    if obj_info:
        context += f" - {obj_info}"
    
    return context


def process_script_reference(elem, script_name, script_usage, scripts):
    """Process a single script reference and categorize its usage"""
    
    # Check for debug mode
    import sys
    debug_mode = '-debug' in sys.argv or '--debug' in sys.argv
    
    # Get the parent path to understand context
    path_elements = []
    current = elem
    for _ in range(20):  # Go up to 20 levels
        parent = current.getparent()
        if parent is None:
            break
        path_elements.append(parent.tag)
        current = parent
    
    path = " > ".join(reversed(path_elements))
    
    # Check various contexts
    found_context = False
    
    # 1. Scripts calling scripts (Perform Script steps)
    # Look for Step in the path, regardless of other elements
    if "Step" in path:
        # Check if this is a Perform Script or Perform Script on Server step
        step_elem = elem.getparent()
        if step_elem is not None and step_elem.tag == "Step":
            step_id = step_elem.attrib.get("id", "")
            step_name = step_elem.attrib.get("name", "")
            
            # Check for Perform Script (id=1) or Perform Script on Server (id=164)
            if step_id in ["1", "164"] or "Perform Script" in step_name:
                # Find the calling script
                calling_script = None
                current = elem
                for i in range(25):
                    parent = current.getparent()
                    if parent is None:
                        break
                    
                    if parent.tag == "Script" and parent.attrib.get("name"):
                        # Get the calling script name
                        potential_name = parent.attrib.get("name")
                        
                        if potential_name in scripts:
                            calling_script = potential_name
                            break
                    current = parent
                
                if calling_script:
                    # Allow self-references for recursive scripts
                    script_usage[script_name]['from_scripts'].append(calling_script)
                    found_context = True
                    return  # Early return to avoid double-counting
    
    # 2. Buttons and layout objects
    # Look for ButtonObj or any layout object that might contain a script
    if any(x in path for x in ["ButtonObj", "ButtonBarSegment", "Object", "Layout"]):
        # Find the layout name
        layout_name = None
        obj_info = ""
        current = elem
        
        for _ in range(20):
            parent = current.getparent()
            if parent is None:
                break
            
            if parent.tag == "Layout":
                layout_name = parent.attrib.get("name", "Unknown Layout")
            elif parent.tag == "ButtonObj":
                obj_info = "Button"
                # Try to get button text from StepText
                step_text = elem.getparent()
                if step_text is not None:
                    text_elem = step_text.find(".//StepText")
                    if text_elem is not None and text_elem.text:
                        # Extract readable part from StepText
                        text = text_elem.text
                        if "Perform Script" in text:
                            text = text.replace("Perform Script", "").strip()
                        obj_info = f"Button (Script: {text[:30]}...)" if len(text) > 30 else f"Button (Script: {text})"
            elif parent.tag == "Object":
                obj_type = parent.attrib.get("type", "Unknown")
                obj_name = parent.attrib.get("name", "")
                if obj_name:
                    obj_info = f"{obj_type} '{obj_name}'"
                else:
                    obj_info = obj_type
            elif parent.tag == "ButtonBarSegment":
                obj_info = "ButtonBar Segment"
            
            current = parent
        
        if layout_name or obj_info:
            button_desc = f"{layout_name or 'Unknown Layout'}"
            if obj_info:
                button_desc += f" - {obj_info}"
            script_usage[script_name]['from_buttons'].append(button_desc)
            found_context = True
            return  # Early return
    
    # 3. Script triggers
    if any(x in path for x in ["Trigger", "ScriptTriggers", "OnRecordLoad", "OnRecordCommit", 
                               "OnLayoutEnter", "OnObjectModify", "OnObjectEnter", "OnObjectExit"]):
        # Find trigger type and context
        trigger_type = None
        context_info = ""
        layout_name = "Unknown Layout"
        object_name = ""
        field_name = ""
        current = elem
        
        # Check parent elements for trigger info
        for _ in range(15):
            parent = current.getparent()
            if parent is None:
                break
            
            # Check various trigger formats
            if parent.tag == "Trigger":
                trigger_type = parent.attrib.get("event", parent.attrib.get("name", "Trigger"))
            elif parent.tag in ["OnRecordLoad", "OnRecordCommit", "OnRecordRevert",
                              "OnLayoutEnter", "OnLayoutExit", "OnLayoutKeystroke",
                              "OnModeEnter", "OnModeExit", "OnViewChange",
                              "OnObjectEnter", "OnObjectExit", "OnObjectModify",
                              "OnObjectKeystroke", "OnObjectSave", "OnObjectValidate",
                              "OnPanelSwitch", "OnTabSwitch", "OnFileAVPlayerChange",
                              "OnGestureTap", "OnExternalCommandReceived",
                              "OnWindowTransaction", "OnFileWindowOpen", "OnFileWindowClose"]:
                trigger_type = parent.tag
            
            # Get context (layout, field, object)
            if parent.tag == "Layout":
                layout_name = parent.attrib.get("name", "Unknown Layout")
            elif parent.tag == "Field":
                field_name = parent.attrib.get("name", "Unknown Field")
                # Look for table
                table_name = "Unknown Table"
                temp_parent = parent
                for _ in range(5):
                    grandparent = temp_parent.getparent()
                    if grandparent is not None and grandparent.tag == "BaseTable":
                        table_name = grandparent.attrib.get("name", table_name)
                        break
                    temp_parent = grandparent
                field_name = f"{table_name}::{field_name}"
            elif parent.tag == "Object":
                obj_type = parent.attrib.get("type", "Unknown")
                obj_name = parent.attrib.get("name", "")
                if obj_name:
                    object_name = obj_name
                else:
                    object_name = obj_type
            
            current = parent
        
        # Build the context description
        if field_name:
            context_info = f"{layout_name} - Field: {field_name}"
        elif object_name:
            context_info = f"{layout_name} - Object: {object_name}"
        else:
            context_info = layout_name
        
        if trigger_type:
            trigger_desc = f"{context_info} - {trigger_type}"
            script_usage[script_name]['from_triggers'].append(trigger_desc)
            found_context = True
            return  # Early return
    
    # 4. Custom menus
    if "CustomMenu" in path:
        menu_info = []
        current = elem
        
        for _ in range(10):
            parent = current.getparent()
            if parent is None:
                break
            
            if parent.tag == "CustomMenuItem":
                menu_info.append(parent.attrib.get("name", "Unknown Item"))
            elif parent.tag == "CustomMenu":
                menu_info.append(parent.attrib.get("name", "Unknown Menu"))
            elif parent.tag == "CustomMenuSet":
                menu_info.append(parent.attrib.get("name", "Unknown Menu Set"))
            
            current = parent
        
        if menu_info:
            menu_desc = " > ".join(reversed(menu_info))
            script_usage[script_name]['from_menus'].append(menu_desc)
            found_context = True
            return  # Early return
    
    # 5. File Options and Window Triggers
    if "FileOptions" in path or "WindowTriggers" in path:
        if any(x in path for x in ["OnOpen", "OnFirstWindowOpen"]):
            script_usage[script_name]['from_other'].append("File Options - OnFirstWindowOpen")
            found_context = True
        elif any(x in path for x in ["OnClose", "OnLastWindowClose"]):
            script_usage[script_name]['from_other'].append("File Options - OnLastWindowClose")
            found_context = True
        elif "OnWindowOpen" in path:
            script_usage[script_name]['from_other'].append("Window Trigger - OnWindowOpen")
            found_context = True
        elif "OnWindowClose" in path:
            script_usage[script_name]['from_other'].append("Window Trigger - OnWindowClose")
            found_context = True
        if found_context:
            return  # Early return
    
    # 6. Web Viewers
    if "ExternalObj" in path or "WebViewer" in path:
        # Find layout context
        layout_name = "Unknown Layout"
        obj_name = "Web Viewer"
        current = elem
        
        for _ in range(10):
            parent = current.getparent()
            if parent is None:
                break
            
            if parent.tag == "Layout":
                layout_name = parent.attrib.get("name", layout_name)
            elif parent.tag == "Object":
                obj_name = parent.attrib.get("name", "Web Viewer")
            
            current = parent
        
        wv_desc = f"{layout_name} - {obj_name}"
        script_usage[script_name]['from_web_viewers'].append(wv_desc)
        found_context = True
        return  # Early return
    
    # 7. Value Lists (less common but possible)
    if "ValueList" in path:
        vl_name = "Unknown Value List"
        current = elem
        
        for _ in range(5):
            parent = current.getparent()
            if parent is None:
                break
            
            if parent.tag == "ValueList":
                vl_name = parent.attrib.get("name", vl_name)
                break
            
            current = parent
        
        script_usage[script_name]['from_value_lists'].append(vl_name)
        found_context = True
        return  # Early return
    
    # If we couldn't identify the context, add to debug
    if not found_context:
        # Get immediate parent info
        parent = elem.getparent()
        parent_info = f"{parent.tag if parent is not None else 'None'}"
        if parent is not None:
            # Add any relevant attributes
            if "name" in parent.attrib:
                parent_info += f" name='{parent.attrib['name']}'"
            if "type" in parent.attrib:
                parent_info += f" type='{parent.attrib['type']}'"
        
        debug_info = f"Path: {path} | Parent: {parent_info} | Element: {elem.tag}"
        script_usage[script_name]['debug_unknown'].append(debug_info)
