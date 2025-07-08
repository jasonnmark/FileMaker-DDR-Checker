from lxml import etree as ET
import re
from collections import defaultdict

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "SQL Usage"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 5

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Error Status": 100,  # New column at the beginning
        "Category": 55,
        "Details": 175,
        "Other Info": 170,
        "Tables": 150,
        "Fields": 245,
        "SQL Text": 280,  # Note: special handling in main script sets this to 40
        "XML Line": 50,
        "Field from XML": 175,
        "Commented?": 100,
        "Raw Field Matches": 200,
        "Missing Tables": 150,
        "Table Exists": 80,
        "Missing Fields": 245,
        "Fields Exist": 80,
        "Base Tables": 175,
        "Debug Info": 200,
        "XML Path": 200,
        "Field Name": 150,
        "Errors": 150,
        "Base Table Name (Debug)": 200,
        "BaseTable Match Tables (Debug)": 200
    }

def apply_styling(ws):
    """Apply custom styling to the SQL Usage worksheet"""
    # Get the color helper functions that were injected
    from openpyxl.styles import PatternFill, Font
    
    # Define category colors
    cat_colors = {
        "Script": get_color_fill('category_script'),
        "Custom Function": get_color_fill('category_custom_function'),
        "Field Calc": get_color_fill('category_field_calc'),
        "Layout Object": get_color_fill('category_layout_object'),
        "Other": get_color_fill('category_other')
    }
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Get row data
        error_status = row[columns.get("Error Status", 0) - 1].value if "Error Status" in columns else None
        category = row[columns.get("Category", 0) - 1].value if "Category" in columns else None
        
        # Color code by error status
        if error_status:
            if error_status == "Missing Table":
                # Apply strong error styling to Error Status cell
                if "Error Status" in columns:
                    row[columns["Error Status"] - 1].fill = get_color_fill('error_strong')
                    row[columns["Error Status"] - 1].font = get_color_font('error_strong', bold=True)
            elif error_status == "Missing Field":
                # Apply error styling
                if "Error Status" in columns:
                    row[columns["Error Status"] - 1].fill = get_color_fill('error')
                    row[columns["Error Status"] - 1].font = get_color_font('error', bold=True)
            elif error_status == "Not Base Table":
                # Apply warning styling
                if "Error Status" in columns:
                    row[columns["Error Status"] - 1].fill = get_color_fill('warning')
                    row[columns["Error Status"] - 1].font = get_color_font('warning', bold=True)
        
        # Category coloring
        if category and category in cat_colors and "Category" in columns:
            row[columns["Category"] - 1].fill = cat_colors[category]
        
        # Missing tables/fields styling
        if "Missing Tables" in columns:
            cell = row[columns["Missing Tables"] - 1]
            if cell.value and cell.value.strip() and cell.value != "None":
                cell.fill = get_color_fill('error')
                cell.font = get_color_font('error', bold=True)
        
        if "Missing Fields" in columns:
            cell = row[columns["Missing Fields"] - 1]
            if cell.value and cell.value.strip() and cell.value != "None":
                cell.fill = get_color_fill('error')
                cell.font = get_color_font('error', bold=True)
        
        # Tables column - color based on existence and warnings
        if "Tables" in columns and "Table Exists" in columns:
            tbl_cell = row[columns["Tables"] - 1]
            exists = str(row[columns["Table Exists"] - 1].value or "").strip().lower()
            
            # Check if table has warning emoji (table occurrence)
            if tbl_cell.value and "⚠️" in str(tbl_cell.value):
                # Warning styling for table occurrences
                tbl_cell.fill = get_color_fill('warning')
                tbl_cell.font = get_color_font('warning', bold=True)
            elif exists == "no":
                # Error styling for missing tables
                tbl_cell.fill = get_color_fill('error_strong')
                tbl_cell.font = get_color_font('error_strong', bold=True)
            else:
                # Muted styling for existing tables
                tbl_cell.font = get_color_font('muted')
        
        # Fields column - color based on existence
        if "Fields" in columns and "Fields Exist" in columns:
            fld_cell = row[columns["Fields"] - 1]
            exists = str(row[columns["Fields Exist"] - 1].value or "").strip().lower()
            
            if exists == "no":
                fld_cell.fill = get_color_fill('error_strong')
                fld_cell.font = get_color_font('error_strong', bold=True)
            else:
                fld_cell.font = get_color_font('muted')
        
        # Commented SQL styling
        if "SQL Text" in columns and "Commented?" in columns:
            commented_cell = row[columns["Commented?"] - 1]
            sql_cell = row[columns["SQL Text"] - 1]
            
            if (commented_cell.value and 
                str(commented_cell.value).strip().lower() == "commented out"):
                # Apply grey color and strikethrough to SQL text
                sql_cell.font = get_color_font('muted', strike=True)
                sql_cell.fill = get_color_fill('muted')
        
        # Format Details column based on Category
        if category == "Script" and "Details" in columns:
            # Bold script names
            row[columns["Details"] - 1].font = Font(bold=True)
            
            # Also bold step number in Other Info if present
            if "Other Info" in columns:
                other_info_cell = row[columns["Other Info"] - 1]
                if other_info_cell.value and str(other_info_cell.value).startswith("Step "):
                    other_info_cell.font = Font(bold=True)

# The rest of the module remains the same...
def process_sql_calls(raw_xml):
    """
    Process all SQL calls in the DDR.
    This is the main SQL processing logic extracted from FilemakerErrorChecker.
    """
    try:
        # Replace double question marks
        raw_xml = raw_xml.replace("??", "*")
        normalized_xml = ''.join(raw_xml.split())
        
        def find_text_line(text):
            """Fast whitespace-agnostic lookup in the one normalized buffer."""
            key = ''.join(text.strip().split())
            idx = normalized_xml.find(key)
            if idx == -1:
                return ""
            return raw_xml[:idx].count('\n') + 1
        
        # Parse the XML
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        root = ET.fromstring(raw_xml.encode('utf-8'), parser)
        
        # Build DDR catalog
        ddr_tables, ddr_fields, table_occurrences_to_base = build_ddr_catalog(root)
        
        results = []
        
        # Regex patterns
        table_regex = re.compile(r'(?:FROM|JOIN)\s+([a-zA-Z0-9_]+)', re.IGNORECASE)
        comment_regex = re.compile(r'^\s*/\*.*?\*/\s*|^\s*--.*?$|/\*.*?\*/', re.MULTILINE | re.DOTALL)
        
        def is_commented_sql(sql_text):
            cleaned = comment_regex.sub('', sql_text)
            return 'ExecuteSQL' not in cleaned
        
        def extract_all_fields_from_sql(sql_text):
            """Extract all field references from SQL"""
            fields = set()
            raw_matches = []
            
            # Find actual ExecuteSQL function calls
            executesql_pattern = r'ExecuteSQL\s*\(\s*["\']([^"\']*)["\']'
            sql_queries = re.findall(executesql_pattern, sql_text, re.IGNORECASE | re.DOTALL)
            
            if not sql_queries:
                return [], ["No ExecuteSQL function calls found"]
            
            for actual_sql in sql_queries:
                raw_matches.append(f"extracted_sql: {actual_sql[:50]}...")
                
                # Extract the main table from FROM clause
                from_match = re.search(r'FROM\s+([a-zA-Z0-9_]+)', actual_sql, re.IGNORECASE)
                if not from_match:
                    continue
                    
                main_table = from_match.group(1)
                raw_matches.append(f"main_table: {main_table}")
                
                # Find alias.field patterns
                alias_field_regex = re.compile(r'([a-zA-Z0-9_]+)\.([a-zA-Z0-9_]+)', re.IGNORECASE)
                alias_matches = alias_field_regex.findall(actual_sql)
                for alias, field in alias_matches:
                    if alias.upper() not in ['SUM', 'COUNT', 'AVG', 'MAX', 'MIN', 'TRIM', 'LEFT', 'RIGHT']:
                        fields.add(f"{alias}.{field}")
                        raw_matches.append(f"alias:{alias}.{field}")
                
                # Find field names in various SQL contexts
                field_patterns = [
                    (r'WHERE\s+([a-zA-Z0-9_]+)\s*[=<>!]', 'WHERE'),
                    (r'AND\s+([a-zA-Z0-9_]+)\s*[=<>!]', 'AND'),
                    (r'OR\s+([a-zA-Z0-9_]+)\s*[=<>!]', 'OR'),
                    (r'SELECT\s+([^\s,()]+)(?:\s*[,)]|\s+FROM)', 'SELECT'),
                    (r'SELECT\s+SUM\s*\(\s*([a-zA-Z0-9_]+)\s*\)', 'SELECT_SUM'),
                    (r'SELECT\s+COUNT\s*\(\s*(?:DISTINCT\s+)?([a-zA-Z0-9_]+)\s*\)', 'SELECT_COUNT'),
                    (r'SELECT\s+AVG\s*\(\s*([a-zA-Z0-9_]+)\s*\)', 'SELECT_AVG'),
                    (r'SELECT\s+MAX\s*\(\s*([a-zA-Z0-9_]+)\s*\)', 'SELECT_MAX'),
                    (r'SELECT\s+MIN\s*\(\s*([a-zA-Z0-9_]+)\s*\)', 'SELECT_MIN'),
                    (r',\s*([a-zA-Z0-9_]+)(?:\s*[,)]|\s+FROM)', 'SELECT_LIST'),
                ]
                
                for pattern, context in field_patterns:
                    matches = re.findall(pattern, actual_sql, re.IGNORECASE)
                    for field in matches:
                        if field.upper() not in [
                            'FROM', 'WHERE', 'SELECT', 'ORDER', 'GROUP', 'BY', 'AS', 'AND', 'OR', 
                            'SUM', 'COUNT', 'AVG', 'MAX', 'MIN', 'DISTINCT', 'NULL', 'TRUE', 'FALSE',
                            'BETWEEN', 'IN', 'NOT', 'IS', 'LIKE', 'EXISTS'
                        ] and not field.isdigit():
                            fields.add(f"{main_table}::{field}")
                            raw_matches.append(f"{context}_in_{main_table}:{field}")
                
                # ORDER BY, GROUP BY
                order_group_patterns = [
                    (r'ORDER\s+BY\s+([a-zA-Z0-9_]+)', 'ORDER_BY'),
                    (r'GROUP\s+BY\s+([a-zA-Z0-9_]+)', 'GROUP_BY'),
                ]
                
                for pattern, context in order_group_patterns:
                    matches = re.findall(pattern, actual_sql, re.IGNORECASE)
                    for field in matches:
                        if not field.isdigit():
                            fields.add(f"{main_table}::{field}")
                            raw_matches.append(f"{context}:{field}")
                
                # Handle JOIN aliases
                join_regex = re.compile(r'JOIN\s+([a-zA-Z0-9_]+)(?:\s+AS\s+([a-zA-Z0-9_]+))?', re.IGNORECASE)
                join_matches = join_regex.findall(actual_sql)
                table_aliases = {main_table: main_table}
                for table, alias in join_matches:
                    table_aliases[alias or table] = table
                
                resolved_fields = set()
                for field_ref in list(fields):
                    if '.' in field_ref and '::' not in field_ref:
                        alias_part, field_part = field_ref.split('.', 1)
                        if alias_part in table_aliases:
                            resolved_fields.add(f"{table_aliases[alias_part]}::{field_part}")
                            fields.remove(field_ref)
                        else:
                            resolved_fields.add(field_ref)
                fields.update(resolved_fields)
            
            return list(fields), raw_matches
        
        def validate_table_field_existence(tables, fields, ddr_tables, ddr_fields, table_occurrences_to_base):
            """Check if tables and fields exist in DDR"""
            missing_tables = []
            missing_fields = []
            table_occurrence_warnings = []
            
            for table in tables:
                if table in table_occurrences_to_base:
                    base_table = table_occurrences_to_base[table]
                    if table != base_table:
                        table_occurrence_warnings.append(table)
                        if base_table not in ddr_tables:
                            missing_tables.append(table)
                    elif table not in ddr_tables:
                        missing_tables.append(table)
                elif table not in ddr_tables:
                    missing_tables.append(table)
            
            for field_ref in fields:
                if '::' in field_ref:
                    table, field = field_ref.split('::', 1)
                    actual_table = table_occurrences_to_base.get(table, table)
                    
                    if actual_table in ddr_tables:
                        if field not in ddr_fields.get(actual_table, set()):
                            if table != actual_table:
                                missing_fields.append(f"{table}::{field} (TO→{actual_table})")
                            else:
                                missing_fields.append(field_ref)
                    else:
                        missing_fields.append(field_ref)
                elif '.' in field_ref:
                    table, field = field_ref.split('.', 1)
                    actual_table = table_occurrences_to_base.get(table, table)
                    
                    if actual_table in ddr_tables:
                        if field not in ddr_fields.get(actual_table, set()):
                            if table != actual_table:
                                missing_fields.append(f"{field_ref} (TO→{actual_table})")
                            else:
                                missing_fields.append(field_ref)
                    else:
                        missing_fields.append(f"{field_ref} (possible alias)")
                else:
                    found = False
                    for tbl, tbl_fields in ddr_fields.items():
                        if field_ref in tbl_fields:
                            found = True
                            break
                    if not found:
                        missing_fields.append(f"{field_ref} (unqualified)")
            
            return missing_tables, missing_fields, table_occurrence_warnings
        
        def extract_field_from_xml_path(node, path):
            """Extract field name and base table name when XML path contains '> Field >'"""
            field_from_xml = ""
            
            if "> Field >" in path:
                field_name = node.attrib.get("name", "")
                table_name = ""
                base_table_name = ""
                
                if not field_name:
                    current = node
                    max_levels = 15
                    level = 0
                    
                    while current is not None and level < max_levels:
                        if current.tag == "Field" and "name" in current.attrib:
                            field_name = current.attrib["name"]
                            if "table" in current.attrib:
                                table_name = current.attrib["table"]
                            break
                        current = current.getparent()
                        level += 1
                
                current = node
                level = 0
                while current is not None and level < 15:
                    if current.tag == "BaseTable" and "name" in current.attrib:
                        base_table_name = current.attrib["name"]
                        break
                    parent = current.getparent()
                    if parent is not None and parent.tag == "BaseTable" and "name" in parent.attrib:
                        base_table_name = parent.attrib["name"]
                        break
                    current = parent
                    level += 1
                
                final_table_name = base_table_name or table_name
                
                if final_table_name and field_name:
                    field_from_xml = f"{final_table_name}::{field_name}"
                elif field_name:
                    field_from_xml = field_name
                elif final_table_name:
                    field_from_xml = f"{final_table_name}::?"
            
            return field_from_xml
        
        def extract_type_and_details(node, path, layout_name, script_name, step_name, step_number, parent_info, sql_line=None):
            try:
                line_info = f"XML_LINE: {sql_line or getattr(node, 'sourceline', 'unknown')} | "
                if script_name:
                    step_info = step_name or "Script Step"
                    if step_number is not None:
                        step_info = f"Step {step_number}: {step_name or ''}"
                    return {
                        "Category": "Script",
                        "Details": script_name,
                        "Other Info": step_info,
                        "Debug Info": f"{line_info}XML_PATH: {path} | Node: {node.tag}, Attributes: {node.attrib}",
                    }
                elif "BaseTableCatalog" in path and "FieldCatalog" in path and "Calculation" in path:
                    base_table_name = ""
                    current = node
                    max_levels = 10
                    level = 0
                    while current is not None and level < max_levels:
                        if current.tag == "BaseTable" and "name" in current.attrib:
                            base_table_name = current.attrib["name"]
                            break
                        current = current.getparent()
                        level += 1
                    
                    table_name = base_table_name or node.attrib.get("table", "Unknown Table")
                    field_name = parent_info.get("name", "Unknown Field") if parent_info else "Unknown Field"
                    return {
                        "Category": "Field Calculation",
                        "Details": f"{table_name}::{field_name}",
                        "Other Info": "Field Calculation",
                        "Debug Info": f"{line_info}XML_PATH: {path} | Node: {node.tag}, Attributes: {node.attrib}",
                    }
                elif "Field" in path and "Calculation" in path:
                    table_name = node.attrib.get("table", "Unknown Table")
                    field_name = parent_info.get("name", "Unknown Field") if parent_info else "Unknown Field"
                    return {
                        "Category": "Field Calculation",
                        "Details": f"{table_name}::{field_name}",
                        "Other Info": f"Layout: {layout_name} Field Calculation",
                        "Debug Info": f"{line_info}XML_PATH: {path} | Node: {node.tag}, Attributes: {node.attrib}",
                    }
                elif "CustomFunctionCatalog" in path:
                    custom_function_name = parent_info.get("name", "Unknown Function") if parent_info else "Unknown Function"
                    return {
                        "Category": "Custom Function",
                        "Details": custom_function_name,
                        "Other Info": "Custom Function",
                        "Debug Info": f"{line_info}XML_PATH: {path} | Node: {node.tag}, Attributes: {node.attrib}",
                    }
                else:
                    return {
                        "Category": "Other",
                        "Details": "Uncategorized",
                        "Other Info": "Unknown",
                        "Debug Info": f"{line_info}XML_PATH: {path} | Node: {node.tag}, Attributes: {node.attrib}",
                    }
            except Exception as e:
                line = getattr(node, 'sourceline', 'unknown')
                return {
                    "Category": "Error",
                    "Details": "Error Parsing Details",
                    "Other Info": "Unknown",
                    "Debug Info": f"XML_LINE: {line} | XML_PATH: {path} | Error: {e}",
                }
        
        def process_object(node, layout_name, ddr_tables, ddr_fields, table_occurrences_to_base):
            try:
                line = getattr(node, 'sourceline', 'unknown')
                line_info = f"XML_LINE: {line} | "
                object_type = node.attrib.get("type", "Unknown Type")
                
                if object_type == "ExternalObject":
                    external_obj = node.find("ExternalObj")
                    if external_obj is not None:
                        type_id = external_obj.attrib.get("typeID", "")
                        if type_id == "WEBV":
                            object_type = "WebViewer"
                        else:
                            object_type = f"External ({type_id})"
                
                bounds = node.find("Bounds")
                object_name = node.attrib.get("name", "")
                if not object_name:
                    object_name_node = node.find("./Styles/CustomStyles/Name")
                    object_name = object_name_node.text if object_name_node is not None else "Unnamed Object"
                
                calculation_paths = [
                    "./HideCondition/Calculation",
                    "./ConditionalFormatting/Calculation",
                    ".//Calculation",
                    "./ExternalObj/Calculation"
                ]
                
                position = ""
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                else:
                    position = "Unknown Position"

                sql_text = ""
                context = ""
                sql_line = line

                for calc_path in calculation_paths:
                    calc_node = node.find(calc_path)
                    if calc_node is not None and calc_node.text and "ExecuteSQL" in calc_node.text:
                        sql_text = calc_node.text.strip()
                        sql_line = getattr(calc_node, 'sourceline', sql_line)
                        
                        if "HideCondition" in calc_path:
                            context = "Hide"
                        elif "ConditionalFormatting" in calc_path:
                            context = "Conditional"
                        elif object_type == "WebViewer":
                            context = "WebViewer"
                        else:
                            context = "Calculation"
                        break

                if sql_text:
                    is_commented = is_commented_sql(sql_text)
                    tables = list(set(table_regex.findall(sql_text)))
                    all_fields, raw_matches = extract_all_fields_from_sql(sql_text)
                    
                    warned_tables = []
                    base_tables = []
                    for table in tables:
                        if table in table_occurrences_to_base:
                            base_table = table_occurrences_to_base[table]
                            if table != base_table:
                                warned_tables.append(f"⚠️ {table}")
                            else:
                                warned_tables.append(table)
                            base_tables.append(base_table)
                        else:
                            warned_tables.append(table)
                            base_tables.append(table)
                    
                    missing_tables, missing_fields, table_occurrence_warnings = validate_table_field_existence(
                        tables, all_fields, ddr_tables, ddr_fields, table_occurrences_to_base
                    )
                    
                    # Determine error status similar to how unused-scripts-check determines status
                    error_status = ""
                    if missing_tables:
                        error_status = "Missing Table"
                    elif missing_fields:
                        error_status = "Missing Field"
                    elif table_occurrence_warnings:
                        error_status = "Not Base Table"
                    
                    error_reasons = []
                    if missing_tables:
                        error_reasons.append("Missing tables")
                    if missing_fields:
                        error_reasons.append("Missing fields")
                    if table_occurrence_warnings:
                        error_reasons.append("Table occurrence")
                    
                    has_errors = bool(error_reasons)
                    error_text = "; ".join(error_reasons) if error_reasons else ""
                    xml_line = sql_line or find_text_line(sql_text)

                    return {
                        "Error Status": error_status,
                        "Category": "Layout Object",
                        "Details": f"{object_type} (Name: {object_name})",
                        "Other Info": f"Layout: {layout_name}, {position}, {context}",
                        "SQL Text": sql_text,
                        "Tables": "\n".join(warned_tables),
                        "Base Tables": "\n".join(base_tables),
                        "Fields": "\n".join(all_fields),
                        "Raw Field Matches": "\n".join(raw_matches),
                        "Commented?": "commented out" if is_commented else "",
                        "Missing Tables": "\n".join(missing_tables) if missing_tables else "None",
                        "Missing Fields": "\n".join(missing_fields) if missing_fields else "None",
                        "Table Exists": "No" if missing_tables else "Yes",
                        "Fields Exist": "No" if missing_fields else "Yes",
                        "Debug Info": f"{line_info}XML_PATH: Layout_Object | Node: {node.tag}, Attributes: {node.attrib}",
                        "XML Path": "Layout_Object",
                        "Field Name": "",
                        "Field from XML": "",
                        "XML Line": xml_line,
                        "Errors": error_text,
                        "Base Table Name (Debug)": "\n".join([f"{t} → {table_occurrences_to_base.get(t, 'N/A')}" for t in tables]),
                        "BaseTable Match Tables (Debug)": "\n".join([f"{t}: {'Yes' if t not in table_occurrences_to_base or t == table_occurrences_to_base.get(t) else 'No'}" for t in tables]),
                    }
                return None

            except Exception as e:
                line = getattr(node, 'sourceline', 'unknown')
                return {
                    "Error Status": "",
                    "Category": "Error",
                    "Details": "Error Parsing Object",
                    "Other Info": f"Layout: {layout_name}",
                    "SQL Text": "",
                    "Tables": "",
                    "Base Tables": "",
                    "Fields": "",
                    "Raw Field Matches": "",
                    "Commented?": "",
                    "Missing Tables": "None",
                    "Missing Fields": "None",
                    "Table Exists": "Unknown",
                    "Fields Exist": "Unknown",
                    "Debug Info": f"{line_info}XML_PATH: Layout_Object_Error | Error: {e}",
                    "XML Path": "Layout_Object_Error",
                    "Field Name": "",
                    "Field from XML": "",
                    "XML Line": line,
                    "Errors": "",
                    "Base Table Name (Debug)": "",
                    "BaseTable Match Tables (Debug)": "",
                }
        
        def search_node(node, path="Root", parent_info=None, layout_name=None, script_name=None, step_name=None, step_number=None, script_steps=None):
            if node.tag == "Layout":
                layout_name = node.attrib.get("name", layout_name)
            if node.tag == "Script":
                script_name = node.attrib.get("name", script_name)
                script_steps = list(node.findall(".//Step"))
            if node.tag == "Step":
                step_name = node.attrib.get("name", step_name)
                try:
                    step_number = int(node.attrib.get("index"))
                except Exception:
                    if script_steps and node in script_steps:
                        step_number = script_steps.index(node) + 1

            is_calculation_in_object = node.tag == "Calculation" and ("> Object >" in path or "> ExternalObject >" in path)
            
            if node.text and "ExecuteSQL" in node.text and not is_calculation_in_object:
                sql_text = node.text.strip()
                
                executesql_pattern = r'ExecuteSQL\s*\(\s*["\']([^"\']*)["\']'
                executesql_matches = list(re.finditer(executesql_pattern, sql_text, re.IGNORECASE | re.DOTALL))
                
                if executesql_matches:
                    if not (node.tag in ['StepText', 'DisplayCalculation'] or 
                        (len(sql_text) < 50 and not sql_text.strip().startswith('ExecuteSQL')) or
                        sql_text.strip().startswith('PatternCount') or
                        sql_text.strip().startswith('If [')):
                        
                        is_commented = is_commented_sql(sql_text)
                        tables = list(set(table_regex.findall(sql_text)))
                        all_fields, raw_matches = extract_all_fields_from_sql(sql_text)
                        
                        missing_tables, missing_fields, table_occurrence_warnings = validate_table_field_existence(
                            tables, all_fields, ddr_tables, ddr_fields, table_occurrences_to_base
                        )
                        
                        # Determine error status similar to how unused-scripts-check determines status
                        error_status = ""
                        if missing_tables:
                            error_status = "Missing Table"
                        elif missing_fields:
                            error_status = "Missing Field"
                        elif table_occurrence_warnings:
                            error_status = "Not Base Table"
                        
                        warned_tables = []
                        base_tables = []
                        for table in tables:
                            if table in table_occurrences_to_base:
                                base_table = table_occurrences_to_base[table]
                                if table != base_table:
                                    warned_tables.append(f"⚠️ {table}")
                                else:
                                    warned_tables.append(table)
                                base_tables.append(base_table)
                            else:
                                warned_tables.append(table)
                                base_tables.append(table)
                        
                        error_reasons = []
                        if missing_tables:
                            error_reasons.append("Missing tables")
                        if missing_fields:
                            error_reasons.append("Missing fields")
                        if table_occurrence_warnings:
                            error_reasons.append("Table occurrence")
                        
                        error_text = "; ".join(error_reasons) if error_reasons else ""
                        xml_line = getattr(node, 'sourceline', None) or find_text_line(sql_text)

                        field_from_xml = extract_field_from_xml_path(node, path)

                        found_field_name = "Unknown Field"
                        field_position = ""
                        field_layout = layout_name or "Unknown Layout"
                        current = node
                        max_levels = 10
                        level = 0
                        
                        while current is not None and level < max_levels:
                            if current.tag == "Field" and "name" in current.attrib:
                                found_field_name = current.attrib["name"]
                                bounds = current.find(".//Bounds")
                                if bounds is not None:
                                    try:
                                        top = round(float(bounds.attrib.get('top', '0')))
                                        left = round(float(bounds.attrib.get('left', '0')))
                                        field_position = f"Top: {top} Left: {left}"
                                    except (ValueError, TypeError):
                                        field_position = ""
                                break
                            current = current.getparent()
                            level += 1

                        details = extract_type_and_details(
                            node, path, layout_name, script_name, step_name, step_number, parent_info, sql_line=xml_line
                        )

                        if found_field_name != "Unknown Field" and "Field Calculation" in details['Category']:
                            table_name = node.attrib.get("table", "Unknown Table")
                            details['Details'] = f"{table_name}::{found_field_name}"

                        field_name_display = found_field_name if "Field Calculation" in details['Category'] and found_field_name != "Unknown Field" else ""
                        
                        if details['Category'] == "Field Calculation":
                            details['Category'] = "Field Calc"
                            if field_from_xml:
                                details['Details'] = field_from_xml
                            details['Other Info'] = ""
                        elif details['Category'] == "Custom Function":
                            details['Other Info'] = ""

                        results.append({
                            "Error Status": error_status,
                            "Category": details['Category'],
                            "Details": details['Details'],
                            "Other Info": details['Other Info'],
                            "SQL Text": sql_text,
                            "Tables": "\n".join(warned_tables),
                            "Base Tables": "\n".join(base_tables),
                            "Fields": "\n".join(all_fields),
                            "Raw Field Matches": "\n".join(raw_matches),
                            "Commented?": "commented out" if is_commented else "",
                            "Missing Tables": "\n".join(missing_tables) if missing_tables else "None",
                            "Missing Fields": "\n".join(missing_fields) if missing_fields else "None",
                            "Table Exists": "No" if missing_tables else "Yes",
                            "Fields Exist": "No" if missing_fields else "Yes",
                            "Debug Info": details['Debug Info'],
                            "XML Path": path,
                            "Field Name": field_name_display,
                            "Field from XML": field_from_xml,
                            "XML Line": xml_line,
                            "Errors": error_text,
                            "Base Table Name (Debug)": "\n".join([f"{t} → {table_occurrences_to_base.get(t, 'N/A')}" for t in tables]),
                            "BaseTable Match Tables (Debug)": "\n".join([f"{t}: {'Yes' if t not in table_occurrences_to_base or t == table_occurrences_to_base.get(t) else 'No'}" for t in tables]),
                        })

            if node.tag in ["Object", "ExternalObject"]:
                obj = process_object(node, layout_name, ddr_tables, ddr_fields, table_occurrences_to_base)
                if obj:
                    results.append(obj)

            for child in node:
                search_node(
                    child,
                    path=f"{path} > {child.tag}",
                    parent_info=node.attrib,
                    layout_name=layout_name,
                    script_name=script_name,
                    step_name=step_name,
                    step_number=step_number,
                    script_steps=script_steps
                )
        
        # Start the search
        search_node(root)
        
        # Sort results: errors first (by Error Status), then by table name
        def sort_key(result):
            # Priority based on Error Status
            error_status = result.get("Error Status", "")
            if error_status == "Missing Table":
                priority = 0
            elif error_status == "Missing Field":
                priority = 1
            elif error_status == "Not Base Table":
                priority = 2
            else:
                priority = 3  # No errors
            
            # Get table name for secondary sort
            tables = result.get("Tables", "")
            first_table = ""
            if tables:
                # Remove warning emoji and get first table
                first_table = tables.split("\n")[0].replace("⚠️ ", "").strip()
            
            return (priority, first_table.lower())
        
        results.sort(key=sort_key)
        
        return results
        
    except Exception as e:
        print(f"Error in process_sql_calls: {e}")
        import traceback
        traceback.print_exc()
        return []

def build_ddr_catalog(root):
    """Build catalogs of tables and fields from DDR"""
    tables = set()
    fields = defaultdict(set)
    table_occurrences_to_base = {}
    
    for table_node in root.findall(".//Table"):
        occurrence_name = table_node.attrib.get("name")
        base_table_name = table_node.attrib.get("baseTable")
        if occurrence_name and base_table_name:
            table_occurrences_to_base[occurrence_name] = base_table_name
    
    for table_occ in root.findall(".//TableOccurrence"):
        occ_name = table_occ.attrib.get("name")
        base_table = table_occ.attrib.get("baseTable")
        if occ_name and base_table:
            table_occurrences_to_base[occ_name] = base_table
    
    for table_node in root.findall(".//BaseTable"):
        table_name = table_node.attrib.get("name")
        if table_name:
            tables.add(table_name)
            for field_node in table_node.findall(".//Field"):
                field_name = field_node.attrib.get("name")
                if field_name:
                    fields[table_name].add(field_name)
    
    for field_catalog in root.findall(".//FieldCatalog/Field"):
        fname = field_catalog.attrib.get("name")
        tname = field_catalog.attrib.get("table")
        if fname and tname:
            tables.add(tname)
            fields[tname].add(fname)
    
    for base_table_catalog in root.findall(".//BaseTableCatalog"):
        for table_entry in base_table_catalog.findall(".//BaseTable"):
            tname = table_entry.attrib.get("name")
            if tname:
                tables.add(tname)
                for fc in table_entry.findall(".//FieldCatalog/Field"):
                    fname = fc.attrib.get("name")
                    if fname:
                        fields[tname].add(fname)
    
    for table_occurrence in root.findall(".//TableOccurrence"):
        bt = table_occurrence.attrib.get("baseTable")
        if bt:
            tables.add(bt)
    
    for table_elem in root.findall(".//Table"):
        tn = table_elem.attrib.get("name")
        if tn:
            tables.add(tn)
    
    return tables, dict(fields), table_occurrences_to_base

def run_check(raw_xml):
    """
    Extract all SQL calls from the raw XML.
    This is the main entry point for the check module.
    """
    try:
        results = process_sql_calls(raw_xml)
        
        # Count errors for debugging (but don't print)
        error_count = sum(1 for r in results if r.get("Errors", "").strip())
        
        # Clean up the results to ensure proper column names
        cleaned_results = []
        for result in results:
            # Make a copy to avoid modifying the original
            cleaned_result = result.copy()
            
            # Rename "Is Commented" to "Commented?" for consistency
            if "Is Commented" in cleaned_result:
                cleaned_result["Commented?"] = cleaned_result.pop("Is Commented")
            
            # Clean display values but keep error indicators
            if cleaned_result.get("Missing Tables") == "None":
                cleaned_result["Missing Tables"] = ""
            if cleaned_result.get("Missing Fields") == "None":
                cleaned_result["Missing Fields"] = ""
                
            cleaned_results.append(cleaned_result)
        
        return cleaned_results
    except Exception as e:
        print(f"Error in SQLCallsCheck.run_check: {e}")
        import traceback
        traceback.print_exc()