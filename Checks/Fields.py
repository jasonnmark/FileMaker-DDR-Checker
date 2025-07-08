from lxml import etree as ET
import re
from collections import defaultdict
import ahocorasick

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Field Usage"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 4  # Place after Custom Function Usage but before SQL Usage

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Table Name": 175,
        "Field Name": 175,
        "Status": 85,
        "Usage Count": 80,
        "XML Count": 65,
        "Used in Layouts": 250,
        "Used in Scripts": 250,
        "Used in Calculations": 250,
        "Used in SQL": 250,
        "Used in Other": 200
    }

def apply_styling(ws):
    """Apply custom styling to the Field Usage worksheet"""
    from openpyxl.styles import PatternFill, Font
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Check status for field categories
        if "Status" in columns:
            status_cell = row[columns["Status"] - 1]
            status_value = status_cell.value if status_cell.value else ""
            
            if status_value == "Not Used":
                # Highlight entire row for unused fields in red
                for cell in row:
                    cell.fill = get_color_fill('error')
                
                # Make status cell bold with error font
                status_cell.font = get_color_font('error', bold=True)
                
                # Also make table name and field name bold
                if "Table Name" in columns:
                    row[columns["Table Name"] - 1].font = get_color_font('error', bold=True)
                if "Field Name" in columns:
                    row[columns["Field Name"] - 1].font = get_color_font('error', bold=True)
            
            elif status_value in ["System", "Cached", "Comment", "Imported"]:
                # Highlight system/cached/comment/imported fields in yellow
                for cell in row:
                    cell.fill = get_color_fill('warning')
                
                # Make status cell bold with warning font
                status_cell.font = get_color_font('warning', bold=True)
                
                # Also make table name and field name bold
                if "Table Name" in columns:
                    row[columns["Table Name"] - 1].font = get_color_font('warning', bold=True)
                if "Field Name" in columns:
                    row[columns["Field Name"] - 1].font = get_color_font('warning', bold=True)
        
        # Color code usage count
        if "Usage Count" in columns:
            count_cell = row[columns["Usage Count"] - 1]
            try:
                count = int(count_cell.value or 0)
                if count == 0:
                    count_cell.font = get_color_font('error', bold=True)
                elif count >= 10:
                    count_cell.font = get_color_font('success', bold=True)
                elif count >= 5:
                    count_cell.font = Font(bold=True, color="388E3C")  # Medium green
            except:
                pass
                
        # Color code XML Count - green if > 2, similar to script usage
        if "XML Count" in columns:
            xml_count_cell = row[columns["XML Count"] - 1]
            try:
                count = int(xml_count_cell.value or 0)
                if count > 2:
                    xml_count_cell.font = get_color_font('success', bold=True)
                    xml_count_cell.fill = PatternFill()  # Remove any background fill
                    
                    # If XML count > 2 but usage count is 0, highlight in warning color
                    if "Usage Count" in columns and "Status" in columns:
                        usage_count_cell = row[columns["Usage Count"] - 1]
                        status_cell = row[columns["Status"] - 1]
                        
                        if (int(usage_count_cell.value or 0) == 0 and 
                            status_cell.value == "Not Used"):
                            # This might be a false negative - field appears in XML but wasn't detected
                            xml_count_cell.fill = get_color_fill('warning')
                            xml_count_cell.font = get_color_font('warning', bold=True)
            except:
                pass

def run_check(raw_xml):
    """
    Find all fields and track where they are used.
    Flag fields that aren't used anywhere.
    Special handling for system fields and cached fields.
    Sort so unused fields appear at the top, followed by system/cached fields, then used fields.
    """
    print("Starting Field Usage Check")  # Debug line
    try:
        # Parse the XML
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        root = ET.fromstring(raw_xml.encode('utf-8'), parser)
        
        # Dictionary to store all fields by table
        fields = defaultdict(dict)
        
        # Dictionary to store field usage by type
        field_usage = defaultdict(lambda: {
            'from_layouts': [],
            'from_scripts': [],
            'from_calculations': [],
            'from_sql': [],
            'from_portals': [],
            'from_value_lists': [],
            'from_relationships': [],
            'from_web_viewers': [],
            'from_other': []
        })
        
        # Build table occurrences to base table mapping
        # This helps us resolve table occurrences to their base tables
        table_occurrences_to_base = {}
        
        for table_node in root.findall(".//Table"):
            occurrence_name = table_node.attrib.get("name")
            base_table_name = table_node.attrib.get("baseTable")
            if occurrence_name and base_table_name:
                table_occurrences_to_base[occurrence_name] = base_table_name
        
        for table_occ in root.findall(".//TableOccurrence"):
            occ_name = table_occ.attrib.get("name")
            base_table = table_occ.attrib.get("baseTable")
            if occ_name and base_table:
                table_occurrences_to_base[occ_name] = base_table
        
        # First, find all fields in all tables
        print("  Finding all fields...")
        
        # Method 1: Look in BaseTable nodes
        for table_node in root.findall(".//BaseTable"):
            table_name = table_node.attrib.get("name")
            if table_name:
                print(f"    Found table: {table_name}")
                for field_node in table_node.findall(".//Field"):
                    field_name = field_node.attrib.get("name")
                    if field_name:
                        fields[table_name][field_name] = {
                            "name": field_name,
                            "id": field_node.attrib.get("id", ""),
                            "type": field_node.attrib.get("dataType", "")
                        }
        
        # Method 2: Look in FieldCatalog nodes
        for field_catalog in root.findall(".//FieldCatalog/Field"):
            field_name = field_catalog.attrib.get("name")
            table_name = field_catalog.attrib.get("table")
            if field_name and table_name:
                # Get the base table if it's a table occurrence
                base_table = table_occurrences_to_base.get(table_name, table_name)
                fields[base_table][field_name] = {
                    "name": field_name,
                    "id": field_catalog.attrib.get("id", ""),
                    "type": field_catalog.attrib.get("dataType", "")
                }
        
        # Method 3: Look in BaseTableCatalog nodes
        for base_table_catalog in root.findall(".//BaseTableCatalog"):
            for table_entry in base_table_catalog.findall(".//BaseTable"):
                table_name = table_entry.attrib.get("name")
                if table_name:
                    for fc in table_entry.findall(".//FieldCatalog/Field"):
                        field_name = fc.attrib.get("name")
                        if field_name:
                            fields[table_name][field_name] = {
                                "name": field_name,
                                "id": fc.attrib.get("id", ""),
                                "type": fc.attrib.get("dataType", "")
                            }
        
        # Count total fields found
        total_fields = sum(len(table_fields) for table_fields in fields.values())
        print(f"  Found {total_fields} fields in {len(fields)} tables")
        
        # If no fields found, return empty results
        if total_fields == 0:
            print("  No fields found in XML")
            return []
        
        # Count XML occurrences for each field
        print("  Counting XML occurrences...")
        xml_counts = {}
        
        # Initialize counts for all fields
        for table_name, table_fields in fields.items():
            for field_name in table_fields:
                field_key = f"{table_name}::{field_name}"
                
                # Skip very short field names to avoid too many false positives
                if len(field_name) <= 2:
                    xml_counts[field_key] = 0
                    continue
                
                # Simple direct count of field name occurrences in the raw XML
                # This ignores table context and just counts all instances
                count = raw_xml.count(field_name)
                xml_counts[field_key] = count
        
        print(f"    XML counting complete")
        
        # Now track field usage
        print("  Tracking field usage...")
        
        # 1. Check for fields used in layouts
        print("    Checking layouts...")
        for layout in root.findall(".//Layout"):
            layout_name = layout.attrib.get("name", "Unknown Layout")
            
            # Process regular field objects on the layout
            for field in layout.findall(".//Object[@type='Field']"):
                field_obj = field.find(".//Field")
                if field_obj is not None:
                    field_name = field_obj.attrib.get("name", "")
                    table_name = field_obj.attrib.get("table", "")
                    
                    # Skip empty fields
                    if not field_name or not table_name:
                        continue
                    
                    # Get the base table if it's a table occurrence
                    base_table = table_occurrences_to_base.get(table_name, table_name)
                    
                    # Get position for context
                    position = ""
                    bounds = field.find(".//Bounds")
                    if bounds is not None:
                        try:
                            top = round(float(bounds.attrib.get('top', '0')))
                            left = round(float(bounds.attrib.get('left', '0')))
                            position = f"Top: {top} Left: {left}"
                        except (ValueError, TypeError):
                            position = "Unknown Position"
                    
                    if base_table in fields and field_name in fields[base_table]:
                        usage_desc = f"Field - {layout_name} {position}"
                        field_key = f"{base_table}::{field_name}"
                        field_usage[field_key]['from_layouts'].append(usage_desc)
                
                # Check for PlaceholderText in field objects
                placeholder_text = field.find(".//PlaceholderText")
                if placeholder_text is not None:
                    # Check both Calculation and DisplayCalculation
                    calc = placeholder_text.find(".//Calculation")
                    if calc is not None and calc.text:
                        # Get position for context
                        position = ""
                        bounds = field.find(".//Bounds")
                        if bounds is not None:
                            try:
                                top = round(float(bounds.attrib.get('top', '0')))
                                left = round(float(bounds.attrib.get('left', '0')))
                                position = f"Top: {top} Left: {left}"
                            except (ValueError, TypeError):
                                position = "Unknown Position"
                        
                        context = f"PlaceholderText - {layout_name} {position}"
                        
                        # Process field references in the placeholder calculation
                        find_field_references(
                            calc.text,
                            None,  # No specific table context
                            context,
                            fields,
                            field_usage,
                            table_occurrences_to_base
                        )
                    
                    # Also check DisplayCalculation for field references
                    display_calc = placeholder_text.find(".//DisplayCalculation")
                    if display_calc is not None:
                        # Look for FieldRef chunks
                        for chunk in display_calc.findall(".//Chunk[@type='FieldRef']"):
                            field_ref = chunk.find(".//Field")
                            if field_ref is not None:
                                ref_table = field_ref.attrib.get("table", "")
                                ref_field = field_ref.attrib.get("name", "")
                                
                                if ref_table and ref_field:
                                    base_table = table_occurrences_to_base.get(ref_table, ref_table)
                                    if base_table in fields and ref_field in fields[base_table]:
                                        # Get position for context
                                        position = ""
                                        bounds = field.find(".//Bounds")
                                        if bounds is not None:
                                            try:
                                                top = round(float(bounds.attrib.get('top', '0')))
                                                left = round(float(bounds.attrib.get('left', '0')))
                                                position = f"Top: {top} Left: {left}"
                                            except (ValueError, TypeError):
                                                position = "Unknown Position"
                                        
                                        usage_desc = f"PlaceholderText - {layout_name} {position}"
                                        field_key = f"{base_table}::{ref_field}"
                                        field_usage[field_key]['from_layouts'].append(usage_desc)
            
            # Process portal fields
            for portal in layout.findall(".//Object[@type='Portal']"):
                portal_obj = portal.find(".//Portal")
                if portal_obj is None:
                    continue
                
                # Get portal position
                position = ""
                bounds = portal.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Get portal table
                portal_table = portal_obj.attrib.get("table", "")
                if not portal_table:
                    continue
                
                # Get base table for portal
                portal_base_table = table_occurrences_to_base.get(portal_table, portal_table)
                
                # Process portal filter calculations
                portal_filter = portal_obj.find(".//FilterCalc")
                if portal_filter is not None:
                    calc = portal_filter.find(".//Calculation")
                    if calc is not None and calc.text:
                        # Find all field references in the calculation
                        find_field_references(
                            calc.text, 
                            portal_base_table, 
                            f"Portal Filter - {layout_name} {position}", 
                            fields, 
                            field_usage, 
                            table_occurrences_to_base
                        )
                
                # Process fields in the portal
                for portal_field in portal.findall(".//Object[@type='Field']"):
                    field_obj = portal_field.find(".//Field")
                    if field_obj is not None:
                        field_name = field_obj.attrib.get("name", "")
                        field_table = field_obj.attrib.get("table", portal_table)
                        
                        if not field_name or not field_table:
                            continue
                        
                        # Get base table for field
                        field_base_table = table_occurrences_to_base.get(field_table, field_table)
                        
                        # Get field position relative to portal
                        field_position = position
                        field_bounds = portal_field.find(".//Bounds")
                        if field_bounds is not None:
                            try:
                                f_top = round(float(field_bounds.attrib.get('top', '0')))
                                f_left = round(float(field_bounds.attrib.get('left', '0')))
                                field_position = f"{position} (Field: Top: {f_top} Left: {f_left})"
                            except (ValueError, TypeError):
                                pass
                        
                        if field_base_table in fields and field_name in fields[field_base_table]:
                            usage_desc = f"Portal Field - {layout_name} {field_position}"
                            field_key = f"{field_base_table}::{field_name}"
                            field_usage[field_key]['from_layouts'].append(usage_desc)
            
            # Check for Web Viewers with field references
            for web_viewer in layout.findall(".//Object[@type='ExternalObject']"):
                # Find Web Viewers (typeID=WEBV)
                external_obj = web_viewer.find(".//ExternalObj")
                if external_obj is not None and external_obj.attrib.get("typeID", "") == "WEBV":
                    # Get position
                    position = ""
                    bounds = web_viewer.find(".//Bounds")
                    if bounds is not None:
                        try:
                            top = round(float(bounds.attrib.get('top', '0')))
                            left = round(float(bounds.attrib.get('left', '0')))
                            position = f"Top: {top} Left: {left}"
                        except (ValueError, TypeError):
                            position = "Unknown Position"
                    
                    # Look for field references in calculations
                    for calc in web_viewer.findall(".//Calculation"):
                        if calc.text:
                            # Web viewers can reference fields from any table
                            for table_name, table_fields in fields.items():
                                for field_name in table_fields:
                                    if f"{table_name}::{field_name}" in calc.text:
                                        usage_desc = f"WebViewer - {layout_name} {position}"
                                        field_key = f"{table_name}::{field_name}"
                                        field_usage[field_key]['from_web_viewers'].append(usage_desc)
                                    elif f"{table_name}." in calc.text and f".{field_name}" in calc.text:
                                        # Check for table.field syntax
                                        usage_desc = f"WebViewer - {layout_name} {position}"
                                        field_key = f"{table_name}::{field_name}"
                                        field_usage[field_key]['from_web_viewers'].append(usage_desc)
                    
                    # Check URL formula (which may contain field references)
                    calc = external_obj.find(".//WebObj/URLCalc/Calculation")
                    if calc is not None and calc.text:
                        for table_name, table_fields in fields.items():
                            for field_name in table_fields:
                                if f"{table_name}::{field_name}" in calc.text:
                                    usage_desc = f"WebViewer URL - {layout_name} {position}"
                                    field_key = f"{table_name}::{field_name}"
                                    field_usage[field_key]['from_web_viewers'].append(usage_desc)
                                elif f"{table_name}." in calc.text and f".{field_name}" in calc.text:
                                    # Check for table.field syntax
                                    usage_desc = f"WebViewer URL - {layout_name} {position}"
                                    field_key = f"{table_name}::{field_name}"
                                    field_usage[field_key]['from_web_viewers'].append(usage_desc)
            
            # Check for Button calculations that reference fields
            for button in layout.findall(".//Object[@type='Button']"):
                # Get position
                position = ""
                bounds = button.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Button name for better context
                button_name = button.attrib.get("name", "Unnamed Button")
                context = f"Button '{button_name}' - {layout_name} {position}"
                
                # Look for script calculations in buttons
                for script_step in button.findall(".//Script"):
                    script_name = script_step.attrib.get("name", "")
                    if script_name:
                        # Button that runs a script - check parameters for field references
                        params = script_step.findall(".//Parameter")
                        for param in params:
                            if param.text:
                                # Field parameters in script calls
                                if "field:" in param.text.lower():
                                    # Parse "field:table::field" or similar patterns
                                    match = re.search(r'field:([^:]+)::([^;"\s]+)', param.text, re.IGNORECASE)
                                    if match:
                                        table_name, field_name = match.groups()
                                        base_table = table_occurrences_to_base.get(table_name, table_name)
                                        if base_table in fields and field_name in fields[base_table]:
                                            usage_desc = f"{context} (Script param)"
                                            field_key = f"{base_table}::{field_name}"
                                            field_usage[field_key]['from_layouts'].append(usage_desc)
                
                # Check button calculations
                for calc in button.findall(".//Calculation"):
                    if calc.text:
                        # Scan all tables for field references
                        for table_name, table_fields in fields.items():
                            for field_name in table_fields:
                                if f"{table_name}::{field_name}" in calc.text:
                                    usage_desc = f"{context} (Calculation)"
                                    field_key = f"{table_name}::{field_name}"
                                    field_usage[field_key]['from_calculations'].append(usage_desc)
                                elif f"{table_name}." in calc.text and f".{field_name}" in calc.text:
                                    # Check for table.field syntax
                                    usage_desc = f"{context} (Calculation)"
                                    field_key = f"{table_name}::{field_name}"
                                    field_usage[field_key]['from_calculations'].append(usage_desc)
            
            # Check for Text objects with field references in FieldList elements
            for text_obj in layout.findall(".//Object[@type='Text']"):
                # Get position
                position = ""
                bounds = text_obj.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Check for field references in FieldList elements
                field_list = text_obj.find(".//TextObj/FieldList")
                if field_list is not None:
                    for field_ref in field_list.findall(".//Field"):
                        field_name = field_ref.attrib.get("name", "")
                        table_name = field_ref.attrib.get("table", "")
                        
                        if field_name and table_name:
                            # Handle emoji table names like 🌎_TableName
                            base_table = table_occurrences_to_base.get(table_name, table_name)
                            
                            if base_table in fields and field_name in fields[base_table]:
                                usage_desc = f"Text Object - {layout_name} {position}"
                                field_key = f"{base_table}::{field_name}"
                                field_usage[field_key]['from_layouts'].append(usage_desc)
                
                # Also check for field references in character style data
                for data_element in text_obj.findall(".//TextObj/CharacterStyleVector/Style/Data"):
                    if data_element.text:
                        # Look for field references like <<field_name>>
                        matches = re.findall(r'<<([^>]+)>>', data_element.text)
                        for field_name in matches:
                            # Since text content might not have table name, check all tables
                            for table_name, table_fields in fields.items():
                                if field_name in table_fields:
                                    usage_desc = f"Text Merge Field - {layout_name} {position}"
                                    field_key = f"{table_name}::{field_name}"
                                    field_usage[field_key]['from_layouts'].append(usage_desc)
            
            # Check for field references in Conditional Formatting
            for obj_with_cond in layout.findall(".//Object"):
                # Only process objects that have conditional formatting
                cond_formatting = obj_with_cond.find(".//ConditionalFormatting")
                if cond_formatting is None:
                    continue
                
                # Get position
                position = ""
                bounds = obj_with_cond.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                obj_type = obj_with_cond.attrib.get("type", "Unknown")
                
                # Process each conditional formatting item
                for cond_item in cond_formatting.findall(".//Item"):
                    # Check calculation
                    calc = cond_item.find(".//Condition/Calc")
                    if calc is not None and calc.text:
                        # Find field references in the calculation
                        find_field_references(
                            calc.text,
                            None,  # No specific table context
                            f"Conditional Format - {layout_name} {obj_type} {position}",
                            fields,
                            field_usage,
                            table_occurrences_to_base
                        )
            
            # Check for field references in Hide Conditions
            for obj_with_hide in layout.findall(".//Object"):
                # Only process objects that have hide conditions
                hide_condition = obj_with_hide.find(".//HideCondition")
                if hide_condition is None:
                    continue
                
                # Get position
                position = ""
                bounds = obj_with_hide.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                obj_type = obj_with_hide.attrib.get("type", "Unknown")
                
                # Check the hide condition calculation
                hide_calc = hide_condition.find(".//Calculation")
                if hide_calc is not None and hide_calc.text:
                    # Find field references in the calculation
                    find_field_references(
                        hide_calc.text,
                        None,  # No specific table context
                        f"Hide Condition - {layout_name} {obj_type} {position}",
                        fields,
                        field_usage,
                        table_occurrences_to_base
                    )
            
            # Check for field references in Tab Controls
            for tab_control in layout.findall(".//Object[@type='TabControl']"):
                # Get position
                position = ""
                bounds = tab_control.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Check for calculations in tab labels
                for panel in layout.findall(f".//Object[@type='TabPanel']"):
                    panel_name = panel.attrib.get("name", "Unnamed Tab")
                    
                    # Check for calculation in tab name
                    tab_obj = panel.find(".//TabControlObj")
                    if tab_obj is not None:
                        name_calc = tab_obj.find(".//Calculation")
                        if name_calc is not None and name_calc.text:
                            find_field_references(
                                name_calc.text,
                                None,  # No specific table context
                                f"Tab Label - {layout_name} {panel_name} {position}",
                                fields,
                                field_usage,
                                table_occurrences_to_base
                            )
            
            # Check for field references in Slide Controls
            for slide_control in layout.findall(".//Object[@type='SlideControl']"):
                # Get position
                position = ""
                bounds = slide_control.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Check for calculations in slide labels
                for panel in layout.findall(f".//Object[@type='SlidePanel']"):
                    panel_name = panel.attrib.get("name", "Unnamed Slide")
                    
                    # Check for calculation in slide name
                    slide_obj = panel.find(".//SlideControlObj")
                    if slide_obj is not None:
                        name_calc = slide_obj.find(".//Calculation")
                        if name_calc is not None and name_calc.text:
                            find_field_references(
                                name_calc.text,
                                None,  # No specific table context
                                f"Slide Label - {layout_name} {panel_name} {position}",
                                fields,
                                field_usage,
                                table_occurrences_to_base
                            )
            
            # Check for field references in Button Bars
            for button_bar in layout.findall(".//Object[@type='ButtonBar']"):
                # Get position
                position = ""
                bounds = button_bar.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                # Check for calculations in button labels
                button_bar_obj = button_bar.find(".//ButtonBarObj")
                if button_bar_obj is not None:
                    for segment in button_bar_obj.findall(".//Segment"):
                        segment_calc = segment.find(".//Calculation")
                        if segment_calc is not None and segment_calc.text:
                            find_field_references(
                                segment_calc.text,
                                None,  # No specific table context
                                f"Button Bar Label - {layout_name} {position}",
                                fields,
                                field_usage,
                                table_occurrences_to_base
                            )
            
            # Check for field references in Popovers
            for popover in layout.findall(".//Object[@type='Popover']"):
                # Get position
                position = ""
                bounds = popover.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                popover_name = popover.attrib.get("name", "Unnamed Popover")
                
                # Check for title calculation
                popover_obj = popover.find(".//PopoverObj")
                if popover_obj is not None:
                    title_calc = popover_obj.find(".//Calculation")
                    if title_calc is not None and title_calc.text:
                        find_field_references(
                            title_calc.text,
                            None,  # No specific table context
                            f"Popover Title - {layout_name} {popover_name} {position}",
                            fields,
                            field_usage,
                            table_occurrences_to_base
                        )
                    
            # Check for field references in Charts
            for obj in layout.findall(".//Object"):
                chart = obj.find(".//ExternalObj[@typeID='CHRT']")
                if chart is None:
                    continue
                
                # Get position
                position = ""
                bounds = obj.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                chart_name = obj.attrib.get("name", "Unnamed Chart")
                
                # Check series fields
                for field_ref in chart.findall(".//ChartSeries//Field"):
                    field_name = field_ref.attrib.get("name", "")
                    table_name = field_ref.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        if base_table in fields and field_name in fields[base_table]:
                            usage_desc = f"Chart Series - {layout_name} {chart_name} {position}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_layouts'].append(usage_desc)
                
                # Check chart calculations
                for calc in chart.findall(".//Calculation"):
                    if calc.text:
                        find_field_references(
                            calc.text,
                            None,  # No specific table context
                            f"Chart Calculation - {layout_name} {chart_name} {position}",
                            fields,
                            field_usage,
                            table_occurrences_to_base
                        )
        
        # 2. Check for fields used in scripts
        print("    Checking scripts...")
        for script in root.findall(".//Script"):
            script_name = script.attrib.get("name", "Unknown Script")
            
            # Look for field references in script steps
            for step in script.findall(".//Step"):
                step_name = step.attrib.get("name", "Unknown Step")
                step_index = step.attrib.get("index", "")
                step_context = f"Script - {script_name} (Step {step_index}: {step_name})"
                
                # Field parameters in script steps
                for param in step.findall(".//Parameter"):
                    if param.text:
                        # Field parameters like "field:table::field"
                        if "field:" in param.text.lower():
                            match = re.search(r'field:([^:]+)::([^;"\s]+)', param.text, re.IGNORECASE)
                            if match:
                                table_name, field_name = match.groups()
                                base_table = table_occurrences_to_base.get(table_name, table_name)
                                if base_table in fields and field_name in fields[base_table]:
                                    usage_desc = f"{step_context} (Param)"
                                    field_key = f"{base_table}::{field_name}"
                                    field_usage[field_key]['from_scripts'].append(usage_desc)
                
                # Check calculations in steps
                for calc in step.findall(".//Calculation"):
                    if calc.text:
                        # Process all field references in the calculation
                        find_field_references(
                            calc.text, 
                            None,  # No specific table context, check all
                            step_context, 
                            fields, 
                            field_usage, 
                            table_occurrences_to_base
                        )
                
                # Special handling for specific script steps
                if step_name in ["Set Field", "Set Field By Name"]:
                    # Get target field for Set Field steps
                    field_ref = step.find(".//Field")
                    if field_ref is not None:
                        field_name = field_ref.attrib.get("name", "")
                        table_name = field_ref.attrib.get("table", "")
                        
                        if field_name and table_name:
                            base_table = table_occurrences_to_base.get(table_name, table_name)
                            if base_table in fields and field_name in fields[base_table]:
                                usage_desc = f"{step_context} (Target field)"
                                field_key = f"{base_table}::{field_name}"
                                field_usage[field_key]['from_scripts'].append(usage_desc)
                
                # Check for Import/Export steps that specify fields
                if step_name in ["Import Records", "Export Records"]:
                    for field_map in step.findall(".//FieldMapping/Field"):
                        field_name = field_map.attrib.get("name", "")
                        table_name = field_map.attrib.get("table", "")
                        
                        if field_name and table_name:
                            base_table = table_occurrences_to_base.get(table_name, table_name)
                            if base_table in fields and field_name in fields[base_table]:
                                usage_desc = f"{step_context} (Field Mapping)"
                                field_key = f"{base_table}::{field_name}"
                                field_usage[field_key]['from_scripts'].append(usage_desc)
        
        # 3. Check for fields used in calculations
        print("    Checking field calculations...")
        for table_node in root.findall(".//BaseTable"):
            table_name = table_node.attrib.get("name")
            if not table_name:
                continue
            
            # Process calculated fields
            for field_node in table_node.findall(".//Field"):
                field_name = field_node.attrib.get("name")
                if not field_name:
                    continue
                
                # Skip if the field itself is not in our catalog
                # (defensive check in case XML structure is unusual)
                if table_name not in fields or field_name not in fields[table_name]:
                    continue
                
                # Check field calculation
                calc = field_node.find(".//Calculation")
                if calc is not None and calc.text:
                    calc_context = f"Field Calculation - {table_name}::{field_name}"
                    
                    # Add usage for the field itself as target of calculation
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(calc_context + " (Target)")
                    
                    # Find field references within the calculation
                    find_field_references(
                        calc.text, 
                        table_name,  # Has table context
                        calc_context, 
                        fields, 
                        field_usage, 
                        table_occurrences_to_base
                    )
                
                # Also check the DisplayCalculation which contains field references in Chunk elements
                display_calc = field_node.find(".//DisplayCalculation")
                if display_calc is not None:
                    # Target field is also used
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(f"Field Display Calculation - {table_name}::{field_name} (Target)")
                    
                    # Check each Chunk element that references a field
                    for chunk in display_calc.findall(".//Chunk[@type='FieldRef']"):
                        field_ref = chunk.find("Field")
                        if field_ref is not None:
                            ref_field_name = field_ref.attrib.get("name", "")
                            ref_table_name = field_ref.attrib.get("table", "")
                            
                            if ref_field_name and ref_table_name:
                                base_table = table_occurrences_to_base.get(ref_table_name, ref_table_name)
                                if base_table in fields and ref_field_name in fields[base_table]:
                                    usage_desc = f"Field Display Calculation - {table_name}::{field_name}"
                                    ref_field_key = f"{base_table}::{ref_field_name}"
                                    field_usage[ref_field_key]['from_calculations'].append(usage_desc)
                
                # Check auto-enter calculation
                auto_enter = field_node.find(".//AutoEnter/Calculation")
                if auto_enter is not None and auto_enter.text:
                    calc_context = f"Auto-Enter Calculation - {table_name}::{field_name}"
                    
                    # Target field is also used
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(calc_context + " (Target)")
                    
                    find_field_references(
                        auto_enter.text, 
                        table_name,  # Has table context
                        calc_context, 
                        fields, 
                        field_usage, 
                        table_occurrences_to_base
                    )
                
                # Check validation calculation
                validation = field_node.find(".//Validation/Calculation")
                if validation is not None and validation.text:
                    calc_context = f"Validation Calculation - {table_name}::{field_name}"
                    
                    # Target field is also used
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(calc_context + " (Target)")
                    
                    find_field_references(
                        validation.text, 
                        table_name,  # Has table context
                        calc_context, 
                        fields, 
                        field_usage, 
                        table_occurrences_to_base
                    )
        
        # 4. Check for fields used in custom functions
        print("    Checking custom functions...")
        for cf_node in root.findall(".//CustomFunctionCatalog/CustomFunction"):
            cf_name = cf_node.attrib.get("name", "Unknown Function")
            
            calc_node = cf_node.find(".//Calculation")
            if calc_node is not None and calc_node.text:
                calc_context = f"Custom Function - {cf_name}"
                find_field_references(
                    calc_node.text, 
                    None,  # No specific table context, check all
                    calc_context, 
                    fields, 
                    field_usage, 
                    table_occurrences_to_base
                )
        
        # 5. Check for fields used in SQL statements
        print("    Checking SQL usage...")
        # Pattern to match ExecuteSQL calls
        sql_pattern = re.compile(r'ExecuteSQL\s*\(\s*["\']([^"\']*)["\']', re.IGNORECASE | re.DOTALL)
        
        # Function to process a calculation and extract SQL queries
        def process_sql_in_calculation(calc_text, context):
            if not calc_text or "ExecuteSQL" not in calc_text:
                return
            
            # Find all ExecuteSQL calls
            sql_queries = sql_pattern.findall(calc_text)
            
            for sql in sql_queries:
                # Extract tables from FROM clauses
                table_matches = re.findall(r'FROM\s+([a-zA-Z0-9_]+)', sql, re.IGNORECASE)
                
                # Process each table
                for table_name in table_matches:
                    # Check if this is a table occurrence
                    base_table = table_occurrences_to_base.get(table_name, table_name)
                    
                    if base_table not in fields:
                        continue
                    
                    # Look for fields from this table in SQL
                    for field_name in fields[base_table]:
                        # Look for table.field syntax
                        if f"{table_name}.{field_name}" in sql:
                            usage_desc = f"SQL - {context}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_sql'].append(usage_desc)
                        
                        # Look for direct field reference (WHERE field = value)
                        # This is less reliable but catches some cases
                        field_patterns = [
                            (r'WHERE\s+' + re.escape(field_name) + r'\s*[=<>!]', 'WHERE'),
                            (r'AND\s+' + re.escape(field_name) + r'\s*[=<>!]', 'AND'),
                            (r'OR\s+' + re.escape(field_name) + r'\s*[=<>!]', 'OR'),
                            (r'SELECT\s+' + re.escape(field_name) + r'\s*(?:,|\s+FROM)', 'SELECT'),
                            (r'ORDER\s+BY\s+' + re.escape(field_name), 'ORDER BY'),
                            (r'GROUP\s+BY\s+' + re.escape(field_name), 'GROUP BY'),
                        ]
                        
                        for pattern, context_type in field_patterns:
                            if re.search(pattern, sql, re.IGNORECASE):
                                usage_desc = f"SQL ({context_type}) - {context}"
                                field_key = f"{base_table}::{field_name}"
                                field_usage[field_key]['from_sql'].append(usage_desc)
        
        # Scan all calculations for SQL
        for calc in root.findall(".//Calculation"):
            if calc.text and "ExecuteSQL" in calc.text:
                # Determine context
                context = "Unknown"
                
                # Check for parent script
                parent = calc.getparent()
                for _ in range(10):  # Limit search depth
                    if parent is None:
                        break
                    
                    if parent.tag == "Script":
                        script_name = parent.attrib.get("name", "Unknown Script")
                        context = f"Script {script_name}"
                        break
                    elif parent.tag == "Step":
                        step_name = parent.attrib.get("name", "Unknown Step")
                        step_index = parent.attrib.get("index", "")
                        
                        # Look for parent script
                        script_parent = parent
                        for _ in range(5):
                            script_parent = script_parent.getparent()
                            if script_parent is None:
                                break
                            if script_parent.tag == "Script":
                                script_name = script_parent.attrib.get("name", "Unknown Script")
                                context = f"Script {script_name} (Step {step_index}: {step_name})"
                                break
                        break
                    elif parent.tag == "CustomFunction":
                        context = f"Custom Function {parent.attrib.get('name', 'Unknown Function')}"
                        break
                    elif parent.tag == "Field":
                        field_name = parent.attrib.get("name", "Unknown Field")
                        # Look for table
                        table_parent = parent
                        for _ in range(5):
                            table_parent = table_parent.getparent()
                            if table_parent is None:
                                break
                            if table_parent.tag == "BaseTable":
                                table_name = table_parent.attrib.get("name", "Unknown Table")
                                context = f"Field Calculation {table_name}::{field_name}"
                                break
                        break
                    elif parent.tag == "Object":
                        obj_type = parent.attrib.get("type", "Unknown Object")
                        # Look for layout
                        layout_parent = parent
                        for _ in range(5):
                            layout_parent = layout_parent.getparent()
                            if layout_parent is None:
                                break
                            if layout_parent.tag == "Layout":
                                layout_name = layout_parent.attrib.get("name", "Unknown Layout")
                                context = f"Layout Object - {layout_name} ({obj_type})"
                                break
                        break
                    
                    parent = parent.getparent()
                
                # Process SQL in this calculation
                process_sql_in_calculation(calc.text, context)
        
        # 6. Check for fields used in relationships
        print("    Checking relationships...")
        for relationship in root.findall(".//Relationship"):
            rel_name = relationship.attrib.get("name", "Unknown Relationship")
            
            # Check for fields in relationship definitions
            for field_pair in relationship.findall(".//FieldPair"):
                # First field in the pair
                field1 = field_pair.find(".//Field[1]")
                if field1 is not None:
                    field_name = field1.attrib.get("name", "")
                    table_name = field1.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        if base_table in fields and field_name in fields[base_table]:
                            usage_desc = f"Relationship - {rel_name}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_relationships'].append(usage_desc)
                
                # Second field in the pair
                field2 = field_pair.find(".//Field[2]")
                if field2 is not None:
                    field_name = field2.attrib.get("name", "")
                    table_name = field2.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        if base_table in fields and field_name in fields[base_table]:
                            usage_desc = f"Relationship - {rel_name}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_relationships'].append(usage_desc)
        
        # 7. Check for fields used in value lists
        print("    Checking value lists...")
        for vl in root.findall(".//ValueList"):
            vl_name = vl.attrib.get("name", "Unknown Value List")
            
            # Check for field-based value lists
            field_ref = vl.find(".//Field")
            if field_ref is not None:
                field_name = field_ref.attrib.get("name", "")
                table_name = field_ref.attrib.get("table", "")
                
                if field_name and table_name:
                    base_table = table_occurrences_to_base.get(table_name, table_name)
                    if base_table in fields and field_name in fields[base_table]:
                        usage_desc = f"Value List - {vl_name}"
                        field_key = f"{base_table}::{field_name}"
                        field_usage[field_key]['from_value_lists'].append(usage_desc)
            
            # Check for related value lists (which use fields)
            for related_set in vl.findall(".//RelatedSet"):
                # Primary field
                primary = related_set.find(".//PrimaryField")
                if primary is not None:
                    field_name = primary.attrib.get("name", "")
                    table_name = primary.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        if base_table in fields and field_name in fields[base_table]:
                            usage_desc = f"Value List (Primary) - {vl_name}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_value_lists'].append(usage_desc)
                
                # Display field
                display = related_set.find(".//DisplayField")
                if display is not None:
                    field_name = display.attrib.get("name", "")
                    table_name = display.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        if base_table in fields and field_name in fields[base_table]:
                            usage_desc = f"Value List (Display) - {vl_name}"
                            field_key = f"{base_table}::{field_name}"
                            field_usage[field_key]['from_value_lists'].append(usage_desc)
        
        # 8. NEW: Check for text objects with field references in merge format <<field_name>>
        print("    Checking text object merge fields...")
        all_text_objs = root.findall(".//CharacterStyleVector/Style/Data")
        for text_obj in all_text_objs:
            if text_obj.text:
                # Look for merge fields in format <<field_name>>
                merge_fields = re.findall(r'<<([^>]+)>>', text_obj.text)
                if merge_fields:
                    # Try to determine context (layout, position)
                    context = "Text Object"
                    parent = text_obj.getparent()
                    for _ in range(10):  # Limit search depth
                        if parent is None:
                            break
                        if parent.tag == "Object":
                            obj_type = parent.attrib.get("type", "Unknown")
                            # Try to find layout
                            layout_parent = parent
                            for _ in range(5):
                                layout_parent = layout_parent.getparent()
                                if layout_parent is None:
                                    break
                                if layout_parent.tag == "Layout":
                                    layout_name = layout_parent.attrib.get("name", "Unknown Layout")
                                    context = f"Text Merge Field - {layout_name} ({obj_type})"
                                    break
                            break
                        parent = parent.getparent()
                    
                    # For each merge field, check all tables for matching field
                    for field_name in merge_fields:
                        # Since we don't have table context, check all tables
                        for table_name, table_fields in fields.items():
                            if field_name in table_fields:
                                usage_desc = context
                                field_key = f"{table_name}::{field_name}"
                                field_usage[field_key]['from_layouts'].append(usage_desc)
        
        # Build results
        print("  Building results...")
        results = []
        
        # List of system fields to mark specially
        system_fields = [
            "CreatedBy",
            "DateCreated_c",
            "ModificationTimestamp",
            "CreationTimestamp",
            "ModifiedBy",
            "PrimaryKey",
            "Count_s"
        ]
        
        # First identify fields that should be marked as "Cached"
        # These are fields ending in "_c" with matching "*_cache" fields in the same table
        cached_fields = set()  # Will store table::field keys for cached fields
        
        # Find all fields ending in "_c" and their potential cache fields
        for table_name, table_fields in fields.items():
            c_fields = {}
            cache_fields = {}
            
            # Collect fields ending in "_c" and "*_cache"
            for field_name in table_fields:
                if field_name.endswith("_c"):
                    base_name = field_name[:-2]  # Remove the "_c"
                    c_fields[base_name] = field_name
                elif field_name.endswith("_cache"):
                    base_name = field_name[:-6]  # Remove the "_cache"
                    cache_fields[base_name] = field_name
            
            # Mark matching pairs as cached
            for base_name in c_fields:
                if base_name in cache_fields:
                    cached_fields.add(f"{table_name}::{c_fields[base_name]}")
                    cached_fields.add(f"{table_name}::{cache_fields[base_name]}")
        
        # Process all fields
        for table_name, table_fields in sorted(fields.items()):
            for field_name, field_info in sorted(table_fields.items()):
                field_key = f"{table_name}::{field_name}"
                usage = field_usage.get(field_key, {
                    'from_layouts': [],
                    'from_scripts': [],
                    'from_calculations': [],
                    'from_sql': [],
                    'from_portals': [],
                    'from_value_lists': [],
                    'from_relationships': [],
                    'from_web_viewers': [],
                    'from_other': []
                })
                
                # Calculate total usage
                total_usage = sum(len(usage[key]) for key in usage)
                
                # Format usage lists
                def format_usage_list(usage_list, max_items=3):
                    if not usage_list:
                        return ""
                    
                    unique_items = list(dict.fromkeys(usage_list))  # Remove duplicates while preserving order
                    
                    if len(unique_items) <= max_items:
                        return "\n".join(unique_items)
                    else:
                        shown_items = unique_items[:max_items]
                        remaining = len(unique_items) - max_items
                        shown_items.append(f"...and {remaining} more")
                        return "\n".join(shown_items)
                
                # Combine other sources
                other_sources = (
                    usage['from_portals'] + 
                    usage['from_value_lists'] + 
                    usage['from_relationships'] + 
                    usage['from_web_viewers'] + 
                    usage['from_other']
                )
                
                # Determine status with special handling for various field types
                if field_name.startswith("#"):
                    status = "Comment"
                    status_sort = 2.5  # Between System and Imported
                elif field_key in cached_fields:
                    status = "Cached"
                    status_sort = 1  # After "Not Used", before "System"
                elif field_name in system_fields:
                    status = "System"
                    status_sort = 2  # After "Cached", before "Comment"
                # Check if the field is from an import table (table name has "Import" in it)
                elif "Import" in table_name or "import" in table_name or table_name.startswith("imp_"):
                    status = "Imported"
                    status_sort = 2.75  # Between Comment and Used
                elif total_usage == 0:
                    status = "Not Used"
                    status_sort = 0  # First in sort order
                else:
                    status = "Used"
                    status_sort = 3  # Last in sort order
                
                # Get XML count
                xml_count = xml_counts.get(field_key, 0)
                
                # Add the result
                results.append({
                    "Table Name": table_name,
                    "Field Name": field_name,
                    "Status": status,
                    "Usage Count": total_usage,
                    "XML Count": xml_count,
                    "Used in Layouts": format_usage_list(usage['from_layouts']),
                    "Used in Scripts": format_usage_list(usage['from_scripts']),
                    "Used in Calculations": format_usage_list(usage['from_calculations']),
                    "Used in SQL": format_usage_list(usage['from_sql']),
                    "Used in Other": format_usage_list(other_sources),
                    "_sort_key": (
                        status_sort,     # Status priority (0=Not Used, 1=System/Cached, 2=Used)
                        table_name.lower(),  # Then by table
                        field_name.lower()   # Then by field
                    )
                })
        
        # Sort by status, table name, field name
        results.sort(key=lambda x: x['_sort_key'])
        
        # Remove sort key before returning
        for result in results:
            del result['_sort_key']
        
        # Apply column ordering to ensure consistent output
        ordered_columns = [
            "Table Name",
            "Field Name", 
            "Status",
            "Usage Count",
            "XML Count",
            "Used in Layouts",
            "Used in Scripts",
            "Used in Calculations", 
            "Used in SQL",
            "Used in Other"
        ]
        
        # Create a new list with columns in the desired order
        ordered_results = []
        for result in results:
            ordered_result = {}
            for column in ordered_columns:
                if column in result:
                    ordered_result[column] = result[column]
                else:
                    # Ensure all columns are present even if missing in the original
                    if column == "XML Count":
                        ordered_result[column] = xml_counts.get(f"{result['Table Name']}::{result['Field Name']}", 0)
                    else:
                        ordered_result[column] = ""  # Default to empty string
            ordered_results.append(ordered_result)
        
        print(f"Final result has {len(ordered_results)} rows")
        if ordered_results:
            print(f"Columns in first row: {list(ordered_results[0].keys())}")
        
        return ordered_results
        
    except Exception as e:
        print(f"Error in UnusedFieldsCheck: {e}")
        import traceback
        traceback.print_exc()
        return []


def find_field_references(calc_text, table_context, usage_context, fields, field_usage, table_occurrences_to_base):
    """
    Find field references in calculation text.
    
    Args:
        calc_text: The calculation text to search
        table_context: The current table context (or None if global)
        usage_context: Description of where the calculation is used
        fields: Dictionary of all fields by table
        field_usage: Dictionary to update with usage info
        table_occurrences_to_base: Mapping of table occurrences to base tables
    """
    if not calc_text:
        return
    
    # 1. Look for direct references with table::field syntax
    # This is the most reliable method - handle emojis and special chars
    table_field_pattern = re.compile(r'([a-zA-Z0-9_🌎🧑‍🎓_🔗👥]+)::([a-zA-Z0-9_]+)', re.IGNORECASE)
    matches = table_field_pattern.findall(calc_text)
    
    for table_name, field_name in matches:
        # Check if this is a table occurrence
        base_table = table_occurrences_to_base.get(table_name, table_name)
        
        if base_table in fields and field_name in fields[base_table]:
            field_key = f"{base_table}::{field_name}"
            field_usage[field_key]['from_calculations'].append(usage_context)
        elif table_name in fields and field_name in fields[table_name]:
            # Try direct table match if base table mapping fails
            field_key = f"{table_name}::{field_name}"
            field_usage[field_key]['from_calculations'].append(usage_context)
    
    # 2. If we have a table context, look for unqualified field references
    if table_context:
        # Look for all fields from this table that might be referenced directly
        for field_name in fields.get(table_context, {}):
            # Check if the field name appears as a word (not part of another word)
            # This is a bit less reliable but catches most cases
            pattern = r'\b' + re.escape(field_name) + r'\b'
            if re.search(pattern, calc_text):
                field_key = f"{table_context}::{field_name}"
                field_usage[field_key]['from_calculations'].append(usage_context)
    
    # 3. Look for table.field syntax (common in SQL or some calculations)
    table_dot_field_pattern = re.compile(r'([a-zA-Z0-9_🌎🧑‍🎓_🔗👥]+)\.([a-zA-Z0-9_]+)', re.IGNORECASE)
    dot_matches = table_dot_field_pattern.findall(calc_text)
    
    for table_name, field_name in dot_matches:
        # Skip common functions that use dot notation
        if table_name.lower() in ['get', 'let', 'set', 'abs', 'sin', 'cos', 'tan', 'exp', 'log', 'min', 'max']:
            continue
            
        # Check if this is a table occurrence
        base_table = table_occurrences_to_base.get(table_name, table_name)
        
        if base_table in fields and field_name in fields[base_table]:
            field_key = f"{base_table}::{field_name}"
            field_usage[field_key]['from_calculations'].append(usage_context)
        elif table_name in fields and field_name in fields[table_name]:
            # Try direct table match if base table mapping fails
            field_key = f"{table_name}::{field_name}"
            field_usage[field_key]['from_calculations'].append(usage_context)
            
    # 4. Look for special cases like globals with emoji
    # Handle special case for global variables - they might not be detected with standard patterns
    if "Globals" in calc_text or "🌎Globals" in calc_text:
        # Use a more permissive pattern for globals
        globals_pattern = re.compile(r'(?:🌎Globals|Globals)::([a-zA-Z0-9_]+)', re.IGNORECASE)
        global_matches = globals_pattern.findall(calc_text)
        
        for field_name in global_matches:
            # Try different versions of Globals table name
            for global_table in ["Globals", "🌎Globals"]:
                if global_table in fields and field_name in fields[global_table]:
                    field_key = f"{global_table}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(usage_context)
                    break
                    
    # 5. Look for Chunk/Field references in DisplayCalculation elements
    # These are often in XML but not in the calculation text
    chunk_pattern = re.compile(r'<Chunk[^>]*>.*?<Field[^>]*table="([^"]*)"[^>]*name="([^"]*)"', re.DOTALL)
    chunk_matches = chunk_pattern.findall(calc_text)
    
    for table_name, field_name in chunk_matches:
        base_table = table_occurrences_to_base.get(table_name, table_name)
        
        if base_table in fields and field_name in fields[base_table]:
            field_key = f"{base_table}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (DisplayCalculation)")
        elif table_name in fields and field_name in fields[table_name]:
            field_key = f"{table_name}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (DisplayCalculation)")
                    
    # 6. Special case for Field ID references
    # FileMaker sometimes uses field IDs in calculations instead of names
    id_pattern = re.compile(r'<Field [^>]*?id="(\d+)"', re.DOTALL)
    id_matches = id_pattern.findall(calc_text)
    
    if id_matches:
        # We need to check all fields to find matching IDs
        for table_name, table_fields in fields.items():
            for field_name, field_info in table_fields.items():
                if isinstance(field_info, dict) and 'id' in field_info:
                    field_id = field_info['id']
                    if field_id in id_matches:
                        field_key = f"{table_name}::{field_name}"
                        field_usage[field_key]['from_calculations'].append(f"{usage_context} (Field ID)")
    
    # 7. Look for fields in placeholders (common in find mode)
    if "PlaceholderText" in calc_text or "findMode" in calc_text:
        # This is likely a placeholder calculation for find mode
        # Check all fields since placeholders often reference globals
        for table_name, table_fields in fields.items():
            for field_name in table_fields:
                if field_name in calc_text:
                    # This is a loose match, but useful for placeholders
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(f"{usage_context} (PlaceholderText)")
    
    # 8. Check for FieldRef nodes which explicitly reference fields
    # <Chunk type="FieldRef"><Field table="..." name="..."></Field></Chunk>
    field_ref_pattern = re.compile(r'<Chunk type="FieldRef">.*?<Field [^>]*?table="([^"]*)"[^>]*?name="([^"]*)"', re.DOTALL)
    field_ref_matches = field_ref_pattern.findall(calc_text)
    
    for table_name, field_name in field_ref_matches:
        base_table = table_occurrences_to_base.get(table_name, table_name)
        
        if base_table in fields and field_name in fields[base_table]:
            field_key = f"{base_table}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (FieldRef)")
        elif table_name in fields and field_name in fields[table_name]:
            field_key = f"{table_name}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (FieldRef)")
    
    # 9. Handle field references in XML attributes
    # Fields can be referenced in various attributes like table="..." name="..."
    xml_field_pattern = re.compile(r'<Field [^>]*?table="([^"]*)"[^>]*?name="([^"]*)"', re.DOTALL)
    xml_field_matches = xml_field_pattern.findall(calc_text)
    
    for table_name, field_name in xml_field_matches:
        base_table = table_occurrences_to_base.get(table_name, table_name)
        
        if base_table in fields and field_name in fields[base_table]:
            field_key = f"{base_table}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (XML Reference)")
        elif table_name in fields and field_name in fields[table_name]:
            field_key = f"{table_name}::{field_name}"
            field_usage[field_key]['from_calculations'].append(f"{usage_context} (XML Reference)")
            
    # 10. Check for merge field references <<field_name>>
    merge_field_pattern = re.compile(r'<<([^>]+)>>')
    merge_matches = merge_field_pattern.findall(calc_text)
    
    if merge_matches:
        for field_name in merge_matches:
            # Since we don't have table context for these, check all tables
            for table_name, table_fields in fields.items():
                if field_name in table_fields:
                    field_key = f"{table_name}::{field_name}"
                    field_usage[field_key]['from_calculations'].append(f"{usage_context} (Merge Field)")
