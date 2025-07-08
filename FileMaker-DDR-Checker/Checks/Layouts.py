from lxml import etree as ET
import re
from collections import defaultdict

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Layout Usage"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 1  # Make this the first sheet

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Layout Name": 250,
        "Path": 300,
        "Total Usage": 80,
        "Status": 85,
        "Used in Scripts": 250,
        "Used in Buttons": 250,
        "Used in Triggers": 250,
        "Used in Other": 200
    }

def apply_styling(ws):
    """Apply custom styling to the Layout Usage worksheet"""
    from openpyxl.styles import PatternFill, Font
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Get path to check for special folders
        path_value = ""
        if "Path" in columns:
            path_cell = row[columns["Path"] - 1]
            path_value = str(path_cell.value or "").lower()
        
        # Check status for unused layouts
        if "Status" in columns:
            status_cell = row[columns["Status"] - 1]
            status_value = str(status_cell.value or "")
            
            # Check if path contains "delete" or "debug"
            if "delete" in path_value or "debug" in path_value:
                # Apply warning styling for layouts in delete/debug folders
                for cell in row:
                    cell.fill = get_color_fill('warning')
                
                # Update status to reflect special folder
                if "NOT USED" in status_value:
                    if "delete" in path_value:
                        status_cell.value = "⚠️ In Delete Folder"
                    elif "debug" in path_value:
                        status_cell.value = "⚠️ In Debug Folder"
                else:
                    if "delete" in path_value:
                        status_cell.value = "In Delete Folder"
                    elif "debug" in path_value:
                        status_cell.value = "In Debug Folder"
                
                status_cell.font = get_color_font('warning', bold=True)
                
                # Also make layout name bold
                if "Layout Name" in columns:
                    name_cell = row[columns["Layout Name"] - 1]
                    name_cell.font = get_color_font('warning', bold=True)
                    
            elif "NOT USED" in status_value:
                # Standard error styling for unused layouts not in special folders
                for cell in row:
                    cell.fill = get_color_fill('error')
                
                status_cell.font = get_color_font('error', bold=True)
                
                # Also make layout name bold
                if "Layout Name" in columns:
                    name_cell = row[columns["Layout Name"] - 1]
                    name_cell.font = get_color_font('error', bold=True)
        
        # Color code total usage count
        if "Total Usage" in columns:
            count_cell = row[columns["Total Usage"] - 1]
            try:
                count = int(count_cell.value or 0)
                if count == 0:
                    count_cell.font = get_color_font('error', bold=True)
                elif count >= 10:
                    count_cell.font = get_color_font('success', bold=True)
                elif count >= 5:
                    count_cell.font = Font(bold=True, color="388E3C")  # Medium green
            except:
                pass

def run_check(raw_xml):
    """
    Find all layouts and track where they are referenced.
    Flag layouts that aren't referenced anywhere.
    Sort so unused layouts appear at the top.
    """
    try:
        # Parse the XML
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        root = ET.fromstring(raw_xml.encode('utf-8'), parser)
        
        # Dictionary to store all layouts
        layouts = {}
        
        # Dictionary to store layout usage by type
        layout_usage = defaultdict(lambda: {
            'from_scripts': [],
            'from_buttons': [],
            'from_triggers': [],
            'from_relationships': [],
            'from_value_lists': [],
            'from_other': []
        })
        
        # First, build the layout path mapping from LayoutCatalog
        layout_paths = {}
        group_names = set()  # Track all group/folder names
        
        # Find the LayoutCatalog
        layout_catalog = root.find(".//LayoutCatalog")
        if layout_catalog is not None:
            def process_catalog_element(elem, current_path=""):
                """Process elements in the LayoutCatalog to build paths"""
                if elem.tag == "Group":
                    # This is a folder
                    group_name = elem.attrib.get("name", "")
                    if group_name:
                        group_names.add(group_name)  # Track folder names
                        new_path = f"{current_path} > {group_name}" if current_path else group_name
                        # Process children with the new path
                        for child in elem:
                            process_catalog_element(child, new_path)
                elif elem.tag == "Layout":
                    # This is a layout reference
                    layout_id = elem.attrib.get("id", "")
                    if layout_id:
                        # Store the path for this layout ID
                        layout_paths[layout_id] = current_path if current_path else "Top Level"
                else:
                    # Process any other elements
                    for child in elem:
                        process_catalog_element(child, current_path)
            
            # Process all elements in the catalog
            for child in layout_catalog:
                process_catalog_element(child)
        
        print(f"Debug: Found {len(group_names)} folder names: {sorted(group_names)}")
        
        # Now find all actual Layout definitions
        # The actual layout definitions are Layout elements that are NOT inside LayoutCatalog
        # and are not just references
        
        all_layouts = root.findall(".//Layout")
        print(f"Debug: Found {len(all_layouts)} total Layout elements in document")
        
        for layout in all_layouts:
            layout_name = layout.attrib.get("name", "Unknown Layout")
            layout_id = layout.attrib.get("id", "")
            
            # Skip if the layout name is actually a folder name
            if layout_name in group_names:
                print(f"Skipping '{layout_name}' - it's a folder name")
                continue
            
            # Check if this layout is inside LayoutCatalog by walking up the tree
            is_in_catalog = False
            current = layout
            for _ in range(20):  # Limit depth to prevent infinite loops
                parent = current.getparent()
                if parent is None:
                    break
                if parent.tag == "LayoutCatalog":
                    is_in_catalog = True
                    break
                current = parent
            
            if is_in_catalog:
                continue  # Skip catalog references
            
            # Also skip if parent is a Step or other reference context
            immediate_parent = layout.getparent()
            if immediate_parent is not None and immediate_parent.tag in ["Step", "Parameter", "ButtonObj", "Relationship"]:
                continue
            
            # This should be an actual layout definition
            # Get the path from our mapping, default to "Top Level"
            path = layout_paths.get(layout_id, "Top Level")
            
            layouts[layout_name] = {
                "id": layout_id,
                "name": layout_name,
                "path": path
            }
        
        print(f"Found {len(layouts)} layouts")
        
        # Debug: Check if any layout names match folder names
        if layout_paths:
            folder_names = set()
            for path in layout_paths.values():
                if path and path != "Top Level":
                    # Extract folder names from paths
                    parts = path.split(" > ")
                    folder_names.update(parts)
            
            # Check for layouts that have the same name as folders
            layout_names = set(layouts.keys())
            overlapping_names = layout_names.intersection(folder_names)
            if overlapping_names:
                print(f"Warning: Found {len(overlapping_names)} names that are both layouts and folders:")
                for name in sorted(overlapping_names):
                    print(f"  - {name}")
        
        # Count layouts by path for debugging
        path_counts = defaultdict(int)
        for layout_info in layouts.values():
            path_counts[layout_info["path"]] += 1
        
        print("\nLayouts by path:")
        for path, count in sorted(path_counts.items()):
            print(f"  {path}: {count} layouts")
        
        # 1. Check for layouts referenced in scripts (Go to Layout steps)
        for script in root.findall(".//Script"):
            script_name = script.attrib.get("name", "Unknown Script")
            
            # Look for Go to Layout steps
            for step in script.findall(".//Step[@name='Go to Layout']"):
                # Find the layout reference
                layout_ref = step.find(".//Layout")
                if layout_ref is not None:
                    referenced_layout = layout_ref.attrib.get("name")
                    if referenced_layout and referenced_layout in layouts:
                        location = f"{script_name}"
                        layout_usage[referenced_layout]['from_scripts'].append(location)
                
                # Also check for layout references in parameters
                for param in step.findall(".//Parameter"):
                    layout_name_attr = param.attrib.get("layout")
                    if layout_name_attr and layout_name_attr in layouts:
                        location = f"{script_name}"
                        layout_usage[layout_name_attr]['from_scripts'].append(location)
            
            # Also check for other navigation steps that might reference layouts
            navigation_steps = [
                "Go to Related Record",
                "New Window",
                "Select Window"
            ]
            
            for step_name in navigation_steps:
                for step in script.findall(f".//Step[@name='{step_name}']"):
                    layout_ref = step.find(".//Layout")
                    if layout_ref is not None:
                        referenced_layout = layout_ref.attrib.get("name")
                        if referenced_layout and referenced_layout in layouts:
                            location = f"{script_name} ({step_name})"
                            layout_usage[referenced_layout]['from_scripts'].append(location)
        
        # 2. Check for layouts referenced in buttons
        for layout in root.findall(".//Layout"):
            source_layout_name = layout.attrib.get("name", "Unknown Layout")
            
            for obj in layout.findall(".//Object"):
                obj_type = obj.attrib.get("type", "Unknown")
                obj_name = obj.attrib.get("name", "")
                
                # Check for script steps in objects that reference layouts
                for step in obj.findall(".//Step"):
                    if step.attrib.get("name") == "Go to Layout":
                        layout_ref = step.find(".//Layout")
                        if layout_ref is not None:
                            referenced_layout = layout_ref.attrib.get("name")
                            if referenced_layout and referenced_layout in layouts:
                                button_desc = f"{source_layout_name}"
                                if obj_name:
                                    button_desc += f" - {obj_name}"
                                else:
                                    button_desc += f" - {obj_type}"
                                layout_usage[referenced_layout]['from_buttons'].append(button_desc)
                
                # Check SingleStep for button actions
                single_step = obj.find(".//SingleStep")
                if single_step is not None:
                    step = single_step.find(".//Step[@name='Go to Layout']")
                    if step is not None:
                        layout_ref = step.find(".//Layout")
                        if layout_ref is not None:
                            referenced_layout = layout_ref.attrib.get("name")
                            if referenced_layout and referenced_layout in layouts:
                                button_desc = f"{source_layout_name}"
                                if obj_name:
                                    button_desc += f" - {obj_name}"
                                else:
                                    button_desc += f" - {obj_type}"
                                layout_usage[referenced_layout]['from_buttons'].append(button_desc)
        
        # 3. Check for layout triggers that might go to other layouts
        for layout in root.findall(".//Layout"):
            source_layout_name = layout.attrib.get("name", "Unknown Layout")
            
            # Check layout-level triggers
            trigger_types = [
                "OnRecordLoad", "OnRecordCommit", "OnRecordRevert",
                "OnLayoutEnter", "OnLayoutExit", "OnLayoutKeystroke",
                "OnModeEnter", "OnModeExit", "OnViewChange"
            ]
            
            for trigger_type in trigger_types:
                trigger = layout.find(f".//{trigger_type}")
                if trigger is not None:
                    # Check if the trigger script goes to a layout
                    for step in trigger.findall(".//Step[@name='Go to Layout']"):
                        layout_ref = step.find(".//Layout")
                        if layout_ref is not None:
                            referenced_layout = layout_ref.attrib.get("name")
                            if referenced_layout and referenced_layout in layouts:
                                trigger_desc = f"{source_layout_name} - {trigger_type}"
                                layout_usage[referenced_layout]['from_triggers'].append(trigger_desc)
        
        # 4. Check for layouts referenced in relationships (Go to Related Record)
        for relationship in root.findall(".//Relationship"):
            rel_name = relationship.attrib.get("name", "Unknown Relationship")
            
            # Check if relationship has a layout reference
            layout_ref = relationship.find(".//Layout")
            if layout_ref is not None:
                referenced_layout = layout_ref.attrib.get("name")
                if referenced_layout and referenced_layout in layouts:
                    layout_usage[referenced_layout]['from_relationships'].append(rel_name)
        
        # 5. Check for layouts in value lists (related values)
        for vl in root.findall(".//ValueList"):
            vl_name = vl.attrib.get("name", "Unknown Value List")
            
            # Check if value list references a layout
            layout_ref = vl.find(".//Layout")
            if layout_ref is not None:
                referenced_layout = layout_ref.attrib.get("name")
                if referenced_layout and referenced_layout in layouts:
                    layout_usage[referenced_layout]['from_value_lists'].append(vl_name)
        
        # 6. Check for default/startup layouts in File Options
        for file_option in root.findall(".//FileOptions"):
            # Default layout
            default_layout = file_option.find(".//DefaultLayout")
            if default_layout is not None:
                layout_name = default_layout.attrib.get("name")
                if layout_name and layout_name in layouts:
                    layout_usage[layout_name]['from_other'].append("File Options - Default Layout")
        
        # 7. Check for layouts in custom menus
        for menu_set in root.findall(".//CustomMenuSet"):
            menu_set_name = menu_set.attrib.get("name", "Unknown Menu Set")
            
            for menu in menu_set.findall(".//CustomMenu"):
                menu_name = menu.attrib.get("name", "Unknown Menu")
                
                for menu_item in menu.findall(".//CustomMenuItem"):
                    item_name = menu_item.attrib.get("name", "Unknown Item")
                    
                    # Check for Go to Layout in menu item
                    for step in menu_item.findall(".//Step[@name='Go to Layout']"):
                        layout_ref = step.find(".//Layout")
                        if layout_ref is not None:
                            referenced_layout = layout_ref.attrib.get("name")
                            if referenced_layout and referenced_layout in layouts:
                                menu_desc = f"{menu_set_name} > {menu_name} > {item_name}"
                                layout_usage[referenced_layout]['from_other'].append(menu_desc)
        
        # Build results
        results = []
        all_layout_results = []
        
        # Process all layouts
        for layout_name in layouts:
            usage = layout_usage.get(layout_name, {
                'from_scripts': [],
                'from_buttons': [],
                'from_triggers': [],
                'from_relationships': [],
                'from_value_lists': [],
                'from_other': []
            })
            
            total_usage = sum(len(usage[key]) for key in usage)
            layout_info = layouts[layout_name]
            
            # Format usage lists
            def format_usage_list(usage_list, max_items=3):
                if not usage_list:
                    return ""
                
                unique_items = list(dict.fromkeys(usage_list))  # Remove duplicates while preserving order
                
                if len(unique_items) <= max_items:
                    return "\n".join(unique_items)
                else:
                    shown_items = unique_items[:max_items]
                    remaining = len(unique_items) - max_items
                    shown_items.append(f"...and {remaining} more")
                    return "\n".join(shown_items)
            
            # Combine other sources
            other_sources = usage['from_relationships'] + usage['from_value_lists'] + usage['from_other']
            
            # Determine status based on usage and path
            path_lower = layout_info.get('path', '').lower()
            
            if total_usage == 0:
                if "delete" in path_lower:
                    status = "⚠️ In Delete Folder"
                elif "debug" in path_lower:
                    status = "⚠️ In Debug Folder"
                else:
                    status = "⚠️ NOT USED"
            else:
                if "delete" in path_lower:
                    status = "In Delete Folder"
                elif "debug" in path_lower:
                    status = "In Debug Folder"
                else:
                    status = "Active"
            
            all_layout_results.append({
                "Layout Name": layout_name,
                "Path": layout_info.get('path', 'Top Level'),  # Use get() with default
                "Total Usage": total_usage,
                "Status": status,
                "Used in Scripts": format_usage_list(usage['from_scripts']),
                "Used in Buttons": format_usage_list(usage['from_buttons']),
                "Used in Triggers": format_usage_list(usage['from_triggers']),
                "Used in Other": format_usage_list(other_sources),
                "_sort_key": (
                    # Sort by status priority
                    0 if "NOT USED" in status else  # NOT USED first (red)
                    1 if "Debug" in status else      # Debug second (yellow)
                    2 if "Delete" in status else     # Delete third (yellow)
                    3,                               # Active last (no color)
                    layout_info.get('path', '').lower(),  # Then by path
                    layout_name.lower()  # Then by name
                )
            })
        
        # Sort by status (errors first), then by path, then by name
        all_layout_results.sort(key=lambda x: x['_sort_key'])
        
        # Remove sort key before returning
        for result in all_layout_results:
            del result['_sort_key']
            results.append(result)
        
        return results
        
    except Exception as e:
        print(f"Error in UnusedLayoutsCheck: {e}")
        import traceback
        traceback.print_exc()
        return []