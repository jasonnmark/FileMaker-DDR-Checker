from lxml import etree as ET
import re
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Color

# Import the color functions from the common module
try:
    from common_styles import get_color_fill, get_color_font
except ImportError:
    # Define fallback functions if common_styles is not available
    def get_color_fill(color_type):
        color_map = {
            'category_script': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            'category_custom_function': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            'category_field_calc': PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
            'category_layout_object': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
            'category_other': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
            'warning': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            'error': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        }
        return color_map.get(color_type, PatternFill())
    
    def get_color_font(color_type, bold=False):
        color_map = {
            'warning': Font(color="F57C00", bold=bold),
            'error': Font(color="D32F2F", bold=bold),
            'muted': Font(color="757575", bold=bold)
        }
        if color_type in color_map:
            return color_map[color_type]
        return Font(color="000000", bold=bold)

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Unknown References"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 7  # Place after SQL Usage

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Status": 100,
        "Type": 100,
        "Context": 150,       # Renamed from Table/Layout to Context
        "Location": 400,
        "Commented": 50,      # New narrow column
        "Error": 300,
        "XML Line": 50,
        "Details": 200
    }

def apply_styling(ws):
    """Apply custom styling to the Unknown References worksheet"""
    
    # Define type colors (same as SQL.py category colors)
    type_colors = {
        "Script": get_color_fill('category_script'),
        "Custom Function": get_color_fill('category_custom_function'),
        "Field Calc": get_color_fill('category_field_calc'),
        "Layout Object": get_color_fill('category_layout_object'),
        "Layout Field": get_color_fill('category_layout_object'),
        "Value List": get_color_fill('category_other'),
        "Relationship": get_color_fill('category_other'),
        "Other": get_color_fill('category_other')
    }
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Check status for coloring
        if "Status" in columns:
            status_cell = row[columns["Status"] - 1]
            status_value = status_cell.value if status_cell.value else ""
            
            if status_value in ["To Delete", "Commented Out", "Temp/Debug"]:
                # Apply warning (yellow) styling to entire row
                for cell in row:
                    cell.fill = get_color_fill('warning')
                
                status_cell.font = get_color_font('warning', bold=True)
            else:
                # Error (red) styling for active errors
                status_cell.font = get_color_font('error', bold=True)
        
        # Get type value
        if "Type" in columns:
            type_cell = row[columns["Type"] - 1]
            type_value = type_cell.value if type_cell.value else ""
            
            # Apply type coloring only if not already yellow
            status_value = row[columns.get("Status", 0) - 1].value if "Status" in columns else ""
            if status_value not in ["To Delete", "Commented Out", "Temp/Debug"] and type_value in type_colors:
                type_cell.fill = type_colors[type_value]
            
            # Make type cell bold
            type_cell.font = Font(bold=True)
        
        # Error column - show in error color only if not yellow status
        if "Error" in columns:
            error_cell = row[columns["Error"] - 1]
            if error_cell.value:
                status_value = row[columns.get("Status", 0) - 1].value if "Status" in columns else ""
                if status_value not in ["To Delete", "Commented Out", "Temp/Debug"]:
                    error_cell.fill = get_color_fill('error')
                    error_cell.font = get_color_font('error', bold=True)
                else:
                    error_cell.font = get_color_font('warning', bold=True)
        
        # Location column - style based on Commented status
        if "Location" in columns and "Commented" in columns:
            location_cell = row[columns["Location"] - 1]
            commented_cell = row[columns["Commented"] - 1]
            
            # If commented out, make location text grey
            if commented_cell.value == "Yes":
                location_cell.font = get_color_font('muted')
            else:
                # Black text for non-commented locations
                location_cell.font = Font(color="000000")
        
        # Context column - normal styling
        if "Context" in columns:
            context_cell = row[columns["Context"] - 1]
            # Apply standard font
            context_cell.font = Font(color="000000")

def run_check(raw_xml, catalogs=None):
    """
    Find all references to unknown/deleted items in the DDR.
    This includes unknown scripts, tables, fields, layouts, etc.
    Now uses shared catalogs for consistency.
    """
    
    # Use shared catalogs if provided (new style)
    if catalogs:
        known_scripts = set(catalogs['scripts'].keys())
        known_layouts = set(catalogs['layouts'].keys())
        known_tables = set(catalogs['tables'].keys())
        known_fields = catalogs['fields_by_table']
        known_custom_functions = set(catalogs['custom_functions'].keys())
        known_value_lists = set(catalogs['value_lists'].keys())
        known_relationships = set(catalogs['relationships'].keys())
        table_occurrences_to_base = catalogs['table_occurrences']
        script_paths = catalogs['script_paths']
        layout_paths = catalogs['layout_paths']
        root = catalogs['root']
    else:
        # If no catalogs provided, we can't run this check
        print("Warning: Unknown References check requires shared catalogs")
        return []
    
    def determine_status(location, error, calc_text=None, is_script_step=False, step_elem=None):
        """Determine the status of an unknown reference"""
        location_lower = location.lower()
        error_lower = error.lower()
        
        # Check if in "to delete" folder
        if "to delete" in location_lower or "todelete" in location_lower or ">delete" in location_lower:
            return "To Delete"
        
        # Check if script or layout path contains "to delete"
        if "Script:" in location:
            script_name = location.split("Script: ")[1].split(",")[0]
            if script_name in script_paths:
                path = script_paths[script_name].lower()
                if "to delete" in path or "todelete" in path or ">delete" in path:
                    return "To Delete"
        
        if "Layout:" in location:
            layout_name = location.split("Layout: ")[1].split(",")[0].split(" Top:")[0].strip()
            if layout_name in layout_paths:
                path = layout_paths[layout_name].lower()
                if "to delete" in path or "todelete" in path or ">delete" in path:
                    return "To Delete"
        
        # Check if contains "temp" or "debug"
        if "temp" in location_lower or "debug" in location_lower:
            return "Temp/Debug"
        if "temp" in error_lower or "debug" in error_lower:
            return "Temp/Debug"
        
        # Check if script step is commented out
        if is_script_step and step_elem is not None:
            # Check if step is disabled
            if step_elem.attrib.get("enable", "True") == "False":
                return "Commented Out"
        
        # Check if field calculation is commented out
        if calc_text:
            # Simple check for comment patterns
            if calc_text.strip().startswith("/*") or calc_text.strip().startswith("//"):
                return "Commented Out"
            # Check if entire calc is wrapped in /* */
            stripped = calc_text.strip()
            if stripped.startswith("/*") and stripped.endswith("*/"):
                return "Commented Out"
        
        # Default to active error
        return "Active Error"
    
    def is_commented(status, is_script_step=False, step_elem=None, calc_text=None):
        """Check if the item is commented out"""
        # First check if status is already "Commented Out"
        if status == "Commented Out":
            return True
        
        # For script steps, check if disabled
        if is_script_step and step_elem is not None:
            if step_elem.attrib.get("enable", "True") == "False":
                return True
        
        # For calculations, check for comment patterns
        if calc_text:
            stripped = calc_text.strip()
            if stripped.startswith("/*") or stripped.startswith("//"):
                return True
            if stripped.startswith("/*") and stripped.endswith("*/"):
                return True
        
        return False
    
    try:
        results = []
        
        # Now search for unknown references
        
        # Check script steps for unknown scripts, layouts, and fields
        for script in root.findall(".//Script"):
            script_name = script.attrib.get("name", "Unknown Script")
            
            for step in script.findall(".//Step"):
                step_name = step.attrib.get("name", "")
                step_index = step.attrib.get("index", "")
                step_id = step.attrib.get("id", "")
                
                # Use step index for display (this is the step number)
                step_number = step_index if step_index else step_id
                
                # Check Perform Script steps
                if step_name == "Perform Script":
                    # Look for script reference
                    script_ref = step.find(".//Script")
                    if script_ref is not None:
                        ref_script_name = script_ref.attrib.get("name", "")
                        if ref_script_name and ref_script_name not in known_scripts:
                            xml_line = getattr(script_ref, 'sourceline', 'unknown')
                            location = f'Step {step_number}: {step_name}'
                            error = f'Unknown script: "{ref_script_name}"'
                            status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                            
                            results.append({
                                "Status": status,
                                "Type": "Script",
                                "Context": script_name,
                                "Location": location,
                                "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Referenced script not found"
                            })
                
                # Check Go to Layout steps
                if step_name == "Go to Layout":
                    layout_ref = step.find(".//Layout")
                    if layout_ref is not None:
                        ref_layout_name = layout_ref.attrib.get("name", "")
                        if ref_layout_name and ref_layout_name not in known_layouts:
                            xml_line = getattr(layout_ref, 'sourceline', 'unknown')
                            location = f'Step {step_number}: {step_name}'
                            error = f'Unknown layout: "{ref_layout_name}"'
                            status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                            
                            results.append({
                                "Status": status,
                                "Type": "Script",
                                "Context": script_name,
                                "Location": location,
                                "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Referenced layout not found"
                            })
                
                # Check all Field references in any script step
                # Skip SQL-related steps since SQL errors are handled by a separate check
                if step_name not in ["Execute SQL"]:
                    for field_ref in step.findall(".//Field"):
                        field_name = field_ref.attrib.get("name", "")
                        table_name = field_ref.attrib.get("table", "")
                        
                        if field_name and table_name:
                            # Get base table
                            base_table = table_occurrences_to_base.get(table_name, table_name)
                            
                            # Check if table exists
                            if base_table not in known_tables:
                                xml_line = getattr(field_ref, 'sourceline', 'unknown')
                                location = f'Step {step_number}: {step_name}'
                                error = f'Unknown table: "{table_name}"'
                                status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Script",
                                    "Context": script_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Table not found"
                                })
                            # Check if field exists in the table
                            elif field_name not in known_fields.get(base_table, set()):
                                xml_line = getattr(field_ref, 'sourceline', 'unknown')
                                location = f'Step {step_number}: {step_name}'
                                error = f'Unknown field: "{table_name}::{field_name}"'
                                status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Script",
                                    "Context": script_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Field not found in table"
                                })
                
                # Check calculations in script steps (like Set Field, Set Variable, etc.)
                # Skip ExecuteSQL calculations since SQL errors are handled separately
                for calc in step.findall(".//Calculation"):
                    if calc.text and "ExecuteSQL" not in calc.text:
                        # Look for table::field references
                        table_field_pattern = re.compile(r'([a-zA-Z0-9_🌎🧑‍🎓_🔗👥]+)::([a-zA-Z0-9_]+)')
                        matches = table_field_pattern.findall(calc.text)
                        
                        for ref_table, ref_field in matches:
                            base_table = table_occurrences_to_base.get(ref_table, ref_table)
                            
                            if base_table not in known_tables:
                                xml_line = getattr(calc, 'sourceline', 'unknown')
                                location = f'Step {step_number}: {step_name}'
                                error = f'Unknown table in calculation: "{ref_table}"'
                                status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Script",
                                    "Context": script_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Table referenced in calculation not found"
                                })
                            elif ref_field not in known_fields.get(base_table, set()):
                                xml_line = getattr(calc, 'sourceline', 'unknown')
                                location = f'Step {step_number}: {step_name}'
                                error = f'Unknown field in calculation: "{ref_table}::{ref_field}"'
                                status = determine_status(f'Script: {script_name}, {location}', error, is_script_step=True, step_elem=step)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Script",
                                    "Context": script_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, is_script_step=True, step_elem=step) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Field referenced in calculation not found"
                                })
        
        # Check layouts for unknown field references
        for layout in root.findall(".//Layout"):
            layout_name = layout.attrib.get("name", "Unknown Layout")
            
            # Check field objects
            for field_obj in layout.findall(".//Object[@type='Field']"):
                # Get position
                position = ""
                bounds = field_obj.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                field_ref = field_obj.find(".//Field")
                if field_ref is not None:
                    field_name = field_ref.attrib.get("name", "")
                    table_name = field_ref.attrib.get("table", "")
                    
                    if field_name and table_name:
                        # Get base table
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        
                        # Check if table exists
                        if base_table not in known_tables:
                            xml_line = getattr(field_ref, 'sourceline', 'unknown')
                            location = f'Field Object {position}'
                            error = f'Unknown table: "{table_name}"'
                            status = determine_status(f'Layout: {layout_name}', error)
                            
                            results.append({
                                "Status": status,
                                "Type": "Layout Field",
                                "Context": layout_name,
                                "Location": location,
                                "Commented": "Yes" if is_commented(status) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Table not found"
                            })
                        # Check if field exists in the table
                        elif field_name not in known_fields.get(base_table, set()):
                            xml_line = getattr(field_ref, 'sourceline', 'unknown')
                            location = f'Field Object {position}'
                            error = f'Unknown field: "{table_name}::{field_name}"'
                            status = determine_status(f'Layout: {layout_name}', error)
                            
                            results.append({
                                "Status": status,
                                "Type": "Layout Field",
                                "Context": layout_name,
                                "Location": location,
                                "Commented": "Yes" if is_commented(status) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Field not found in table"
                            })
            
            # Check buttons for unknown script references
            for button in layout.findall(".//Object[@type='Button']"):
                # Get position
                position = ""
                bounds = button.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                button_name = button.attrib.get("name", "Unnamed Button")
                
                # Check for script references in buttons
                for script_ref in button.findall(".//Script"):
                    ref_script_name = script_ref.attrib.get("name", "")
                    if ref_script_name and ref_script_name not in known_scripts:
                        xml_line = getattr(script_ref, 'sourceline', 'unknown')
                        location = f'Button: {button_name} {position}'
                        error = f'Unknown script: "{ref_script_name}"'
                        status = determine_status(f'Layout: {layout_name}', error)
                        
                        results.append({
                            "Status": status,
                            "Type": "Layout Object",
                            "Context": layout_name,
                            "Location": location,
                            "Commented": "Yes" if is_commented(status) else "No",
                            "Error": error,
                            "XML Line": xml_line,
                            "Details": "Script referenced by button not found"
                        })
                
                # Check SingleStep for script references
                single_step = button.find(".//SingleStep")
                if single_step is not None:
                    step = single_step.find(".//Step[@name='Perform Script']")
                    if step is not None:
                        script_ref = step.find(".//Script")
                        if script_ref is not None:
                            ref_script_name = script_ref.attrib.get("name", "")
                            if ref_script_name and ref_script_name not in known_scripts:
                                xml_line = getattr(script_ref, 'sourceline', 'unknown')
                                location = f'Button: {button_name} {position}'
                                error = f'Unknown script: "{ref_script_name}"'
                                status = determine_status(f'Layout: {layout_name}', error)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Layout Object",
                                    "Context": layout_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Script referenced by button not found"
                                })
            
            # Check portals for unknown table references
            for portal in layout.findall(".//Object[@type='Portal']"):
                # Get position
                position = ""
                bounds = portal.find(".//Bounds")
                if bounds is not None:
                    try:
                        top = round(float(bounds.attrib.get('top', '0')))
                        left = round(float(bounds.attrib.get('left', '0')))
                        position = f"Top: {top} Left: {left}"
                    except (ValueError, TypeError):
                        position = "Unknown Position"
                
                portal_obj = portal.find(".//Portal")
                if portal_obj is not None:
                    portal_table = portal_obj.attrib.get("table", "")
                    if portal_table:
                        base_table = table_occurrences_to_base.get(portal_table, portal_table)
                        if base_table not in known_tables:
                            xml_line = getattr(portal_obj, 'sourceline', 'unknown')
                            location = f'Portal {position}'
                            error = f'Unknown portal table: "{portal_table}"'
                            status = determine_status(f'Layout: {layout_name}', error)
                            
                            results.append({
                                "Status": status,
                                "Type": "Layout Object",
                                "Context": layout_name,
                                "Location": location,
                                "Commented": "Yes" if is_commented(status) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Portal table not found"
                            })
        
        # Check field calculations for unknown field references
        for table_node in root.findall(".//BaseTable"):
            table_name = table_node.attrib.get("name", "Unknown Table")
            
            for field_node in table_node.findall(".//Field"):
                field_name = field_node.attrib.get("name", "Unknown Field")
                
                # Check field calculation
                calc = field_node.find(".//Calculation")
                if calc is not None and calc.text:
                    # Skip calculations with ExecuteSQL since SQL errors are handled separately
                    if "ExecuteSQL" not in calc.text:
                        # Look for table::field references
                        table_field_pattern = re.compile(r'([a-zA-Z0-9_🌎🧑‍🎓_🔗👥]+)::([a-zA-Z0-9_]+)')
                        matches = table_field_pattern.findall(calc.text)
                        
                        for ref_table, ref_field in matches:
                            base_table = table_occurrences_to_base.get(ref_table, ref_table)
                            
                            if base_table not in known_tables:
                                xml_line = getattr(calc, 'sourceline', 'unknown')
                                location = f'Field: {field_name}'
                                error = f'Unknown table in calculation: "{ref_table}"'
                                status = determine_status(f'Field: {table_name}::{field_name}', error, calc_text=calc.text)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Field Calc",
                                    "Context": table_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, calc_text=calc.text) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Table referenced in calculation not found"
                                })
                            elif ref_field not in known_fields.get(base_table, set()):
                                xml_line = getattr(calc, 'sourceline', 'unknown')
                                location = f'Field: {field_name}'
                                error = f'Unknown field in calculation: "{ref_table}::{ref_field}"'
                                status = determine_status(f'Field: {table_name}::{field_name}', error, calc_text=calc.text)
                                
                                results.append({
                                    "Status": status,
                                    "Type": "Field Calc",
                                    "Context": table_name,
                                    "Location": location,
                                    "Commented": "Yes" if is_commented(status, calc_text=calc.text) else "No",
                                    "Error": error,
                                    "XML Line": xml_line,
                                    "Details": "Field referenced in calculation not found"
                                })
        
        # Check custom functions for unknown references
        for cf_node in root.findall(".//CustomFunction"):
            cf_name = cf_node.attrib.get("name", "Unknown Function")
            
            calc_node = cf_node.find(".//Calculation")
            if calc_node is not None and calc_node.text:
                # Check for custom function calls
                cf_pattern = re.compile(r'(\w+)\s*\(')
                potential_cf_calls = cf_pattern.findall(calc_node.text)
                
                for potential_cf in potential_cf_calls:
                    # Skip known FileMaker functions
                    fm_functions = [
                        'If', 'Let', 'Get', 'Set', 'Sum', 'Count', 'Max', 'Min', 'Average',
                        'Date', 'Time', 'Timestamp', 'Year', 'Month', 'Day', 'Hour', 'Minute',
                        'Second', 'Left', 'Right', 'Middle', 'Length', 'Position', 'Substitute',
                        'Trim', 'Upper', 'Lower', 'Proper', 'TextStyleAdd', 'TextStyleRemove',
                        'PatternCount', 'Filter', 'FilterValues', 'GetValue', 'ValueCount',
                        'List', 'IsEmpty', 'IsValid', 'Case', 'Choose', 'Evaluate', 'Extend',
                        'Lookup', 'Last', 'GetField', 'GetFieldName', 'GetLayoutObjectAttribute',
                        'Self', 'GetNthRecord', 'GetRepetition'
                    ]
                    
                    if (potential_cf not in fm_functions and 
                        potential_cf not in known_custom_functions):
                        xml_line = getattr(calc_node, 'sourceline', 'unknown')
                        location = f'Custom Function: {cf_name}'
                        error = f'Unknown custom function: "{potential_cf}"'
                        status = determine_status(location, error, calc_text=calc_node.text)
                        
                        results.append({
                            "Status": status,
                            "Type": "Custom Function",
                            "Context": "",  # Custom functions don't belong to tables/layouts
                            "Location": location,
                            "Commented": "Yes" if is_commented(status, calc_text=calc_node.text) else "No",
                            "Error": error,
                            "XML Line": xml_line,
                            "Details": "Custom function call not found"
                        })
        
        # Check value lists for unknown field references
        for vl in root.findall(".//ValueList"):
            vl_name = vl.attrib.get("name", "Unknown Value List")
            
            # Check field-based value lists
            field_ref = vl.find(".//Field")
            if field_ref is not None:
                field_name = field_ref.attrib.get("name", "")
                table_name = field_ref.attrib.get("table", "")
                
                if field_name and table_name:
                    base_table = table_occurrences_to_base.get(table_name, table_name)
                    
                    if base_table not in known_tables:
                        xml_line = getattr(field_ref, 'sourceline', 'unknown')
                        location = f'Value List: {vl_name}'
                        error = f'Unknown table: "{table_name}"'
                        status = determine_status(location, error)
                        
                        results.append({
                            "Status": status,
                            "Type": "Value List",
                            "Context": "",  # Value lists don't belong to tables/layouts
                            "Location": location,
                            "Commented": "Yes" if is_commented(status) else "No",
                            "Error": error,
                            "XML Line": xml_line,
                            "Details": "Table referenced in value list not found"
                        })
                    elif field_name not in known_fields.get(base_table, set()):
                        xml_line = getattr(field_ref, 'sourceline', 'unknown')
                        location = f'Value List: {vl_name}'
                        error = f'Unknown field: "{table_name}::{field_name}"'
                        status = determine_status(location, error)
                        
                        results.append({
                            "Status": status,
                            "Type": "Value List",
                            "Context": "",  # Value lists don't belong to tables/layouts
                            "Location": location,
                            "Commented": "Yes" if is_commented(status) else "No",
                            "Error": error,
                            "XML Line": xml_line,
                            "Details": "Field referenced in value list not found"
                        })
        
        # Check relationships for unknown field references
        for relationship in root.findall(".//Relationship"):
            rel_name = relationship.attrib.get("name", "Unknown Relationship")
            
            # Check field pairs
            for field_pair in relationship.findall(".//FieldPair"):
                for i, field_ref in enumerate(field_pair.findall(".//Field"), 1):
                    field_name = field_ref.attrib.get("name", "")
                    table_name = field_ref.attrib.get("table", "")
                    
                    if field_name and table_name:
                        base_table = table_occurrences_to_base.get(table_name, table_name)
                        
                        if base_table not in known_tables:
                            xml_line = getattr(field_ref, 'sourceline', 'unknown')
                            location = f'Relationship: {rel_name}, Field {i}'
                            error = f'Unknown table: "{table_name}"'
                            status = determine_status(location, error)
                            
                            results.append({
                                "Status": status,
                                "Type": "Relationship",
                                "Context": "",  # Relationships don't belong to tables/layouts
                                "Location": location,
                                "Commented": "Yes" if is_commented(status) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Table referenced in relationship not found"
                            })
                        elif field_name not in known_fields.get(base_table, set()):
                            xml_line = getattr(field_ref, 'sourceline', 'unknown')
                            location = f'Relationship: {rel_name}, Field {i}'
                            error = f'Unknown field: "{table_name}::{field_name}"'
                            status = determine_status(location, error)
                            
                            results.append({
                                "Status": status,
                                "Type": "Relationship",
                                "Context": "",  # Relationships don't belong to tables/layouts
                                "Location": location,
                                "Commented": "Yes" if is_commented(status) else "No",
                                "Error": error,
                                "XML Line": xml_line,
                                "Details": "Field referenced in relationship not found"
                            })
        
        # Sort results by Type (Scripts first, then Field Calc, then Layout Field, then others), then Context, then Location
        type_order = {
            "Script": 0,
            "Field Calc": 1,
            "Layout Field": 2,
            "Layout Object": 3,
            "Custom Function": 4,
            "Value List": 5,
            "Relationship": 6,
            "Other": 7
        }
        
        status_order = {
            "Active Error": 0,
            "To Delete": 1,
            "Commented Out": 2,
            "Temp/Debug": 3
        }
        
        results.sort(key=lambda x: (
            type_order.get(x["Type"], 999),
            x["Context"],
            x["Location"],
            status_order.get(x["Status"], 999)
        ))
        
        return results
        
    except Exception as e:
        print(f"Error in UnknownCheck: {e}")
        import traceback
        traceback.print_exc()
        return []