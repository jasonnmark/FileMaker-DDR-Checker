from lxml import etree as ET
import re
from collections import defaultdict

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Table Occurrences"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 10

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Table Occurrence": 200,
        "Base Table": 200,
        "Same as Base": 70,
        "XML Occurrences": 70,
        "Usage Count": 70,
        "Usage Locations": 400,
        "Relationships": 300
    }

def apply_styling(ws):
    """Apply custom styling to the Table Occurrences worksheet"""
    from openpyxl.styles import PatternFill, Font
    
    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx
    
    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Check if usage count is 0 to apply red background to entire row
        if "Usage Count" in columns:
            usage_count_cell = row[columns["Usage Count"] - 1]
            if usage_count_cell.value == 0:
                # Apply error color (red background) to entire row
                for cell in row:
                    cell.fill = get_color_fill('error')
                
                # Make usage count and table occurrence name bold
                usage_count_cell.font = get_color_font('error', bold=True)
                if "Table Occurrence" in columns:
                    to_cell = row[columns["Table Occurrence"] - 1]
                    to_cell.font = get_color_font('error', bold=True)
        
        # Additionally highlight rows where table occurrence name differs from base table
        if "Same as Base" in columns:
            same_cell = row[columns["Same as Base"] - 1]
            if same_cell.value == "No":
                # Only apply warning color if not already red from zero usage
                if "Usage Count" not in columns or row[columns["Usage Count"] - 1].value != 0:
                    same_cell.fill = get_color_fill('warning')
                    same_cell.font = get_color_font('warning', bold=True)
                    
                    # Also highlight the table occurrence name
                    if "Table Occurrence" in columns:
                        to_cell = row[columns["Table Occurrence"] - 1]
                        to_cell.font = get_color_font('warning', bold=True)

def run_check(raw_xml, catalogs=None):
    """
    Extract all table occurrences and their base tables from the raw XML.
    Returns a list of dictionaries suitable for Excel output.
    Now can use shared catalogs for consistency.
    """
    try:
        # Use shared catalogs if provided (new style)
        if catalogs:
            table_occurrences = catalogs['table_occurrences']
            root = catalogs['root']
        else:
            # Fall back to parsing ourselves (old style)
            parser = ET.XMLParser(remove_blank_text=True, recover=True)
            root = ET.fromstring(raw_xml.encode('utf-8'), parser)
            
            # Dictionary to store table occurrences
            table_occurrences = {}
            
            # First, check for Table elements with name and baseTable attributes
            for table_node in root.findall(".//Table"):
                occurrence_name = table_node.attrib.get("name")
                base_table_name = table_node.attrib.get("baseTable")
                if occurrence_name and base_table_name:
                    table_occurrences[occurrence_name] = base_table_name
            
            # Also check for TableOccurrence elements (different structure in some DDRs)
            for table_occ in root.findall(".//TableOccurrence"):
                occ_name = table_occ.attrib.get("name")
                base_table = table_occ.attrib.get("baseTable")
                if occ_name and base_table:
                    table_occurrences[occ_name] = base_table
            
            # Check in RelationshipGraph for table occurrences
            for table_occurrence in root.findall(".//RelationshipGraph//TableOccurrence"):
                name = table_occurrence.attrib.get("name")
                base_table = table_occurrence.attrib.get("baseTable")
                if name and base_table:
                    table_occurrences[name] = base_table
        
        # Build usage and relationships dictionaries in single passes
        table_usage = defaultdict(list)
        table_relationships = defaultdict(list)
        
        # Collect all relationships in one pass
        for relationship in root.findall(".//Relationship"):
            left_table = relationship.find(".//LeftTable")
            right_table = relationship.find(".//RightTable")
            
            if left_table is not None and right_table is not None:
                left_name = left_table.attrib.get("name")
                right_name = right_table.attrib.get("name")
                
                if left_name and right_name:
                    table_relationships[left_name].append(f"→ {right_name}")
                    table_relationships[right_name].append(f"← {left_name}")
        
        # Collect all script usage in one pass
        for script in root.findall(".//Script"):
            script_name = script.attrib.get("name", "Unknown Script")
            
            for step in script.findall(".//Step"):
                step_id = step.attrib.get("id", "")
                step_name = step.attrib.get("name", "Unknown Step")
                
                # Check Set Field steps
                if step_name == "Set Field":
                    field = step.find(".//Field")
                    if field is not None:
                        table_name = field.attrib.get("table")
                        if table_name:
                            table_usage[table_name].append(f"Script '{script_name}' Step {step_id}: Set Field")
                
                # Check Set Field By Name steps
                elif step_name == "Set Field By Name":
                    calc = step.find(".//Calculation")
                    if calc is not None and calc.text:
                        for to_name in table_occurrences:
                            if to_name in calc.text:
                                table_usage[to_name].append(f"Script '{script_name}' Step {step_id}: Set Field By Name")
                
                # Check Execute SQL steps and other steps with SQL
                elif step_name in ["Execute SQL", "Set Field", "Set Variable", "If", "Exit Loop If", "Set Field By Name"]:
                    # Check for SQL in calculations within these steps
                    for calc in step.findall(".//Calculation"):
                        if calc is not None and calc.text and "ExecuteSQL" in calc.text:
                            # Use improved SQL parsing
                            sql_patterns = [
                                r'FROM\s+([^\s,;]+)',  # Basic FROM table
                                r'FROM\s+"([^"]+)"',    # FROM "table"
                                r'FROM\s+\'([^\']+)\'', # FROM 'table'
                                r'JOIN\s+([^\s,;]+)',   # JOIN table
                                r'JOIN\s+"([^"]+)"',    # JOIN "table"
                                r'JOIN\s+\'([^\']+)\''  # JOIN 'table'
                            ]
                            
                            found_tables = set()
                            for pattern in sql_patterns:
                                matches = re.findall(pattern, calc.text, re.IGNORECASE | re.MULTILINE)
                                for match in matches:
                                    # Clean up the table name
                                    table_ref = match.strip().strip('"').strip("'")
                                    # Remove any WHERE/ORDER/etc that might have been captured
                                    table_ref = table_ref.split()[0] if table_ref else ""
                                    if table_ref and table_ref in table_occurrences:
                                        found_tables.add(table_ref)
                            
                            # Add usage for each found table
                            for table_ref in found_tables:
                                sql_snippet = calc.text[:100] + "..." if len(calc.text) > 100 else calc.text
                                table_usage[table_ref].append(f"Script '{script_name}' Step {step_id}: {step_name} (ExecuteSQL)")
                
                # Original Execute SQL step handling (kept for backwards compatibility)
                if step_name == "Execute SQL":
                    calc = step.find(".//Calculation")
                    if calc is not None and calc.text:
                        for to_name in table_occurrences:
                            if to_name in calc.text and f"Script '{script_name}' Step {step_id}: {step_name} (ExecuteSQL)" not in table_usage[to_name]:
                                sql_snippet = calc.text[:100] + "..." if len(calc.text) > 100 else calc.text
                                table_usage[to_name].append(f"Script '{script_name}' Step {step_id}: Execute SQL - {sql_snippet}")
                
                # Check Go to Layout steps
                elif step_name == "Go to Layout":
                    layout = step.find(".//Layout")
                    if layout is not None:
                        table_ref = layout.attrib.get("table")
                        if table_ref:
                            layout_name = layout.attrib.get("name", "Unknown")
                            table_usage[table_ref].append(f"Script '{script_name}' Step {step_id}: Go to Layout '{layout_name}'")
                
                # Check any other Field references
                for field in step.findall(".//Field"):
                    table_name = field.attrib.get("table")
                    if table_name and f"Script '{script_name}' Step {step_id}: {step_name}" not in table_usage[table_name]:
                        table_usage[table_name].append(f"Script '{script_name}' Step {step_id}: {step_name}")
        
        # Collect all layout usage in one pass
        for layout in root.findall(".//Layout"):
            layout_name = layout.attrib.get("name", "Unknown Layout")
            layout_table = layout.attrib.get("table", "")
            
            if layout_table:
                table_usage[layout_table].append(f"Layout '{layout_name}' (Based on Table Occurrence)")
            
            # Also check the Table element within Layout
            table_elem = layout.find(".//Table")
            if table_elem is not None:
                table_name = table_elem.attrib.get("name")
                if table_name and table_name != layout_table:  # Avoid duplicate if same as layout table
                    table_usage[table_name].append(f"Layout '{layout_name}' (Table Reference)")
            
            # Field objects
            for field_obj in layout.findall(".//FieldObj"):
                field = field_obj.find(".//Field")
                if field is not None:
                    table_name = field.attrib.get("table")
                    if table_name:
                        bounds = field_obj.find(".//Bounds")
                        if bounds is not None:
                            top = int(float(bounds.attrib.get("top", "0")))
                            left = int(float(bounds.attrib.get("left", "0")))
                            table_usage[table_name].append(f"Field Layout '{layout_name}' Object FieldObj Top {top} Left {left}")
            
            # Portal objects
            for portal in layout.findall(".//PortalObj"):
                table_alias = portal.find(".//TableAliasKey")
                if table_alias is not None and table_alias.text:
                    bounds = portal.find(".//Bounds")
                    if bounds is not None:
                        top = int(float(bounds.attrib.get("top", "0")))
                        left = int(float(bounds.attrib.get("left", "0")))
                        table_usage[table_alias.text].append(f"Portal Layout '{layout_name}' Object Portal Top {top} Left {left}")
                
                # Portal filter
                filter_calc = portal.find(".//Calculation")
                if filter_calc is not None and filter_calc.text:
                    for to_name in table_occurrences:
                        if to_name in filter_calc.text:
                            bounds = portal.find(".//Bounds")
                            if bounds is not None:
                                top = int(float(bounds.attrib.get("top", "0")))
                                left = int(float(bounds.attrib.get("left", "0")))
                                table_usage[to_name].append(f"Portal Filter Layout '{layout_name}' Object Portal Top {top} Left {left}")
            
            # Check ALL objects for various calculation types
            for obj in layout.findall(".//Object"):
                obj_type = obj.attrib.get("type", "Unknown")
                
                # Check HideCondition calculations
                hide_calc = obj.find(".//HideCondition/Calculation")
                if hide_calc is not None and hide_calc.text:
                    # Check for ExecuteSQL
                    if "ExecuteSQL" in hide_calc.text:
                        sql_patterns = [
                            r'FROM\s+([^\s,;]+)',  # Basic FROM table
                            r'FROM\s+"([^"]+)"',    # FROM "table"
                            r'FROM\s+\'([^\']+)\'', # FROM 'table'
                            r'JOIN\s+([^\s,;]+)',   # JOIN table
                            r'JOIN\s+"([^"]+)"',    # JOIN "table"
                            r'JOIN\s+\'([^\']+)\''  # JOIN 'table'
                        ]
                        
                        found_tables = set()
                        for pattern in sql_patterns:
                            matches = re.findall(pattern, hide_calc.text, re.IGNORECASE | re.MULTILINE)
                            for match in matches:
                                table_ref = match.strip().strip('"').strip("'")
                                table_ref = table_ref.split()[0] if table_ref else ""
                                if table_ref and table_ref in table_occurrences:
                                    found_tables.add(table_ref)
                        
                        for table_ref in found_tables:
                            table_usage[table_ref].append(f"Hide Condition on Layout '{layout_name}' - {obj_type} (ExecuteSQL)")
                    
                    # Also check for direct table references
                    for to_name in table_occurrences:
                        if to_name in hide_calc.text and to_name not in found_tables:
                            table_usage[to_name].append(f"Hide Condition on Layout '{layout_name}' - {obj_type}")
                
                # Check Tooltip calculations
                tooltip_calc = obj.find(".//Tooltip/Calculation")
                if tooltip_calc is not None and tooltip_calc.text:
                    # Check for ExecuteSQL
                    if "ExecuteSQL" in tooltip_calc.text:
                        found_tables = set()
                        for pattern in sql_patterns:
                            matches = re.findall(pattern, tooltip_calc.text, re.IGNORECASE | re.MULTILINE)
                            for match in matches:
                                table_ref = match.strip().strip('"').strip("'")
                                table_ref = table_ref.split()[0] if table_ref else ""
                                if table_ref and table_ref in table_occurrences:
                                    found_tables.add(table_ref)
                        
                        for table_ref in found_tables:
                            table_usage[table_ref].append(f"Tooltip on Layout '{layout_name}' - {obj_type} (ExecuteSQL)")
                    
                    # Check direct references
                    for to_name in table_occurrences:
                        if to_name in tooltip_calc.text and to_name not in found_tables:
                            table_usage[to_name].append(f"Tooltip on Layout '{layout_name}' - {obj_type}")
                
                # Check Conditional Formatting calculations
                for cond_format in obj.findall(".//ConditionalFormatting/Calculation"):
                    if cond_format.text:
                        # Check for ExecuteSQL
                        if "ExecuteSQL" in cond_format.text:
                            found_tables = set()
                            for pattern in sql_patterns:
                                matches = re.findall(pattern, cond_format.text, re.IGNORECASE | re.MULTILINE)
                                for match in matches:
                                    table_ref = match.strip().strip('"').strip("'")
                                    table_ref = table_ref.split()[0] if table_ref else ""
                                    if table_ref and table_ref in table_occurrences:
                                        found_tables.add(table_ref)
                            
                            for table_ref in found_tables:
                                table_usage[table_ref].append(f"Conditional Format on Layout '{layout_name}' - {obj_type} (ExecuteSQL)")
                        
                        # Check direct references
                        for to_name in table_occurrences:
                            if to_name in cond_format.text and to_name not in found_tables:
                                table_usage[to_name].append(f"Conditional Format on Layout '{layout_name}' - {obj_type}")
                
                # Check PlaceholderText calculations (for fields)
                placeholder_calc = obj.find(".//PlaceholderText/Calculation")
                if placeholder_calc is not None and placeholder_calc.text:
                    for to_name in table_occurrences:
                        if to_name in placeholder_calc.text:
                            table_usage[to_name].append(f"Placeholder Text on Layout '{layout_name}' - {obj_type}")
                
                # Check any other Calculation nodes within objects
                for calc in obj.findall(".//Calculation"):
                    # Skip if already processed above
                    parent = calc.getparent()
                    if parent is not None and parent.tag in ["HideCondition", "Tooltip", "ConditionalFormatting", "PlaceholderText"]:
                        continue
                    
                    if calc.text:
                        # Check for ExecuteSQL
                        if "ExecuteSQL" in calc.text:
                            found_tables = set()
                            for pattern in sql_patterns:
                                matches = re.findall(pattern, calc.text, re.IGNORECASE | re.MULTILINE)
                                for match in matches:
                                    table_ref = match.strip().strip('"').strip("'")
                                    table_ref = table_ref.split()[0] if table_ref else ""
                                    if table_ref and table_ref in table_occurrences:
                                        found_tables.add(table_ref)
                            
                            for table_ref in found_tables:
                                context = parent.tag if parent is not None else "Calculation"
                                table_usage[table_ref].append(f"{context} on Layout '{layout_name}' - {obj_type} (ExecuteSQL)")
                        
                        # Check direct references
                        for to_name in table_occurrences:
                            if to_name in calc.text:
                                context = parent.tag if parent is not None else "Calculation"
                                if f"{context} on Layout '{layout_name}' - {obj_type}" not in table_usage[to_name]:
                                    table_usage[to_name].append(f"{context} on Layout '{layout_name}' - {obj_type}")
        
        # The rest of the function remains the same...
        # (collecting value list usage, field calculation usage, converting to results, etc.)
        
        # Convert to list format for Excel output
        results = []
        
        # Pre-compute XML counts for all table occurrences at once
        xml_counts = {}
        for occurrence in table_occurrences:
            # Simple count without regex for performance
            xml_counts[occurrence] = raw_xml.count(occurrence)
        
        for occurrence, base in table_occurrences.items():
            # Get XML count from pre-computed dictionary
            xml_count = xml_counts.get(occurrence, 0)
            
            # Get usage and relationships from pre-built dictionaries
            usage_locations = table_usage.get(occurrence, [])
            relationships = table_relationships.get(occurrence, [])
            
            # Remove duplicates while preserving order
            usage_locations = list(dict.fromkeys(usage_locations))
            
            # Format for display
            usage_text = "\n".join(usage_locations) if usage_locations else ""
            relationships_text = "\n".join(relationships) if relationships else ""
            
            # Check if occurrence matches base table (considering z_ prefix)
            same_as_base = "No"
            if occurrence == base:
                same_as_base = "Yes"
            elif occurrence.startswith("z_") and occurrence[2:] == base:
                same_as_base = "Yes"
            elif base.startswith("z_") and base[2:] == occurrence:
                same_as_base = "Yes"
            
            results.append({
                "Table Occurrence": occurrence,
                "Base Table": base,
                "Same as Base": same_as_base,
                "XML Occurrences": xml_count,
                "Usage Count": len(usage_locations),
                "Usage Locations": usage_text,
                "Relationships": relationships_text
            })
        
        # Sort results: first by usage count (0 first), then by XML occurrences (low to high), then by table occurrence name
        results.sort(key=lambda x: (x["Usage Count"], x["XML Occurrences"], x["Table Occurrence"]))
        
        return results
        
    except Exception as e:
        print(f"Error in Table Occurrences Check: {e}")
        return []
