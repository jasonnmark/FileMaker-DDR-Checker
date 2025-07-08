from lxml import etree as ET
import re
from collections import defaultdict

def get_sheet_name():
    """Return the name for the Excel sheet"""
    return "Custom Function Usage"

def get_sheet_order():
    """Return the order for this sheet (lower numbers appear first)"""
    return 3

def get_column_widths():
    """Return column widths for this sheet"""
    return {
        "Custom Function": 200,
        "Usage Count": 80,
        "Used In": 400,
        "Status": 85
    }

def apply_styling(ws):
    """Apply custom styling to the Custom Function Usage worksheet"""
    from openpyxl.styles import PatternFill, Font

    # Find column indexes
    columns = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            columns[cell.value] = idx

    # Apply row-by-row styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Check status for unused functions
        if "Status" in columns:
            status_cell = row[columns["Status"] - 1]
            if status_cell.value == "Unused":
                # Highlight entire row for unused functions
                for cell in row:
                    cell.fill = get_color_fill('error')

                # Make status cell bold with error font
                status_cell.font = get_color_font('error', bold=True)

                # Also make the "Used In" cell bold and red for emphasis
                if "Used In" in columns:
                    used_in_cell = row[columns["Used In"] - 1]
                    used_in_cell.font = get_color_font('error', bold=True)

        # Color code usage count
        if "Usage Count" in columns:
            count_cell = row[columns["Usage Count"] - 1]
            try:
                count = int(count_cell.value or 0)
                if count == 0:
                    count_cell.font = get_color_font('error', bold=True)
                elif count >= 10:
                    count_cell.font = get_color_font('success', bold=True)
                elif count >= 5:
                    count_cell.font = Font(bold=True, color="388E3C")  # Medium green
            except:
                pass

def get_color_fill(color_name):
    from openpyxl.styles import PatternFill
    colors = {
        "warning": "FFFFCC",
        "success": "CCFFCC",
        "error": "FFCCCC",
    }
    return PatternFill(start_color=colors.get(color_name, "FFFFFF"), end_color=colors.get(color_name, "FFFFFF"), fill_type="solid")

def get_color_font(color_name, bold=False):
    from openpyxl.styles import Font
    colors = {
        "warning": "CC6600",
        "success": "006600",
        "error": "FF0000",
    }
    return Font(color=colors.get(color_name, "000000"), bold=bold)

def run_check(raw_xml):
    """
    Find all custom functions and where they are called.
    Flag the ones that aren't called anywhere.
    Sort so unused functions appear at the top.
    """
    try:
        # Parse the XML
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        root = ET.fromstring(raw_xml.encode('utf-8'), parser)

        # Dictionary to store custom functions
        custom_functions = {}

        # Dictionary to store where each function is used
        function_usage = defaultdict(list)

        cf_called_by = defaultdict(set)  # cf_called_by[called_function] = set of calling_functions

        # First, find all custom functions
        for cf_node in root.findall(".//CustomFunctionCatalog/CustomFunction"):
            cf_name = cf_node.attrib.get("name", "Unknown Function")
            cf_id = cf_node.attrib.get("id", "")
            custom_functions[cf_name] = {
                "id": cf_id,
                "name": cf_name,
                "definition": ""
            }

            # Try to get the function definition
            calc_node = cf_node.find(".//Calculation")
            if calc_node is not None and calc_node.text:
                custom_functions[cf_name]["definition"] = calc_node.text.strip()[:100] + "..." if len(calc_node.text.strip()) > 100 else calc_node.text.strip()

        print(f"Found {len(custom_functions)} custom functions")

        # Pre-compile regex patterns for all custom functions
        function_patterns = {}
        for cf_name in custom_functions:
            # Use word boundaries and escape special regex characters
            # Look for the function name followed by either:
            # 1. Opening parenthesis (function call)
            # 2. Word boundary (used as parameter or reference)
            pattern = re.compile(r'\b' + re.escape(cf_name) + r'(?:\s*\(|\b)', re.IGNORECASE)
            function_patterns[cf_name] = pattern

        # Helper function to check if a calculation contains custom function calls
        def find_custom_function_calls(text, location_info):
            if not text:
                return

            # Check each custom function pattern against the text
            for cf_name, pattern in function_patterns.items():
                if pattern.search(text):
                    function_usage[cf_name].append(location_info)

        # Search in Scripts - batch process steps
        print("  Searching in scripts...")
        for script in root.findall(".//Script"):
            script_name = script.attrib.get("name", "Unknown Script")
            for calc in script.findall(".//Calculation"):
                if calc.text:
                    step = calc.getparent()
                    if step is not None and step.tag == "Step":
                        step_name = step.attrib.get("name", "Unknown Step")
                        step_index = step.attrib.get("index", "?")
                        location = f"Script: {script_name}, Step {step_index}: {step_name}"
                    else:
                        location = f"Script: {script_name}"
                    find_custom_function_calls(calc.text, location)

        # Search in Field Calculations
        print("  Searching in field calculations...")
        for table in root.findall(".//BaseTable"):
            table_name = table.attrib.get("name", "Unknown Table")
            for field in table.findall(".//Field"):
                field_name = field.attrib.get("name", "Unknown Field")
                # Calculation
                calc_node = field.find(".//Calculation")
                if calc_node is not None and calc_node.text:
                    location = f"Field Calc: {table_name}::{field_name}"
                    find_custom_function_calls(calc_node.text, location)
                # Auto-Enter
                auto_enter = field.find(".//AutoEnter/Calculation")
                if auto_enter is not None and auto_enter.text:
                    location = f"Auto-Enter: {table_name}::{field_name}"
                    find_custom_function_calls(auto_enter.text, location)
                # Validation
                validation = field.find(".//Validation/Calculation")
                if validation is not None and validation.text:
                    location = f"Validation: {table_name}::{field_name}"
                    find_custom_function_calls(validation.text, location)

        # Search in Layout Objects
        print("  Searching in layout objects...")
        for layout in root.findall(".//Layout"):
            layout_name = layout.attrib.get("name", "Unknown Layout")
            for obj in layout.findall(".//Object"):
                obj_type = obj.attrib.get("type", "Unknown")
                obj_name = obj.attrib.get("name", "")
                for calc in obj.findall(".//Calculation"):
                    if calc.text:
                        parent = calc.getparent()
                        context = ""
                        if parent is not None:
                            if parent.tag == "HideCondition":
                                context = " (Hide)"
                            elif parent.tag == "ConditionalFormatting":
                                context = " (Conditional)"
                            elif parent.tag == "Tooltip":
                                context = " (Tooltip)"
                        location = f"Layout Object: {layout_name} - {obj_type}"
                        if obj_name:
                            location += f" '{obj_name}'"
                        location += context
                        find_custom_function_calls(calc.text, location)

        # Search in Value Lists
        for vl in root.findall(".//ValueList"):
            vl_name = vl.attrib.get("name", "Unknown Value List")
            calc = vl.find(".//Calculation")
            if calc is not None and calc.text:
                location = f"Value List: {vl_name}"
                find_custom_function_calls(calc.text, location)

        # Search in Privileges
        for priv_set in root.findall(".//PrivilegeSet"):
            priv_name = priv_set.attrib.get("name", "Unknown Privilege Set")
            for calc in priv_set.findall(".//Calculation"):
                if calc.text:
                    location = f"Privilege Set: {priv_name}"
                    find_custom_function_calls(calc.text, location)

        # Search in other Custom Functions (recursive usage)
        print("  Searching in custom functions...")
        cf_patterns = {}
        for cf_name, cf_info in custom_functions.items():
            if cf_info["definition"]:
                cf_patterns[cf_name] = cf_info["definition"]

        # Now check each definition against all other functions
        for cf_name, definition in cf_patterns.items():
            for other_cf, pattern in function_patterns.items():
                if pattern.search(definition):
                    location = f"Custom Function: {cf_name}"
                    function_usage[other_cf].append(location)
                    cf_called_by[other_cf].add(cf_name)

        print("  Building results...")

        results = []

        # First add unused functions
        for cf_name in sorted(custom_functions.keys()):
            callers = cf_called_by.get(cf_name, set())
            # If no usage, or only called by itself, mark as unused
            if cf_name not in function_usage or callers == {cf_name}:
                results.append({
                    "Custom Function": cf_name,
                    "Usage Count": 0,
                    "Used In": "⚠️ NOT USED",
                    "Status": "Unused"
                })

        # Then add used functions, sorted by usage count (most used first)
        used_functions = [(cf_name, len(locations)) for cf_name, locations in function_usage.items()]
        used_functions.sort(key=lambda x: x[1], reverse=True)

        for cf_name, count in used_functions:
            callers = cf_called_by.get(cf_name, set())
            # Skip if only called by itself (already marked as unused above)
            if callers == {cf_name}:
                continue
            locations = function_usage[cf_name]
            callers = callers - {cf_name}
            location_groups = defaultdict(list)
            for loc in locations:
                loc_type = loc.split(":")[0]
                location_groups[loc_type].append(loc)
            formatted_locations = []
            for loc_type in sorted(location_groups.keys()):
                grouped_locs = location_groups[loc_type]
                if len(grouped_locs) <= 3:
                    formatted_locations.extend(grouped_locs)
                else:
                    formatted_locations.extend(grouped_locs[:2])
                    formatted_locations.append(f"...and {len(grouped_locs) - 2} more {loc_type}(s)")
            used_in_note = ""
            if callers:
                used_in_note = f"\nCalled by: {', '.join(sorted(callers))}"
            results.append({
                "Custom Function": cf_name,
                "Usage Count": count,
                "Used In": ("\n".join(formatted_locations[:10]) + used_in_note).strip(),
                "Status": "Active"
            })

        return results

    except Exception as e:
        print(f"Error in CustomFunctionUsageCheck: {e}")
        return []