import sys
import subprocess
import platform
import os
import json
import pickle
import hashlib

# Suppress macOS Tkinter warnings
if platform.system() == "Darwin":
    os.environ['TK_SILENCE_DEPRECATION'] = '1'

from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
import re
from lxml import etree as ET
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import importlib.util

# Check for debug mode and cache mode early
DEBUG_MODE = '-debug' in sys.argv or '--debug' in sys.argv
CACHE_MODE = '-cache' in sys.argv or '--cache' in sys.argv

# Define standard colors and styles that check modules can use
STANDARD_COLORS = {
    'error': {'fill': 'FFEBEE', 'font': 'B71C1C'},  # Red
    'error_strong': {'fill': 'FFCDD2', 'font': 'B71C1C'},  # Darker red
    'warning': {'fill': 'FFF9C4', 'font': 'F57C00'},  # Yellow/Orange
    'success': {'fill': 'C8E6C9', 'font': '2E7D32'},  # Green
    'info': {'fill': 'E3F2FD', 'font': '1976D2'},  # Blue
    'muted': {'fill': 'F5F5F5', 'font': '888888'},  # Grey
    'header': {'fill': '2D3142', 'font': 'FFFFFF'},  # Dark header
    
    # Category colors for SQL sheets
    'category_script': {'fill': 'B7E4C7', 'font': '000000'},
    'category_custom_function': {'fill': 'A9DEF9', 'font': '000000'},
    'category_field_calc': {'fill': 'FFE066', 'font': '000000'},
    'category_layout_object': {'fill': 'FFB4A2', 'font': '000000'},
    'category_other': {'fill': 'D3D3D3', 'font': '000000'},
}

def get_cache_dir():
    """Get the Cache directory path"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cache_dir = os.path.join(script_dir, "Cache")
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    return cache_dir

def get_exports_dir():
    """Get the Exports directory path"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    exports_dir = os.path.join(script_dir, "Exports")
    if not os.path.exists(exports_dir):
        os.makedirs(exports_dir)
    return exports_dir

def get_file_hash(file_path):
    """Get a hash of the file content for cache validation"""
    hasher = hashlib.md5()
    with open(file_path, 'rb') as f:
        # Read in chunks to handle large files
        for chunk in iter(lambda: f.read(4096), b""):
            hasher.update(chunk)
    return hasher.hexdigest()

def get_cache_filename(xml_file):
    """Get the cache filename for a given XML file"""
    file_hash = get_file_hash(xml_file)
    base_name = os.path.splitext(os.path.basename(xml_file))[0]
    return f"{base_name}_{file_hash}.cache"

def get_last_used_file():
    """Get the last used file from cache directory"""
    if not CACHE_MODE:
        return None
    
    cache_dir = get_cache_dir()
    last_file_path = os.path.join(cache_dir, ".last_used_file")
    
    if os.path.exists(last_file_path):
        try:
            with open(last_file_path, 'r') as f:
                last_file = f.read().strip()
                if os.path.exists(last_file):
                    return last_file
        except:
            pass
    
    return None

def save_last_used_file(file_path):
    """Save the last used file path"""
    if not CACHE_MODE:
        return
    
    cache_dir = get_cache_dir()
    last_file_path = os.path.join(cache_dir, ".last_used_file")
    
    try:
        with open(last_file_path, 'w') as f:
            f.write(file_path)
    except:
        pass

def save_to_cache(xml_file, normalized_xml, catalogs):
    """Save normalized XML and catalogs to cache file"""
    if not CACHE_MODE:
        return
    
    cache_dir = get_cache_dir()
    cache_file = os.path.join(cache_dir, get_cache_filename(xml_file))
    
    try:
        # Cache both the normalized XML and the catalogs
        # We'll parse the root from XML when loading to save space
        cache_data = {
            'normalized_xml': normalized_xml,
            'catalogs': {k: v for k, v in catalogs.items() if k not in ['raw_xml', 'root']}
        }
        
        with open(cache_file, 'wb') as f:
            pickle.dump(cache_data, f)
        
        print(f"✓ Saved normalized XML and catalogs to cache: {os.path.basename(cache_file)}")
    except Exception as e:
        print(f"Warning: Could not save cache: {e}")

def load_from_cache(xml_file):
    """Load normalized XML and catalogs from cache file if it exists and is valid"""
    if not CACHE_MODE:
        return None, None
    
    cache_dir = get_cache_dir()
    cache_file = os.path.join(cache_dir, get_cache_filename(xml_file))
    
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'rb') as f:
                cache_data = pickle.load(f)
            
            print(f"✓ Loaded data from cache: {os.path.basename(cache_file)}")
            return cache_data.get('normalized_xml'), cache_data.get('catalogs')
        except Exception as e:
            print(f"Warning: Could not load cache: {e}")
            # Delete corrupted cache file
            try:
                os.remove(cache_file)
            except:
                pass
    
    return None, None

def get_color_fill(color_name):
    """Get a PatternFill object for a standard color"""
    if color_name in STANDARD_COLORS:
        color = STANDARD_COLORS[color_name]['fill']
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None

def get_color_font(color_name, bold=False, strike=False):
    """Get a Font object for a standard color"""
    if color_name in STANDARD_COLORS:
        color = STANDARD_COLORS[color_name]['font']
        return Font(color=color, bold=bold, strike=strike)
    return Font(bold=bold, strike=strike)

def check_and_install(package):
    # Handle special case where package name differs from import name
    import_name = package
    if package == 'pyahocorasick':
        import_name = 'ahocorasick'
    
    try:
        __import__(import_name)
        print(f"✓ {package} found")
    except ImportError:
        print(f"❌ {package} not found")
        response = input(f"Install {package} now? (y/n): ").lower().strip()
        if response in ['y', 'yes']:
            print(f"Installing {package}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"✓ {package} installed successfully!")
            except Exception as e:
                print(f"❌ Installation failed: {e}")
                sys.exit(1)
        else:
            print(f"❌ Cannot continue without {package}")
            sys.exit(1)

print("=== System Check ===")

def check_python_version():
    version = sys.version_info
    if version.major < 3 or (version.major == 3 and version.minor < 7):
        print(f"❌ Python {version.major}.{version.minor} detected.")
        print("This script requires Python 3.7 or higher.")
        print("Please upgrade your Python installation.")
        return False
    else:
        print(f"✓ Python {version.major}.{version.minor}.{version.micro} - Compatible")
        return True

if not check_python_version():
    sys.exit(1)

for dep in ['pandas', 'openpyxl', 'lxml', 'pyahocorasick']:
    check_and_install(dep)

parser = ET.XMLParser(remove_blank_text=True, recover=True)
print("✓ Ready to process DDR files\n")

def replace_emojis_with_plus(text):
    """Replace all emojis in text with + character"""
    import unicodedata
    
    # Count emojis before replacement (for debug mode)
    emoji_count = 0
    
    result = []
    i = 0
    while i < len(text):
        char = text[i]
        
        # Check if this character is an emoji or part of an emoji sequence
        is_emoji = False
        
        # Method 1: Check Unicode category
        category = unicodedata.category(char)
        if category in ['So', 'Cn']:  # Symbol, Other or Unassigned
            # Additional check for emoji properties
            code_point = ord(char)
            
            # Check various emoji ranges
            if (0x1F300 <= code_point <= 0x1F9FF or  # Emoticons, pictographs, etc.
                0x1F000 <= code_point <= 0x1F2FF or  # Mahjong, dominoes, cards
                0x1FA00 <= code_point <= 0x1FAFF or  # Chess symbols, new emojis
                0x2600 <= code_point <= 0x27BF or    # Misc symbols, dingbats  
                0x2300 <= code_point <= 0x23FF or    # Misc technical
                0x2B00 <= code_point <= 0x2BFF or    # Misc symbols and arrows
                0x1F1E6 <= code_point <= 0x1F1FF or  # Regional indicators (flags)
                code_point in [0x200D, 0xFE0F] or     # ZWJ and variation selector
                0x1F3FB <= code_point <= 0x1F3FF or  # Skin tone modifiers
                0xE0020 <= code_point <= 0xE007F or  # Tag characters
                0x2000 <= code_point <= 0x206F or    # General punctuation (includes some emoji)
                0x20D0 <= code_point <= 0x20FF or    # Combining marks
                0x3000 <= code_point <= 0x303F or    # CJK symbols (includes some emoji)
                0xFE00 <= code_point <= 0xFE0F or    # Variation selectors
                0x1F000 <= code_point <= 0x1FFFF or  # All SMP emoji blocks
                0xE0000 <= code_point <= 0xE01EF):   # Tags
                is_emoji = True
        
        # Method 2: Check for specific emoji characters that might be missed
        elif ord(char) > 127:  # Non-ASCII
            # Check if it's a standalone emoji that might not be in 'So' category
            code_point = ord(char)
            if (code_point >= 0x231A or  # From watch emoji onwards
                char in '©®™℗⚠♻☠☢☣⚡☎☏✉✈🎯🎪🎨🎬🎤🎧🎼🎵🎶🎹🎻🎺🎷🎸🎮🃏🎴🀄🎲🎰🎳'):
                is_emoji = True
        
        # Check for emoji sequences (multi-codepoint emojis)
        sequence_length = 1
        if i + 1 < len(text):
            next_char = text[i + 1]
            next_code = ord(next_char)
            
            # Check for variation selector or zero-width joiner
            if next_code in [0xFE0F, 0xFE0E, 0x200D]:
                is_emoji = True
                sequence_length = 2
                
                # Check for longer sequences (e.g., family emojis, profession emojis)
                j = i + 2
                while j < len(text) and j < i + 10:  # Limit sequence check
                    following_code = ord(text[j])
                    if following_code in [0xFE0F, 0xFE0E, 0x200D] or 0x1F000 <= following_code <= 0x1FFFF:
                        sequence_length = j - i + 1
                        j += 1
                    else:
                        break
            
            # Check for flag sequences (regional indicators)
            elif 0x1F1E6 <= ord(char) <= 0x1F1FF and 0x1F1E6 <= next_code <= 0x1F1FF:
                is_emoji = True
                sequence_length = 2
            
            # Check for keycap sequences (e.g., 1️⃣)
            elif char in '0123456789*#' and next_code == 0xFE0F and i + 2 < len(text) and ord(text[i + 2]) == 0x20E3:
                is_emoji = True
                sequence_length = 3
            elif char in '0123456789*#' and next_code == 0x20E3:
                is_emoji = True
                sequence_length = 2
        
        if is_emoji:
            result.append('+')
            i += sequence_length
            emoji_count += 1
        else:
            result.append(char)
            i += 1
    
    # Final pass: catch any remaining suspicious unicode that might be emoji
    final_text = ''.join(result)
    
    # Pattern for any remaining high unicode characters that are likely emojis
    remaining_emoji_pattern = re.compile(r'[\U00010000-\U0010FFFF]+', flags=re.UNICODE)
    
    # Check each match to see if it's likely an emoji
    def check_and_replace(match):
        text = match.group(0)
        # If it's in the emoji ranges, replace it
        for char in text:
            code = ord(char)
            if 0x1F000 <= code <= 0x1FFFF:
                return '+'
        return text
    
    final_text = remaining_emoji_pattern.sub(check_and_replace, final_text)
    
    if DEBUG_MODE and emoji_count > 0:
        print(f"[DEBUG] Replaced {emoji_count} emoji sequences with +")
    
    return final_text

def build_ddr_catalogs(root, raw_xml):
    """
    Build comprehensive catalogs of all DDR entities.
    This is done once and shared with all check modules.
    """
    catalogs = {
        'scripts': {},           # script_name -> {'id': id, 'path': folder_path}
        'layouts': {},           # layout_name -> {'id': id, 'path': folder_path}
        'tables': {},            # base_table_name -> {'id': id, 'fields': {field_name -> field_info}}
        'table_occurrences': {}, # occurrence_name -> base_table_name
        'custom_functions': {},  # cf_name -> {'id': id}
        'value_lists': {},       # vl_name -> {'id': id}
        'relationships': {},     # rel_name -> {'id': id, 'left_table': ..., 'right_table': ...}
        'fields_by_table': defaultdict(dict),  # table_name -> {field_name -> field_info}
        'script_paths': {},      # script_name -> folder_path
        'layout_paths': {},      # layout_name -> folder_path
        'raw_xml': raw_xml,      # Include raw XML for checks that need it
        'root': root            # Include parsed root for checks that need it
    }
    
    print("  Building table occurrences catalog...")
    # 1. Build table occurrence mapping first
    for table_node in root.findall(".//Table"):
        occurrence_name = table_node.attrib.get("name")
        base_table_name = table_node.attrib.get("baseTable")
        if occurrence_name and base_table_name:
            catalogs['table_occurrences'][occurrence_name] = base_table_name
    
    for table_occ in root.findall(".//TableOccurrence"):
        occ_name = table_occ.attrib.get("name")
        base_table = table_occ.attrib.get("baseTable")
        if occ_name and base_table:
            catalogs['table_occurrences'][occ_name] = base_table
    
    # Also check in RelationshipGraph
    for table_occurrence in root.findall(".//RelationshipGraph//TableOccurrence"):
        name = table_occurrence.attrib.get("name")
        base_table = table_occurrence.attrib.get("baseTable")
        if name and base_table:
            catalogs['table_occurrences'][name] = base_table
    
    print("  Building tables and fields catalog...")
    # 2. Build table and field catalogs
    # Method 1: BaseTable nodes
    for table_node in root.findall(".//BaseTable"):
        table_name = table_node.attrib.get("name")
        if table_name:
            table_info = {
                'id': table_node.attrib.get("id", ""),
                'fields': {}
            }
            
            for field_node in table_node.findall(".//Field"):
                field_name = field_node.attrib.get("name")
                if field_name:
                    field_info = {
                        'name': field_name,
                        'id': field_node.attrib.get("id", ""),
                        'type': field_node.attrib.get("dataType", "")
                    }
                    table_info['fields'][field_name] = field_info
                    catalogs['fields_by_table'][table_name][field_name] = field_info
            
            catalogs['tables'][table_name] = table_info
    
    # Method 2: FieldCatalog nodes
    for field_catalog in root.findall(".//FieldCatalog/Field"):
        field_name = field_catalog.attrib.get("name")
        table_name = field_catalog.attrib.get("table")
        if field_name and table_name:
            # Get the base table if it's a table occurrence
            base_table = catalogs['table_occurrences'].get(table_name, table_name)
            
            field_info = {
                'name': field_name,
                'id': field_catalog.attrib.get("id", ""),
                'type': field_catalog.attrib.get("dataType", "")
            }
            
            # Ensure table exists in catalog
            if base_table not in catalogs['tables']:
                catalogs['tables'][base_table] = {'id': '', 'fields': {}}
            
            catalogs['tables'][base_table]['fields'][field_name] = field_info
            catalogs['fields_by_table'][base_table][field_name] = field_info
    
    # Method 3: BaseTableCatalog nodes
    for base_table_catalog in root.findall(".//BaseTableCatalog"):
        for table_entry in base_table_catalog.findall(".//BaseTable"):
            table_name = table_entry.attrib.get("name")
            if table_name:
                if table_name not in catalogs['tables']:
                    catalogs['tables'][table_name] = {'id': table_entry.attrib.get("id", ""), 'fields': {}}
                
                for fc in table_entry.findall(".//FieldCatalog/Field"):
                    field_name = fc.attrib.get("name")
                    if field_name:
                        field_info = {
                            'name': field_name,
                            'id': fc.attrib.get("id", ""),
                            'type': fc.attrib.get("dataType", "")
                        }
                        catalogs['tables'][table_name]['fields'][field_name] = field_info
                        catalogs['fields_by_table'][table_name][field_name] = field_info
    
    print("  Building scripts catalog...")
    # 3. Build script catalog with paths
    for script_catalog in root.findall(".//ScriptCatalog"):
        def process_script_catalog(elem, current_path=""):
            if elem.tag == "Group":
                group_name = elem.attrib.get("name", "")
                if group_name:
                    new_path = f"{current_path} > {group_name}" if current_path else group_name
                    for child in elem:
                        process_script_catalog(child, new_path)
            elif elem.tag == "Script":
                script_id = elem.attrib.get("id", "")
                script_name = elem.attrib.get("name", "")
                if script_id and script_name:
                    catalogs['scripts'][script_name] = {
                        'id': script_id,
                        'path': current_path if current_path else "Top Level"
                    }
                    catalogs['script_paths'][script_name] = current_path if current_path else "Top Level"
            else:
                for child in elem:
                    process_script_catalog(child, current_path)
        
        for child in script_catalog:
            process_script_catalog(child)
    
    # Also add scripts found elsewhere
    for script in root.findall(".//Script"):
        script_name = script.attrib.get("name")
        script_id = script.attrib.get("id", "")
        if script_name and script_name not in catalogs['scripts']:
            catalogs['scripts'][script_name] = {
                'id': script_id,
                'path': "Unknown"
            }
    
    print("  Building layouts catalog...")
    # 4. Build layout catalog with paths
    for layout_catalog in root.findall(".//LayoutCatalog"):
        def process_layout_catalog(elem, current_path=""):
            if elem.tag == "Group":
                group_name = elem.attrib.get("name", "")
                if group_name:
                    new_path = f"{current_path} > {group_name}" if current_path else group_name
                    for child in elem:
                        process_layout_catalog(child, new_path)
            elif elem.tag == "Layout":
                layout_id = elem.attrib.get("id", "")
                layout_name = elem.attrib.get("name", "")
                if layout_id and layout_name:
                    catalogs['layouts'][layout_name] = {
                        'id': layout_id,
                        'path': current_path if current_path else "Top Level"
                    }
                    catalogs['layout_paths'][layout_name] = current_path if current_path else "Top Level"
            else:
                for child in elem:
                    process_layout_catalog(child, current_path)
        
        for child in layout_catalog:
            process_layout_catalog(child)
    
    # Also add layouts found elsewhere
    for layout in root.findall(".//Layout"):
        layout_name = layout.attrib.get("name")
        layout_id = layout.attrib.get("id", "")
        if layout_name and layout_name not in catalogs['layouts']:
            catalogs['layouts'][layout_name] = {
                'id': layout_id,
                'path': "Unknown"
            }
    
    print("  Building other catalogs...")
    # 5. Build custom function catalog
    for cf in root.findall(".//CustomFunction"):
        cf_name = cf.attrib.get("name")
        cf_id = cf.attrib.get("id", "")
        if cf_name:
            catalogs['custom_functions'][cf_name] = {'id': cf_id}
    
    # 6. Build value list catalog
    for vl in root.findall(".//ValueList"):
        vl_name = vl.attrib.get("name")
        vl_id = vl.attrib.get("id", "")
        if vl_name:
            catalogs['value_lists'][vl_name] = {'id': vl_id}
    
    # 7. Build relationship catalog
    for rel in root.findall(".//Relationship"):
        rel_name = rel.attrib.get("name")
        rel_id = rel.attrib.get("id", "")
        if rel_name:
            # Try to get left and right tables
            left_table = None
            right_table = None
            
            left_table_elem = rel.find(".//LeftTable")
            if left_table_elem is not None:
                left_table = left_table_elem.attrib.get("name")
            
            right_table_elem = rel.find(".//RightTable")
            if right_table_elem is not None:
                right_table = right_table_elem.attrib.get("name")
            
            catalogs['relationships'][rel_name] = {
                'id': rel_id,
                'left_table': left_table,
                'right_table': right_table
            }
    
    print("  Catalog building complete!")
    return catalogs

def load_check_module(module_name, file_path):
    """Dynamically load a check module from file path"""
    try:
        spec = importlib.util.spec_from_file_location(module_name, file_path)
        module = importlib.util.module_from_spec(spec)
        
        # Inject the color helpers into the module before executing
        module.STANDARD_COLORS = STANDARD_COLORS
        module.get_color_fill = get_color_fill
        module.get_color_font = get_color_font
        
        spec.loader.exec_module(module)
        return module
    except Exception as e:
        print(f"Error loading check module {module_name}: {e}")
        return None

def parse_ddr(xml_file, base_output_name):
    try:
        # Check if we can use cached data
        cached_xml, cached_catalogs = load_from_cache(xml_file)
        
        if cached_xml and cached_catalogs:
            # We have cached data - use it directly!
            print("\nUsing cached normalized XML and catalogs...")
            raw_xml = cached_xml
            
            # Parse the root from cached XML
            parser = ET.XMLParser(remove_blank_text=True, recover=True)
            root = ET.fromstring(raw_xml.encode('utf-8'), parser)
            
            # Use cached catalogs and add raw_xml and root back
            catalogs = cached_catalogs
            catalogs['raw_xml'] = raw_xml
            catalogs['root'] = root
        else:
            # No cache or cache mode not enabled, process normally
            print("\nProcessing DDR XML...")
            
            # Read and normalize XML
            raw_xml, root = read_and_normalize_xml(xml_file)
            
            # Build comprehensive catalogs once
            print("\nBuilding DDR catalogs...")
            catalogs = build_ddr_catalogs(root, raw_xml)
            
            # Save to cache if cache mode is enabled
            save_to_cache(xml_file, raw_xml, catalogs)
        
        # Print catalog summary
        print(f"\n✓ Found {len(catalogs['scripts'])} scripts")
        print(f"✓ Found {len(catalogs['layouts'])} layouts")
        print(f"✓ Found {len(catalogs['tables'])} tables")
        print(f"✓ Found {len(catalogs['table_occurrences'])} table occurrences")
        print(f"✓ Found {len(catalogs['custom_functions'])} custom functions")
        print(f"✓ Found {len(catalogs['value_lists'])} value lists")
        print(f"✓ Found {len(catalogs['relationships'])} relationships")
        
        total_fields = sum(len(table_info.get('fields', {})) for table_info in catalogs['tables'].values())
        print(f"✓ Found {total_fields} fields across all tables")

        # Load and run all checks from Checks folder
        all_sheets = {}
        sheet_orders = {}  # Track sheet orders
        checks_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Checks")
        
        if os.path.exists(checks_dir):
            print("\nRunning checks...")
            # First, collect all check modules
            check_modules = []
            
            # Scan for all .py files in Checks folder
            for filename in sorted(os.listdir(checks_dir)):
                if filename.endswith('.py') and not filename.startswith('__'):
                    check_path = os.path.join(checks_dir, filename)
                    module_name = filename[:-3]  # Remove .py extension
                    
                    check_module = load_check_module(module_name, check_path)
                    if check_module and hasattr(check_module, 'run_check'):
                        # Get sheet order if defined
                        sheet_order = None
                        if hasattr(check_module, 'get_sheet_order'):
                            try:
                                sheet_order = check_module.get_sheet_order()
                            except:
                                sheet_order = None
                        
                        check_modules.append({
                            'module': check_module,
                            'name': module_name,
                            'path': check_path,
                            'order': float(sheet_order) if sheet_order is not None else float('inf')
                        })
            
            # Sort modules by order, then by name
            check_modules.sort(key=lambda x: (x['order'], x['name']))
            
            # Run checks in sorted order
            total_results = 0
            for check_info in check_modules:
                check_module = check_info['module']
                module_name = check_info['name']
                
                print(f"Running {module_name}...")
                
                try:
                    # Check if the module expects the new catalogs parameter
                    import inspect
                    sig = inspect.signature(check_module.run_check)
                    
                    if len(sig.parameters) >= 2:
                        # New style - pass catalogs
                        check_results = check_module.run_check(raw_xml, catalogs)
                    else:
                        # Old style - just pass raw_xml for backward compatibility
                        check_results = check_module.run_check(raw_xml)
                    
                    if check_results and hasattr(check_module, 'get_sheet_name'):
                        sheet_name = check_module.get_sheet_name()
                    else:
                        # Default sheet name based on module name
                        sheet_name = module_name.replace('_', ' ').replace('Check', '').strip()
                    
                    if check_results:
                        all_sheets[sheet_name] = {
                            'data': check_results,
                            'module': check_module
                        }
                        sheet_orders[sheet_name] = check_info['order']
                        total_results += len(check_results)
                        print(f"✓ {module_name} completed with {len(check_results)} results")
                    else:
                        print(f"  {module_name} returned no results")
                except Exception as e:
                    print(f"Error running {module_name}: {e}")
                    import traceback
                    traceback.print_exc()
            
            # Clear terminal before final output if not in debug mode
            if DEBUG_MODE:
                print("\n[DEBUG] Skipping terminal clear due to debug mode")
                print("=== FileMaker DDR Analysis Complete (Debug Mode) ===")
                print(f"Total issues found: {total_results}")
                print("Debug mode: All output preserved above")
                print()
            else:
                os.system('cls' if os.name == 'nt' else 'clear')
                print("=== FileMaker DDR Analysis Complete ===")
                print(f"Total issues found: {total_results}")
                if CACHE_MODE:
                    print("Cache mode: Enabled")
                print()
            
            # Generate output with all sheets
            if all_sheets:
                generate_output_files(all_sheets, base_output_name, total_results, sheet_orders)
            else:
                print("\nNo results found from any checks.")
        else:
            print(f"\nChecks folder not found at: {checks_dir}")
            print("Please create a 'Checks' folder and add check modules.")

    except ET.ParseError as e:
        print(f"Error parsing XML: {e}")
    except FileNotFoundError:
        print(f"File not found: {xml_file}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def read_and_normalize_xml(xml_file):
    """Read and normalize the XML file, return raw_xml and parsed root"""
    # --- Read & normalize XML with better encoding handling ---
    encodings_to_try = ['utf-8', 'utf-8-sig', 'utf-16', 'utf-16le', 'utf-16be']
    # Detect BOM-based UTF-16 and read accordingly
    with open(xml_file, 'rb') as __f:
        __bom = __f.read(2)
    if __bom in (b'\xff\xfe', b'\xfe\xff'):
        # File is UTF-16, read in one shot and skip further encoding tries
        with open(xml_file, 'r', encoding='utf-16', errors='replace') as __f:
            raw_xml = __f.read()
        print("Successfully read file with utf-16 BOM detection")
    else:
        raw_xml = None

    if raw_xml is None:
        for encoding in encodings_to_try:
            try:
                with open(xml_file, 'r', encoding=encoding, errors='replace') as f:
                    raw_xml = f.read()
                print(f"Successfully read file with {encoding} encoding")
                break
            except UnicodeDecodeError:
                continue
    
    if raw_xml is None:
        print(f"❌ Failed to read '{xml_file}' with any of the tried encodings.")
        sys.exit(1)

    # Replace all emojis with + BEFORE any other processing
    raw_xml = replace_emojis_with_plus(raw_xml)
    
    count_nwea  = raw_xml.count("??")
    print(f'Found {count_nwea} occurrences of "??" in raw_xml')

    # Replace all double question marks
    raw_xml = raw_xml.replace("??", "+")
    
    # Replace smart quotes with regular quotes to normalize script references
    # Using Unicode escapes to avoid encoding issues
    raw_xml = raw_xml.replace("\u201C", '"')  # U+201C " Left double quotation mark
    raw_xml = raw_xml.replace("\u201D", '"')  # U+201D " Right double quotation mark
    raw_xml = raw_xml.replace("\u201E", '"')  # U+201E „ Double low-9 quotation mark
    raw_xml = raw_xml.replace("\u201F", '"')  # U+201F ‟ Double high-reversed-9 quotation mark
    raw_xml = raw_xml.replace("\u2018", "'")  # U+2018 ' Left single quotation mark
    raw_xml = raw_xml.replace("\u2019", "'")  # U+2019 ' Right single quotation mark
    
    # Fix self-closing tags to ensure proper parsing
    # Add space before /> to prevent issues with script name detection
    raw_xml = raw_xml.replace('"/>', '" />')
    raw_xml = raw_xml.replace("'/>", "' />")
    
    # Also add space after semicolons in quotes for better parsing
    raw_xml = raw_xml.replace('";', '" ;')
    raw_xml = raw_xml.replace("';", "' ;")
    
    count_plus = raw_xml.count("+")
    print(f"raw contains '+': {count_plus}")
    
    # Parse the XML once
    parser = ET.XMLParser(remove_blank_text=True, recover=True)
    root = ET.fromstring(raw_xml.encode('utf-8'), parser)
    
    return raw_xml, root

def generate_output_files(all_sheets, base_output_name, total_found, sheet_orders=None):
    """Write results to Excel with formatting."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Output path in Exports folder
    exports_dir = get_exports_dir()
    output_path = os.path.join(exports_dir, f"{os.path.basename(base_output_name)}_full_report.xlsx")

    # Fix duplicate sheet orders by adjusting them
    if sheet_orders:
        # Find duplicate order values and resolve them
        order_values = list(sheet_orders.values())
        seen_orders = set()
        duplicate_orders = [order for order in order_values if order in seen_orders or seen_orders.add(order)]

        if duplicate_orders and DEBUG_MODE:
            print(f"[DEBUG] Found {len(duplicate_orders)} duplicate sheet order values: {duplicate_orders}")
        
        # Fix duplicate orders by adding increments of 10
        if duplicate_orders:
            # Keep track of orders we've seen and adjusted
            fixed_orders = {}
            increment = 10
            
            for sheet_name, order in sheet_orders.items():
                if order in fixed_orders:
                    # This order has already been seen and adjusted for another sheet
                    # Use the next increment (10, 20, 30, etc.)
                    new_order = order + increment
                    sheet_orders[sheet_name] = new_order
                    increment += 10
                    
                    if DEBUG_MODE:
                        print(f"[DEBUG] Adjusted order for '{sheet_name}' from {order} to {new_order}")
                else:
                    # First time seeing this order, record it
                    fixed_orders[order] = True
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sort sheets by order, then by name
        if sheet_orders:
            sorted_sheets = sorted(all_sheets.items(), 
                                 key=lambda x: (sheet_orders.get(x[0], float('inf')), x[0]))
        else:
            sorted_sheets = sorted(all_sheets.items())
        
        # Verify that each sheet is unique (debug only)
        if DEBUG_MODE:
            sheet_names = [name for name, _ in sorted_sheets]
            if len(sheet_names) != len(set(sheet_names)):
                print(f"[DEBUG] WARNING: Duplicate sheet names found: {sheet_names}")
            print(f"[DEBUG] Writing {len(sorted_sheets)} sheets in order:")
            for idx, (name, _) in enumerate(sorted_sheets):
                order = sheet_orders.get(name, float('inf'))
                print(f"[DEBUG]   {idx+1}. '{name}' (order: {order})")
        
        for sheet_name, sheet_info in sorted_sheets:
            sheet_data = sheet_info['data']
            if isinstance(sheet_data, pd.DataFrame):
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Convert list of dicts to DataFrame
                pd.DataFrame(sheet_data).to_excel(writer, sheet_name=sheet_name, index=False)

    # Styling
    wb = load_workbook(output_path)
    
    # Verify all sheets are present (debug only)
    if DEBUG_MODE:
        expected_sheets = [name for name, _ in sorted_sheets]
        actual_sheets = wb.sheetnames
        print(f"[DEBUG] Expected {len(expected_sheets)} sheets, found {len(actual_sheets)} sheets")
        if set(expected_sheets) != set(actual_sheets):
            missing = set(expected_sheets) - set(actual_sheets)
            extra = set(actual_sheets) - set(expected_sheets)
            if missing:
                print(f"[DEBUG] Missing sheets: {missing}")
            if extra:
                print(f"[DEBUG] Unexpected sheets: {extra}")
    
    header_fill = get_color_fill('header')
    header_font = get_color_font('header', bold=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = ws['A2']
        
        # Row height
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30
            
        # Wrap & alignment
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
        # Column widths - get from check module if available
        column_widths = {}
        check_module = None
        
        # Find the check module for this sheet
        if sheet_name in all_sheets:
            check_module = all_sheets[sheet_name]['module']
            if hasattr(check_module, 'get_column_widths'):
                column_widths = check_module.get_column_widths()
        
        # Apply column widths
        for idx, cell in enumerate(ws[1], 1):
            val = cell.value
            col = get_column_letter(idx)
            
            if val in column_widths:
                ws.column_dimensions[col].width = column_widths[val]/7
            elif val == "SQL Text":
                ws.column_dimensions[col].width = 40
            elif val == "":
                ws.column_dimensions[col].width = 3
            else:
                ws.column_dimensions[col].width = 20
                
        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Apply sheet-specific styling from check module
        if check_module and hasattr(check_module, 'apply_styling'):
            check_module.apply_styling(ws)

    wb.save(output_path)

    # Auto-open the file after saving
    auto_open_file(output_path)

    print(f"Full report saved to: {os.path.abspath(output_path)}")
    print(f"\nRun complete: {total_found} results found.")

def auto_open_file(filename):
    try:
        if DEBUG_MODE:
            print(f"[DEBUG] Opening file: {filename}")
        
        if platform.system() == "Darwin":
            subprocess.run(["open", filename])
        elif platform.system() == "Windows":
            os.startfile(filename)
        elif platform.system() == "Linux":
            subprocess.run(["xdg-open", filename])
    except Exception as e:
        if DEBUG_MODE:
            print(f"[DEBUG] Error opening file: {e}")
        pass

def get_input_file():
    """Handle file selection"""
    # Check if we have a cached file to reuse
    if CACHE_MODE:
        last_file = get_last_used_file()
        if last_file:
            print(f"\nLast used file: {os.path.basename(last_file)}")
            response = input("Use this file again? (Y/n): ").lower().strip()
            if response in ['', 'y', 'yes']:
                return last_file
    
    # Create Tk root and bring it to front
    root = Tk()
    root.withdraw()
    
    # Bring the dialog to front on all platforms
    root.lift()
    root.attributes('-topmost', True)
    root.focus_force()
    
    # Show the file dialog
    input_file = askopenfilename(
        title="Select DDR XML File",
        filetypes=[("XML Files", "*.xml")],
        parent=root
    )
    
    # Immediately destroy the root window to prevent "tk" window from lingering
    root.quit()
    root.destroy()
    
    if not input_file:
        print("No file selected. Exiting.")
        sys.exit(0)
    
    # Save this as the last used file if in cache mode
    if CACHE_MODE:
        save_last_used_file(input_file)
    
    return input_file

if __name__ == "__main__":
    # Get input file
    input_file = get_input_file()
    
    print(f"\nProcessing: {os.path.basename(input_file)}")
    if CACHE_MODE:
        print("Cache mode: Enabled - will use cached catalogs if available")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = os.path.splitext(os.path.basename(input_file))[0]
    base_name = os.path.join(script_dir, f"{input_filename}_parsed")

    print("Starting enhanced DDR parse...")
    parse_ddr(input_file, base_name)
